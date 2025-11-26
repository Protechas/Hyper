import os
import re
import sys
import time
import shutil
import winreg
import platform
import openpyxl
import subprocess
import urllib.parse
import tkinter as tk
from enum import Enum
import win32clipboard
from tkinter import messagebox
from selenium import webdriver
import chromedriver_autoinstaller
from openpyxl.styles import Font
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.webdriver import WebDriver
from selenium.webdriver.common.window import WindowTypes
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.remote.webelement import WebElement
import urllib.parse
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC

#####################################################################################################################################################

import re
from difflib import SequenceMatcher

def _strip_qualifiers(s: str) -> str:
    s = (s or '')
    s = re.sub(r'\[[^\]]*\]', '', s)   # remove [HEV], etc.
    s = re.sub(r'\([^)]+\)', '', s)    # remove (Hybrid), etc.
    s = s.replace('-', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s.strip().upper()

# ★ Replace your helper with this version (covers 5500/5500HD and 6500/6500HD) adds to bottom of sheet, unless there is a exact match, its overwriting things that shouldent be

# ── add near your helpers ─────────────────────────────────────────────────────
def _adas_name_norms(txt: str):
    t = (txt or "").upper()
    letters_only = re.sub(r"[^A-Z]", "", t)      # "ACC (4)" -> "ACC"
    alnum       = re.sub(r"[^A-Z0-9]", "", t)    # "ACC (4)" -> "ACC4"
    return letters_only, alnum

def _adas_name_col_index(repair_mode: bool, excel_mode: str):
    # ADAS sheets (OG & NEW): "SME Generic System Name" is Column S (0-based 18) in your screenshots.
    # Return None in Repair mode.
    if repair_mode:
        return None
    return 18

def __add_yellow_text_marker(self, worksheet, year, make, model, system, file_name):
    """
    Apply a dark yellow font color for NO-doc hyperlinks (file names starting with 'No ...').
    """
    from openpyxl.styles import Font, Color
    for row_key, row_num in self.row_index.items():
        yr, mk, mdl, sys = row_key
        if str(yr) == str(year) and mk == make and mdl == model and sys == system:
            hyperlink_col = self.HYPERLINK_COLUMN_INDEX or 12
            cell = worksheet.cell(row=row_num, column=hyperlink_col)
            if cell and cell.hyperlink:
                cell.font = Font(color="9B870C", underline="single")  # dark yellow
            elif cell:
                cell.font = Font(color="9B870C")
            print(f"🟡 Marked NO-doc hyperlink in yellow at {cell.coordinate} → {file_name}")
            break    


def _is_force_bottom_model(model: str) -> bool:
    """
    Models that MUST NOT regex/fuzzy-match into other trims.
    Covers:
      - Chevrolet: Silverado 4500 / 4500HD / 5500 / 5500HD / 6500 / 6500HD
      - Ford: F-450 / F450, F-550 / F550, E-450 / E450
    """
    m = (model or "").upper()
    return bool(re.search(
        r"\b(?:SILVERADO\s*(?:4500|5500|6500)(?:\s*HD)?|F[-\s]?450|F[-\s]?550|E[-\s]?450)\b",
        m
    ))

def _is_force_bottom_combo(year, make, model) -> bool:
    """
    Return True for combos that must NOT regex/fuzzy-match.
    Force-bottom unless there is an exact (Year+Make+Model+System) match.
    """
    y = (str(year) or "").strip()
    m = (make or "").upper()
    mdl_core = re.sub(r"\s+", " ", _strip_qualifiers(model)).upper()
    mdl_core_nopunct = re.sub(r"[^A-Z0-9 ]", "", mdl_core)

    # Existing rules
    if y == "2025" and m == "LAND ROVER" and mdl_core.startswith("RANGE ROVER EVOQUE"):
        return True
    if y == "2023" and m == "LEXUS" and mdl_core.startswith("RX450H"):
        return True
    if y == "2017" and m == "LEXUS" and mdl_core.startswith("NX300"):
        return True
    if y in {"2023", "2024"} and m == "CADILLAC" and mdl_core.startswith("ESCALADE ESV"):
        return True
    if y in {"2023", "2024"} and m == "AUDI" and mdl_core.startswith("SQ3"):
        return True
    if y == "2015" and m == "FIAT" and mdl_core.startswith("500X"):
        return True
    if y == "2025" and m == "GMC" and mdl_core.startswith("HUMMER EV"):
        return True
    if y == "2017" and m == "HONDA" and mdl_core.startswith("CLARITY"):
        return True
    if y == "2021" and m == "JAGUAR" and mdl_core.startswith("I PACE"):
        return True

    # Audi e-tron family
    if y == "2025" and m == "AUDI" and (mdl_core.startswith("Q8 E TRON") or mdl_core.startswith("Q8 E-TRON")):
        return True
    if y in {"2023", "2024"} and m == "AUDI" and (mdl_core.startswith("RS E TRON") or mdl_core.startswith("RS E-TRON")):
        return True

    # Porsche Cayenne Coupe (2024)
    if y == "2024" and m in {"PORSCHE", "PORCSHE"} and (
        mdl_core.startswith("CAYENNE COUPE 9YA") or
        mdl_core.startswith("CAYENNE COUPE 9YB")
    ):
        return True

    # Porsche Cayenne (2023, 2022)
    if y == "2023" and m == "PORSCHE" and (
        mdl_core.startswith("CAYENNE 9YA") or
        mdl_core.startswith("CAYENNE 9YB")
    ):
        return True
    if y == "2022" and m == "PORSCHE" and (
        mdl_core.startswith("CAYENNE 9YA") or
        mdl_core.startswith("CAYENNE 9YB")
    ):
        return True

    # ★ NEW: Porsche Cayenne (2024) — non-coupe 9YA
    if y == "2024" and m == "PORSCHE" and mdl_core.startswith("CAYENNE 9YA"):
        return True

    # Volkswagen ID.7
    if y == "2025" and m == "VOLKSWAGEN" and (
        mdl_core_nopunct.startswith("ID7") or mdl_core.startswith("ID 7")
    ):
        return True

    return False




# ★ NEW: model-family guards to prevent NX300h ↔ ES300 hops, etc.
def _alpha_prefix(s: str) -> str:
    t = re.sub(r'[^A-Z0-9]', '', (s or '').upper())
    m = re.match(r'([A-Z]+)', t)
    return m.group(1) if m else ''

def _model_number_block(s: str) -> str:
    t = re.sub(r'[^A-Z0-9]', '', (s or '').upper())
    m = re.search(r'(\d{2,4})', t)
    return m.group(1) if m else ''

def _cross_family_conflict(a: str, b: str) -> bool:
    """
    True when models share the same number block (e.g., 300) but have different alpha prefixes (NX vs ES).
    Blocks regex/fuzzy/letters-only placements across families; exact matches still allowed.
    """
    ap, bp = _alpha_prefix(a), _alpha_prefix(b)
    an, bn = _model_number_block(a), _model_number_block(b)
    return bool(an and bn and an == bn and ap and bp and ap != bp)


# ★ Add once near your helpers
def _system_missing_text(txt: str) -> bool:
    """
    True when the sheet's System cell isn't a usable ADAS acronym.
    Covers: blank, 'Sys N/A', 'Sys N/A - Place', 'Mapping', 'Mapping Needed', etc.
    """
    t = (txt or "").strip().upper()
    if not t:
        return True
    # common variants
    if "SYS" in t and "N/A" in t:
        return True
    if "MAPPING" in t:
        return True
    if "PLACE" in t and "SYS" in t:
        return True
    return False


def _model_regex_from_excel(model_text: str):
    core = _strip_qualifiers(model_text)
    core = re.sub(r'\s+', r'\\s*', re.escape(core))  # allow optional spaces
    return re.compile(rf'(^|[^A-Z0-9]){core}([^A-Z0-9]|$)', re.IGNORECASE)

def _similar(a: str, b: str) -> float:
    # compare on alphanum-only uppercase
    norm = lambda x: re.sub(r'[^A-Z0-9]', '', (x or '').upper())
    return SequenceMatcher(None, norm(a), norm(b)).ratio()

def _norm_system_index(s: str) -> str:
    # EXACTLY how your index normalizes: keep digits, strip non-alphanum
    return re.sub(r'[^A-Z0-9]', '', (s or '').upper())

def _norm_system_loose(s: str) -> str:
    # fallback: letters only (rare sheets that drop digits)
    return re.sub(r'[^A-Z]', '', (s or '').upper())

def _extract_system_from_filename(file_name: str) -> str:
    """
    Extract the ADAS system acronym from a SharePoint filename.
    Priority:
      1) Parentheses (…) content that is a known ADAS acronym (STRICT PREFERENCE)
      2) Any known acronym anywhere in the string
      3) Brackets […] content (only if no parentheses match) and only if known
      4) Last token (only if known)
    This avoids treating model qualifiers like [EV]/[HEV] as systems.
    """
    name = (file_name or "").upper()
    if re.search(r'(?i)\bg[\s\-]?force\b', name):
        return "G-FORCE"
    # Known acronyms — expandable, includes 1/2 suffix forms you’ve used
    KNOWN = {
        # ADAS (existing)
        "ACC","ACC1","ACC2","ACC3",
        "AEB","AEB1","AEB2","AEB3",
        "AHL",
        "APA","APA1","APA2","APA3",
        "BSW","BSW1","BSW2","BSW3",
        "BUC",
        "LKA","LKA1","LKA2","LKA3",
        "LW",
        "NV",
        "SVC","SVC1","SVC2","SVC3",
        "WAMC",

        # Repair SI (added)
        "YAW",
        "G-Force","GFORCE",         # both dashed and plain
        "SWS",
        "HUD",
        "SRS D&E","SRSDE",          # keep spaced/ampersand + normalized
        "SCI",
        "SRR",
        "TPMS",
        "SBI",
        "EBDE",                     # catches "EBDE (1)/(2)"
        "HDE",                      # catches "HDE (1)/(2)"
        "LGR",
        "PSI",
        "WRL",
        "PCM",
        "TRANS",
        "AIR",
        "ABS",
        "BCM",
        "SAS",
        "HLI",
        "ESC",
        "SRS",
        "KEY",
        "FOB",
        "HVAC",                     # catches "HVAC (1)/(2)"
        "COOL",
        "HEAD",                     # catches "HEAD (1)/(2)"
        "OCS","OCS1","OCS2","OCS3","OCS4",
    }


    # --- 1) Prefer (...) tokens ---
    paren_tokens = re.findall(r"\(([^\)]+)\)", name)
    for tok in paren_tokens:
        t = tok.strip().upper()
        # Normalize light punctuation/spacing for matching
        t_plain = re.sub(r"[^A-Z0-9\-]", "", t)  # keep dash to catch BSM-RCTW
        if t in KNOWN or t_plain in KNOWN:
            return tok.strip().upper()  # return raw token; caller normalizes for the index

    # --- 2) Scan entire string for any known acronym ---
    for k in sorted(KNOWN, key=len, reverse=True):  # longer first (e.g., BSM-RCTW)
        if k in name:
            return k

    # --- 3) Consider [...] tokens only if NOTHING found yet, and only if known ---
    bracket_tokens = re.findall(r"\[([^\]]+)\]", name)
    for tok in bracket_tokens:
        t = tok.strip().upper()
        t_plain = re.sub(r"[^A-Z0-9\-]", "", t)
        if t in KNOWN or t_plain in KNOWN:
            return tok.strip().upper()

    # --- 4) Last token, but only if it’s known ---
    parts = name.split()
    if parts:
        last = parts[-1].strip().upper()
        last_plain = re.sub(r"[^A-Z0-9\-]", "", last)
        if last in KNOWN or last_plain in KNOWN:
            return last

    # No system detected
    return ""



def _system_val_for_row(self, row, repair_mode: bool):
    """
    Return (system_text, system_norm_for_index) for a given openpyxl 'row' (tuple of cells).
    Uses the correct columns for OG vs NEW and Repair vs ADAS.
    NOTE: 'row[i]' here is 0-based indexing (row[0] == Column A).
    """
    # Decide which source column to read the Protech Generic system name from
    if repair_mode:
        # Repair SI
        if self.excel_mode == "new":
            # NEW Repair: Column T (0-based 19)
            sys_cell = row[19] if len(row) > 19 and row[19].value else None
        elif str(self.sharepoint_make).lower() == "toyota":
            # OG Repair for Toyota uses Column E (0-based 4)
            sys_cell = row[4] if len(row) > 4 and row[4].value else None
        else:
            # OG Repair default uses Column D (0-based 3)
            sys_cell = row[3] if len(row) > 3 and row[3].value else None
    else:
        # ADAS SI
        if self.excel_mode == "new":
            # 🔧 NEW ADAS: Column U (0-based 20) after the S→U move
            sys_cell = row[20] if len(row) > 20 and row[20].value else None
        else:
            # OG ADAS: Column E (0-based 4)
            sys_cell = row[18] if len(row) > 18 and row[18].value else None

    sys_text = (str(sys_cell.value).strip().upper() if sys_cell else "")
    sys_norm = re.sub(r"[^A-Z0-9]", "", sys_text)  # EXACTLY like your __build_row_index__
    return sys_text, sys_norm



class SharepointExtractor: 
    """
    Class definition for a SharepointExtractor. 
    Navigates to a given sharepoint location and extracts all the links for files and folders inside of the given location
    """
  
    # Attributes for the SharepointExtractor instance
    sharepoint_make: str = None             # The make of the sharepoint folder in use
    sharepoint_link: str = None             # The link to the sharepoint location for the given make
    excel_file_path: str = None             # The fully qualified path to the excel file holding our ADAS SI
    selenium_driver: WebDriver = None       # The selenium driver instance for this session
    selenium_wait: WebDriverWait = None     # The default wait operator for the webdriver
    
    # Configuration attributes for the sharepoint module names and timeouts
    __MAX_WAIT_TIME__ = 120
    __DEBUG_RUN__ = False

    # Locators used to find objects on the sharepoint folder pages
    __ONEDRIVE_PAGE_NAME_LOCATOR__ = "//li[@data-automationid='breadcrumb-listitem']//div[@data-automationid='breadcrumb-crumb']"
    __ONEDRIVE_TABLE_LOCATOR__ = "//div[@data-automationid='list-page']"
    __ONEDRIVE_TABLE_ROW_LOCATOR__ = "./div[@role='row' and starts-with(@data-automationid, 'row-')]"

    # Collections of system names used for finding correct files and row locations
    __DEFINED_MODULE_NAMES__ = [
        'ACC', 'SCC', 'AEB', 'AHL', 'APA', 'BSW', 'BSW/RCTW', 'BSW-RCTW',
        'BSW & RCTW', 'BSW RCTW', 'BSW-RCT W', 'BSW RCT W', 'BSM-RCTW', 'BSW-RTCW', 'BSW_RCTW',
        'BCW-RCTW', 'BUC', 'LKA', 'LW', 'NV', 'SVC', 'WAMC', 'FRS', 'PDS', 'RRS', 'WSC', 
    
        # 🔧 Repair SI modules added below
        'YAW', 'G-Force', 'SWS', 'HUD', 'SRS D&E', 'SCI', 'SRR', 'TPMS', 'SBI',
        'EBDE (1)', 'EBDE (2)', 'HDE (1)', 'HDE (2)', 'LGR', 'PSI', 'WRL',
        'PCM', 'TRANS', 'AIR', 'ABS', 'BCM', 'SAS', 'HLI', 'ESC','SRS',
        'KEY', 'FOB', 'HVAC (1)', 'HVAC (2)', 'COOL', 'HEAD (1)', 'HEAD (2)',
        "OCS","OCS2","OCS3","OCS4"
    ]

    __ROW_SEARCH_TERMS__ = ['LKAS', 'FCW/LDW', 'Multipurpose', 'Cross Traffic Alert', 'Side Blind Zone Alert', 'Lane Change Alert', 'Blind Spot Warning (BSW)', 'Surround Vision Camera', 'Video Processing', 'Pending Further Research',]
    __ADAS_SYSTEMS_WHITELIST__ = [
        'FCW/LDW',
        'FCW-LDW',
        'Multipurpose Camera',
        'Multipurpose',
        'WL',
        '[WL]',
        'Forward Collision Warning/Lane Departure Warning (FCW/LDW)',
        'Blind Spot Warning (BSW)',
        'Cross Traffic Alert',
        ' Side Blind Zone Alert',
        'Side Blind Zone Alert',
        'Surround Vision Camera',
        'Video Processing'
    ]
    SPECIFIC_HYPERLINKS = {
    # ADAS SI Manual Placeholders
    #"2013 GMC Acadia (BSW-RCTW 1)-PL-PW072NLB.pdf": "L79",
    #"2016 Acura RLX (LKA 1) [FCW-LDW].pdf": "L248",
    #"2016 Acura RLX (LKA 1) [Multipurpose].pdf": "L249",
    #"2012 Volkswagen CC (ACC 1).pdf": "L11",
    #"2013 Volkswagen CC (ACC 1).pdf": "L83",
    #"2014 Volkswagen CC (ACC 1).pdf": "L155",
    #"2015 Volkswagen CC (ACC 1).pdf": "L227",
    #"2016 Volkswagen CC (ACC 1).pdf": "L299",
    #"2017 Volkswagen CC (ACC 1).pdf": "L371",
    #"2022 Kis Stinger (ACC 2)": "L956",
    #"2023 Kia Niro EV (ACC 2)": "L1010",
    #"2023 Kia Niro EV (AEB 2)": "L1011",
    #"2023 Kia Niro EV (APA 1).pdf": "L1013",
    #"2023 Kia Niro EV (BSW 1).pdf": "L1014",
    #"2023 Kia Niro EV (BUC).pdf": "L1015",
    #"2023 Kia Niro EV (LKA 1).pdf": "L1016",
    #"2023 Kia Niro HEV (ACC 2)": "L1019",
    #"2023 Kia Niro HEV (AEB 2)": "L1020",
    #"2023 Kia Niro HEV (ACC 2).pdf": "L155", Dupliacate entry
    #"2023 Kia Niro HEV (APA 1).pdf": "L1022",
    #"2023 Kia Niro HEV (BSW 1).pdf": "L1023",
    #"2023 Kia Niro HEV (BUC).pdf": "L1024",
    #"2023 Kia Niro HEV (LKA 1).pdf": "L1025",
    #"2023 Kia Niro PHEV (ACC 2)": "L1028",
    #"2023 Kia Niro PHEV (AEB 2)": "L1029",
    #"2023 Kia Niro PHEV (APA 1).pdf": "L1031",
    #"2023 Kia Niro PHEV (BSW 1).pdf": "L1032",
    #"2023 Kia Niro PHEV (BUC).pdf": "L1033",
    #"2023 Kia Niro PHEV (LKA 1).pdf": "L1034",
    #"2023 Kia Sorento HEV (ACC 2)": "L1064",
    #"2023 Kia Sorento HEV (AEB 2)": "L1065",
    #"2023 Kia Sorento HEV (BSW 1)": "L1068",
    #"2023 Kia Sorento HEV (SVC 1)": "L1072",
    #"2023 Kia Sorento HEV (APA 1).pdf": "L1067",
    #"2023 Kia Sorento HEV (BUC).pdf": "L1069",
    #"2023 Kia Sorento HEV (LKA 1).pdf": "L1070",
    #"2023 Kia Sorento PHEV (BSW 1)": "L1077",
    #"2023 Kia Sorento PHEV (SVC 1)": "L1081",
    #"2023 Kia Sorento PHEV (ACC 2).pdf": "L1073",
    #"2023 Kia Sorento PHEV (AEB 2).pdf": "L1074",
    #"2023 Kia Sorento PHEV (APA 1).pdf": "L1076",
    #"2023 Kia Sorento PHEV (BUC).pdf": "L1078",
    #"2023 Kia Sorento PHEV (LKA 1).pdf": "L1079",
    #"2023 Kia Sportage (ACC 2)": "L1091",
    #"2023 Kia Sportage (AEB 2)": "L1092",
    #"2023 Kia Sportage (APA 1)": "L1094",
    #"2023 Kia Sportage (SVC 1)": "L1099",
    #"2023 Kia Sportage (BSW 1).pdf": "L1095",
    #"2023 Kia Sportage (BUC).pdf": "L1096",
    #"2023 Kia Sportage (LKA 1).pdf": "L1097",
    #"2023 Kia Sportage NG5 (AEB 2)": "L1101", #misspealt, Its NQ5, not NG5
    #"2023 Kia Sportage NQ5 (ACC 2)": "L1100",
    #"2023 Kia Sportage NQ5 (APA 1)": "L1103",
    #"2023 Kia Sportage NQ5 (BSW 1).pdf": "L1104",
    #"2023 Kia Sportage NQ5 (BUC).pdf": "L1105",
    #"2023 Kia Sportage NQ5 (LKA 1).pdf": "L1106",
    #"2023 Kia Sportage NQ5 (SVC 1).pdf": "L1108",
    #"2023 Kia Sportage PHEV (ACC 2)": "L1118",
    #"2023 Kia Sportage PHEV (AEB 2)": "L1119",
    #"2023 Kia Sportage PHEV (SVC 1)": "L1126",
    #"2023 Kia Sportage PHEV (APA 1).pdf": "L1121",
    #"2023 Kia Sportage PHEV (BSW 1).pdf": "L1122",
    #"2023 Kia Sportage PHEV (BUC).pdf": "L123",
    #"2023 Kia Sportage PHEV (LKA 1).pdf": "L1124", 
    #"2024 Kia Carnival (ACC 2).pdf": "L1145", # MPV?????
    #"2024 Kia Carnival (AEB 2).pdf": "L1146",
    #"2024 Kia Carnival (APA 1).pdf": "L1148",
    #"2024 Kia Carnival (BSW 1).pdf": "L1149",
    #"2024 Kia Carnival (BUC).pdf": "L1150",
    #"2024 Kia Carnival (LKA 1).pdf": "L1151",
    #"2024 Kia Carnival (SVC 1).pdf": "L1153",
    
    #Repair SI Manual Placeholders
    "2020 Kia Niro EV (DE PHEV)(G-Force)":    "H2842",
    "2020 Kia Niro EV (DE PHEV)(SAS)":        "H2840",
    "2020 Kia Niro EV (DE PHEV)(YAW)":        "H2841",
    "2020 Kia Niro EV (DE PHEV)(SWS).pdf":    "H2843",
    
    #2021 Kia Niro (Excel) and Below dosent match, but placing links anyway (it is named completely wrong in the Excel as a normal model should never reflect the EV version)
    "2021 Kia Niro EV (DE PHEV)(G-Force)":    "H2413",
    "2021 Kia Niro EV (DE PHEV)(SAS)":        "H2411",
    "2021 Kia Niro EV (DE PHEV)(YAW)":        "H2412",
    "2021 KIA Niro EV (DE PHEV)(ESC).pdf":    "H2418",
    "2021 KIA Niro EV (DE PHEV)(HLI).pdf":    "H2422",
    "2021 KIA Niro EV (DE PHEV)(SCI).pdf":    "H2420",
    "2021 KIA Niro EV (DE PHEV)(SRR).pdf":    "H2421",
    "2021 KIA Niro EV (DE PHEV)(SRS D&E).pdf":"H2419",
    "2021 Kia Niro EV (DE PHEV)(SWS).pdf":    "H2414",
    "2021 KIA Niro EV (DE PHEV)(TPMS).pdf":   "H2423",

    #2022 Kia Niro (Excel) and Below dosent match, but placing links anyway (it is named completely wrong in the Excel as a normal model should never reflect the EV version)
    "2022 Kia Niro EV (G-Force)":             "H2050",
    "2022 Kia Niro EV (SAS)":                 "H2048",
    "2022 Kia Niro EV (SRR)":                 "H2058",
    "2022 Kia Niro EV (YAW)":                 "H2049",
    "2022 Kia Niro EV (ESC).pdf":             "H2055",
    "2022 Kia Niro EV (HLI).pdf":             "H2059",
    "2022 Kia Niro EV (SCI).pdf":             "H2057",
    "2022 Kia Niro EV (SRS D&E).pdf":         "H2056",
    "2022 Kia Niro EV (SWS).pdf":             "H2051",
    "2022 Kia Niro EV (TPMS).pdf":            "H2060",

    "2023 Kia Niro (EV)(G-Force)":            "H1423",
    "2023 Kia Niro (EV)(SAS)":                "H1421",
    "2023 Kia Niro (EV)(YAW)":                "H1422",
    "2023 Kia Niro (EV) (SWS).pdf":           "H1424",
    "2023 Kia Niro (EV)(ESC).pdf":            "H1428",
    "2023 Kia Niro (EV)(HLI).pdf":            "H1432",
    "2023 Kia Niro (EV)(SCI).pdf":            "H1430",
    "2023 Kia Niro (EV)(SRR).pdf":            "H1431",
    "2023 Kia Niro (EV)(SRS D&E).pdf":        "H1429",
    "2023 Kia Niro (EV)(TPMS).pdf":           "H1433",

    "2023 Kia Niro (HEV)(G-Force)":           "H1456",
    "2023 Kia Niro (HEV)(YAW)":               "H1455",
    "2023 Kia Niro (HEV) (SAS).pdf":          "H1454",
    "2023 Kia Niro (HEV) (SWS).pdf":          "H1457",
    "2023 Kia Niro (HEV)(ESC).pdf":           "H1461",
    "2023 Kia Niro (HEV)(HLI).pdf":           "H1465",
    "2023 Kia Niro (HEV)(SCI).pdf":           "H1463",
    "2023 Kia Niro (HEV)(SRR).pdf":           "H1464",
    "2023 Kia Niro (HEV)(SRS D&E).pdf":       "H1462",
    "2023 Kia Niro (HEV)(TPMS).pdf":          "H1466",

    "2023 Kia Niro (PHEV)(G-Force)":          "H1489",
    "2023 Kia Niro (PHEV)(YAW)":              "H1488",
    "2023 Kia Niro (PHEV) (SAS).pdf":         "H1487",
    "2023 Kia Niro (PHEV) (SWS).pdf":         "H1490",
    "2023 Kia Niro (PHEV)(ESC).pdf":          "H1494",
    "2023 Kia Niro (PHEV)(HLI).pdf":          "H1498",
    "2023 Kia Niro (PHEV)(SCI).pdf":          "H1496",
    "2023 Kia Niro (PHEV)(SRR).pdf":          "H1497",
    "2023 Kia Niro (PHEV)(SRS D&E).pdf":      "H1495",
    "2023 Kia Niro (PHEV)(TPMS).pdf":         "H1499",

    "2023 Kia Sorento (HEV)(ESC)":            "H1626",
    "2023 Kia Sorento (HEV)(G-Force)":        "H1621",
    "2023 Kia Sorento (HEV)(SAS)":            "H1619",
    "2023 Kia Sorento (HEV)(SRR)":            "H1629",
    "2023 Kia Sorento (HEV)(YAW)":            "H1620",
    "2023 Kia Sorento (HEV)(HLI).pdf":        "H1630",
    "2023 Kia Sorento (HEV)(SCI).pdf":        "H1628",
    "2023 Kia Sorento (HEV)(SRS D&E).pdf":    "H1627",
    "2023 Kia Sorento (HEV)(SWS).pdf":        "H1622",
    "2023 Kia Sorento (HEV)(TPMS).pdf":       "H1631",

    "2023 Kia Sorento (PHEV)(SAS)":           "H1652",
    "2023 Kia Sorento PHEV (ESC)":            "H1659",
    "2023 Kia Sorento PHEV (SRR)":            "H1662",
    "2023 Kia Sorento (PHEV)(G-Force).pdf":   "H1654",
    "2023 Kia Sorento (PHEV)(HLI).pdf":       "H1663",
    "2023 Kia Sorento (PHEV)(SCI).pdf":       "H1661",
    "2023 Kia Sorento (PHEV)(SRS D&E).pdf":   "H1660",
    "2023 Kia Sorento (PHEV)(SWS).pdf":       "H1655",
    "2023 Kia Sorento (PHEV)(TPMS).pdf":      "H1664",
    "2023 Kia Sorento (PHEV)(YAW).pdf":       "H1653",

    "2023 Kia Sportage (HEV)(G-Force)":       "H1720",
    "2023 Kia Sportage (HEV)(YAW)":           "H1719",
    "2023 Kia Sportage HEV (SRR)":            "H1728",
    "2023 Kia Sportage (HEV)(ESC).pdf":       "H1725",
    "2023 Kia Sportage (HEV)(HLI).pdf":       "H1729",
    "2023 Kia Sportage (HEV)(SAS).pdf":       "H1718",
    "2023 Kia Sportage (HEV)(SCI).pdf":       "H1727",
    "2023 Kia Sportage (HEV)(SRS D&E).pdf":   "H1726",
    "2023 Kia Sportage (HEV)(SWS).pdf":       "H1721",
    "2023 Kia Sportage (HEV)(TPMS).pdf":      "H1730",

    "2023 Kia Sportage (NQ5)(ESC)":           "H1758",
    "2023 Kia Sportage (NQ5)(G-Force)":       "H1753",
    "2023 Kia Sportage (NQ5)(SAS)":           "H1751",
    "2023 Kia Sportage (NQ5)(SRR)":           "H1761",
    "2023 Kia Sportage (NQ5)(YAW)":           "H1752",
    "2023 Kia Sportage (NQ5)(HLI).pdf":       "H1762",
    "2023 Kia Sportage (NQ5)(SCI).pdf":       "H1760",
    "2023 Kia Sportage (NQ5)(SRS D&E).pdf":   "H1759",
    "2023 Kia Sportage (NQ5)(SWS).pdf":       "H1754",
    "2023 Kia Sportage (NQ5)(TPMS).pdf":      "H1763",

    "2023 Kia Sportage (NQ5A)(ESC)":          "H1791",
    "2023 Kia Sportage (NQ5A)(G-Force)":      "H1786",
    "2023 Kia Sportage (NQ5A)(SRR)":          "H1794",
    "2023 Kia Sportage (NQ5A)(YAW)":          "H1785",
    "2023 Kia Sportage (NQ5A)(HLI).pdf":      "H1795",
    "2023 Kia Sportage (NQ5A)(SAS).pdf":      "H1784",
    "2023 Kia Sportage (NQ5A)(SCI).pdf":      "H1793",
    "2023 Kia Sportage (NQ5A)(SRS D&E).pdf":  "H1792",
    "2023 Kia Sportage (NQ5A)(SWS).pdf":      "H1787",
    "2023 Kia Sportage (NQ5A)(TPMS).pdf":     "H1796",

    "2023 Kia Sportage (PHEV)(SRR)":          "H1827",
    "2023 Kia Sportage (PHEV)(ESC).pdf":      "H1824",
    "2023 Kia Sportage (PHEV)(G-Force).pdf":  "H1819",
    "2023 Kia Sportage (PHEV)(HLI).pdf":      "H1828",
    "2023 Kia Sportage (PHEV)(SAS).pdf":      "H1817",
    "2023 Kia Sportage (PHEV)(SCI).pdf":      "H1826",
    "2023 Kia Sportage (PHEV)(SRS D&E).pdf":  "H1825",
    "2023 Kia Sportage (PHEV)(SWS).pdf":      "H1820",
    "2023 Kia Sportage (PHEV)(TPMS).pdf":     "H1829",
    "2023 Kia Sportage (PHEV)(YAW).pdf":      "H1818"
   
    # Add more mappings as needed
    }

    REPAIR_SYNONYMS = {
        # Core driver-assist sensors
        "Steering Angle Sensor":           "SAS",
        "Yaw Rate Sensor":                 "YAW",
        "G Force Sensor":                  "G-Force",
        "Seat Weight Sensor":              "SWS",
        "Adaptive Head Lamps":             "AHL",
        "Night Vision":                    "NV",
        "Heads Up Display":                "HUD",
        "Electronic Stability Control Relearn": "ESC",
        "Airbag Disengagement/Engagement": "SRS D&E",
        "Steering Column Inspection":      "SCI",
        "Steering Rack Relearn":           "SRR",
        "Headlamp Initialization":         "HLI",
        "Tire Pressure Monitor Relearn":   "TPMS",
        "Seat Belt Inspection":            "SBI",

        # Electric/hybrid battery systems
        "Battery Disengagement":           "EBDE (1)",
        "Battery Engagement":              "EBDE (2)",
        "Hybrid Disengagement":            "HDE (1)",
        "Hybrid Engagement":               "HDE (2)",

        # Miscellaneous vehicle-level relearns
        "Liftgate Relearn":                "LGR",
        "Power Seat Initialization":       "PSI",
        "Window Relearn":                  "WRL",

        # Module programming routines
        "Powertrain Control Module Program":       "PCM",
        "Transmission Control Module Program":     "TRANS",
        "Airbag Control Module Program":           "AIR",
        "Antilock Brake Control Module Program":   "ABS",
        "Body Control Module Program":             "BCM",
        "Key Program":                             "KEY",
        "Key FOB Relearn":                         "FOB",

        # HVAC & coolant
        "Heating, Air Conditioning, Ventilation EVAC":    "HVAC (1)",
        "Heating, Air Conditioning, Ventilation Recharge":"HVAC (2)",
        "Coolant Services":                               "COOL",

        # Headset resets
        "Headset Reset (Spring Style)":       "HEAD (1)",
        "Headset Reset (Squib Style)":        "HEAD (2)",
    }

    HYPERLINK_COLUMN_INDEX = 12  # Default is Column L (can change to 11 for K, etc.)

    #################################################################################################################################################

    # Class objects holding information about files and folders in a given sharepoint location
    class EntryTypes(Enum):
        """Enumeration holding the different types of entries in a sharepoint location"""
        UNDEFINED = "N/A",
        FILE_ENTRY = "File",
        FOLDER_ENTRTY = "Folder"
    class SharepointEntry:
        """Base type for a file or folder in a sharepoint location"""
        
        # Attributes for a sharepoint entry object
        entry_name: str = None                                  # The name of the entry in the sharepoint location
        entry_link: str = None                                  # The link to the entry in the sharepoint location
        entry_heirarchy: str = None                             # The folder path/heirarchy to the entry in our sharepoint location        
        entry_type: 'SharepointExtractor.EntryType' = None      # The type of entry in the sharepoint location
        
        def __init__(self, name: str, heirarchy: str, link: str, type: 'SharepointExtractor.EntryType') -> 'SharepointExtractor.SharepointEntry':
            """
            CTOR for building a new SharepointEntry object
            
            ----------------------------------------------
            
            name: str
                The name of the entry in the sharepoint folder
            heirarchy: str  
                The heirachy path to the entry in the sharepoint folder
            link: str
                The link to the entry in the sharepoint folder
            type: SharepointExtractor.EntryType
                The type of entry in our sharepoint folder
            """
            
            # Assign properties of the entry to this instance and exit out 
            self.entry_name = name
            self.entry_link = link
            self.entry_type = type
            self.entry_heirarchy = heirarchy

    #################################################################################################################################################  

    def __init__(self, sharepoint_link: str, excel_file_path: str, debug_run: bool = False) -> 'SharepointExtractor':
        """
        CTOR for a new SharepointExtractor. Takes the link to the requested sharepoint location 
        and prepares to extract all file and folder links
        
        ----------------------------------------------
        
        sharepoint_link: str 
            The link to the sharepoint location for the given make
        excel_file_path: str 
            The fully qualified path to the excel file holding our ADAS SI
        debug_run: bool
            When true, we don't actually get any file links.
            Useful for quickly testing operations without waiting for links to generate.
            Defaults to false.
        """

        # ── Header utils (header-only; no fallbacks) ─────────────────────────────────
        def _norm_hdr(self, s: str) -> str:
            # Normalize header text: trim, uppercase, collapse inner spaces
            import re
            return re.sub(r"\s+", " ", str(s).strip().upper())
    
        # ── Header-only helpers ─────────────────────────────────────────────
        def _header_colmap_(self, ws):
            """Return 1-based indices for required headers; no fallbacks."""
            import re
            def norm(s): return re.sub(r"\s+", " ", str(s).strip().upper())
        
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = {}
            for i, c in enumerate(header_cells):
                if c.value is None: 
                    continue
                headers[norm(c.value)] = i + 1  # 1-based
        
            def pick(*names):
                for n in names:
                    key = norm(n)
                    if key in headers:
                        return headers[key]
                return None
        
            colmap = {
                "year":   pick("Year"),
                "make":   pick("Make"),
                "model":  pick("Model"),
                "system": pick(  # ADAS + Repair
                    "SME Generic System Name",
                    "Protech Generic System Name",
                    "Generic System Name",
                    "System Name",
                    "System",
                ),
                "hyperlink": pick(  # includes “Service Information”
                    "Hyperlink", "Link", "URL",
                    "Service Information", "Service Information (URL)",
                    "SI", "SI Link", "SI URL",
                ),
            }
            missing = [k for k in ("year","make","model","system") if not colmap.get(k)]
            if missing:
                raise ValueError(f"Missing required header(s): {', '.join(missing)}")
            return colmap
        
        def _cell_val_upper(self, row_tuple, one_based_idx):
            if not one_based_idx:
                return ""
            i = one_based_idx - 1
            if i < 0 or i >= len(row_tuple):
                return ""
            v = row_tuple[i].value
            return (str(v).strip().upper() if v is not None else "")



        self.mode = sys.argv[4] if len(sys.argv) > 4 else "adas"
        self.repair_mode = self.mode == "repair"
        self.selected_adas = sys.argv[3].split(",") if len(sys.argv) > 3 else []
        self.cleanup_mode = sys.argv[5] == "cleanup" if len(sys.argv) > 5 else False
        self.excel_mode = sys.argv[6] if len(sys.argv) > 6 else "og"
        self.broken_entries = []  # ← Store broken hyperlinks here for cleanup mode
        
        # Always set system_col and hyperlink index based on mode
        if self.repair_mode and self.excel_mode == "og":
            self.system_col, self.HYPERLINK_COLUMN_INDEX = 4, 8
        elif not self.repair_mode and self.excel_mode == "og":
            self.system_col, self.HYPERLINK_COLUMN_INDEX = 5, 12
        elif not self.repair_mode and self.excel_mode == "new":
            self.system_col, self.HYPERLINK_COLUMN_INDEX = 21, 11
        else:
            print("⚠️ Unsupported mode/Excel combination in cleanup mode")
            self.system_col, self.HYPERLINK_COLUMN_INDEX = None, None
          
        # Store attributes for the Extractor on this instance
        self.__DEBUG_RUN__ = debug_run
        self.sharepoint_links = sharepoint_link.split('||') if '||' in sharepoint_link else [sharepoint_link]
        self.sharepoint_link = self.sharepoint_links[0]
        self.excel_file_path = excel_file_path
        self.selected_adas = sys.argv[3].split(",") if len(sys.argv) > 3 else []

        # Check installed Chrome version
        def get_chrome_version():
            try:
                output = subprocess.check_output(
                    r'reg query "HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon" /v version',
                    shell=True, stderr=subprocess.DEVNULL, text=True
                )
                match = re.search(r"version\s+REG_SZ\s+([^\s]+)", output)
                return match.group(1) if match else None
            except Exception:
                return None
        
        def get_local_chromedriver_version():
            base = os.path.dirname(chromedriver_autoinstaller.__file__)
            if not os.path.exists(base):
                return None
            for folder in os.listdir(base):
                if folder.isdigit():
                    return folder
            return None
        
        chrome_version = get_chrome_version()
        driver_version = get_local_chromedriver_version()
        
        if chrome_version and driver_version:
            if not chrome_version.startswith(driver_version):
                mismatch_path = os.path.join(os.path.dirname(chromedriver_autoinstaller.__file__), driver_version)
                if os.path.exists(mismatch_path):
                    print(f"Deleting mismatched ChromeDriver v{driver_version} for Chrome v{chrome_version}")
                    shutil.rmtree(mismatch_path)
        
        # Always install the correct version
        chromedriver_autoinstaller.install()
               
        # Then just start Chrome normally:
        self.selenium_driver = webdriver.Chrome(
            options=self.__generate_chrome_options__()
        )
        self.selenium_wait = WebDriverWait(self.selenium_driver, 10)

        # Navigate to the main SharePoint page for Acura
        print("Navigating to main SharePoint page link now...")
        self.selenium_driver.get(self.sharepoint_link)
 
        # Wait until the element with the specified XPath is found, or until 60 seconds have passed
        try: WebDriverWait(self.selenium_driver, self.__MAX_WAIT_TIME__).until(EC.presence_of_element_located((By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__)))
        except: 
            print(f"The element was not found within {self.__MAX_WAIT_TIME__} seconds.")
            raise Exception(f"ERROR! Failed to find valid login state after {self.__MAX_WAIT_TIME__} seconds!")
        
        # Find the make of the folder for the current sharepoint link and store it.
        self.sharepoint_make = self.selenium_driver.find_elements(By.XPATH, self.__ONEDRIVE_PAGE_NAME_LOCATOR__)[-1].get_attribute("innerText").strip()
        print(f"Configured new SharepointExtractor for {self.sharepoint_make} correctly!")
        
        if self.sharepoint_make.lower() == "toyota" and self.repair_mode:
           self.HYPERLINK_COLUMN_INDEX = 10  # Excel column J
              
    def __cleanup_across_all_links__(self) -> tuple[list, list]:
        """
        Batch resolve self.broken_entries across ALL self.sharepoint_links in a single pass.
        Visits each link once, then navigates Year → Model and matches as many systems as possible.
        Returns ([], unique_matches)
        """
        from collections import defaultdict
    
        print("🔍 Clean up Mode: Resolving broken links across all SharePoint links in a single pass...")
    
        if not getattr(self, "broken_entries", None):
            print("ℹ️ No broken entries present; nothing to resolve.")
            return [], []
    
        # Group Excel problems by Year → Model → [Systems]
        grouped = defaultdict(lambda: defaultdict(list))
        for _, (yr, mk, mdl, sys) in self.broken_entries:
            # Normalize Excel system through REPAIR_SYNONYMS if available
            for desc, acronym in getattr(self, "REPAIR_SYNONYMS", {}).items():
                normalized = acronym.replace(" ", "").replace("&", "").replace("-", "").upper()
                if normalized == sys.strip().upper():
                    sys = acronym
                    break
            grouped[str(yr).strip()][str(mdl).strip()].append(str(sys).strip())
    
        resolved = set()           # (year, model, system)
        matched_files = []
    
        for root_link in getattr(self, "sharepoint_links", [self.sharepoint_link]):
            try:
                self.selenium_driver.get(root_link)
                time.sleep(1.0)
            except Exception as e:
                print(f"⚠️ Could not navigate to link: {root_link} → {e}")
                continue
    
            # List year folders once for this root
            try:
                year_folders, _ = self.__get_folder_rows__()
            except Exception as e:
                print(f"⚠️ Could not read top-level year folders for link: {e}")
                continue
    
            for yr, models in grouped.items():
                target_year = next((f for f in year_folders if yr == f.entry_name.strip()), None)
                if not target_year:
                    continue
    
                # Enter year folder once
                self.selenium_driver.get(target_year.entry_link)
                time.sleep(0.8)
    
                # List model folders once for this year
                try:
                    model_folders, _ = self.__get_folder_rows__()
                except Exception as e:
                    print(f"⚠️ Could not read model folders under '{yr}': {e}")
                    continue
    
                for mdl, sys_list in models.items():
                    target_model = next((f for f in model_folders if mdl.upper() == f.entry_name.strip().upper()), None)
                    if not target_model:
                        continue
    
                    # 👇 NEW: announce each attempt BEFORE navigating into the model folder/files
                    for sys_name in list(sys_list):
                        if (yr, mdl, sys_name) in resolved:
                            continue
                        print(f"🔗 Attempting to gather link for {yr} {self.sharepoint_make} {mdl} ({sys_name})")
    
                    # Enter model folder once (navigation happens AFTER the attempt logs)
                    self.selenium_driver.get(target_model.entry_link)
                    time.sleep(0.8)
    
                    # Grab file rows once
                    try:
                        table = WebDriverWait(self.selenium_driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__))
                        )
                        rows = table.find_elements(By.XPATH, self.__ONEDRIVE_TABLE_ROW_LOCATOR__)
                    except Exception as e:
                        print(f"⚠️ Failed to list files in {yr}/{mdl}: {e}")
                        continue
    
                    # Prepare row names cache for matching
                    row_names = [self.__get_row_name__(r) for r in rows]
    
                    for sys_name in list(sys_list):
                        if (yr, mdl, sys_name) in resolved:
                            continue
    
                        base_sys = re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", sys_name).strip()
                        regex = re.compile(
                            rf"(?<![A-Za-z0-9])\(?{re.escape(base_sys)}(?:\s*\d+)?\)?(?![A-Za-z0-9])",
                            re.IGNORECASE
                        )
    
                        direct_row = None
                        no_doc_row = None
    
                        for row, name in zip(rows, row_names):
                            if name.lower().startswith("no ") and regex.search(name):
                                no_doc_row = row
    
                            if not regex.search(name):
                                continue
    
                            # Year match
                            ym = re.search(r"(20\d{2})", name)
                            if not ym or ym.group(1).strip() != yr:
                                continue
    
                            # Model presence (clean parentheses and year, compare upper)
                            cleaned = re.sub(r"\(.*?\)", "", name)
                            cleaned = re.sub(r"(20\d{2})", "", cleaned).replace(".pdf", "").strip().upper()
                            # ★ NEW: allow small model typos in file names
                            clean_model = mdl.strip().upper()
                            if clean_model not in cleaned.upper():
                                if _similar(clean_model, cleaned) < 0.75:
                                    continue

    
                            direct_row = row
                            break
    
                        if direct_row:
                            link = self.__get_encrypted_link__(direct_row)
                            if link:
                                matched_files.append(
                                    SharepointExtractor.SharepointEntry(
                                        name=self.__get_row_name__(direct_row),
                                        heirarchy=self.__get_entry_heirarchy__(direct_row),
                                        link=link,
                                        type=SharepointExtractor.EntryTypes.FILE_ENTRY
                                    )
                                )
                                resolved.add((yr, mdl, sys_name))
                                print(f"✅ Direct match: {yr} {self.sharepoint_make} {mdl} ({sys_name})")
                                continue
    
                        if no_doc_row:
                            orig_name = self.__get_row_name__(no_doc_row)
                            link = self.__get_encrypted_link__(no_doc_row)
                            forced = f"{yr} {self.sharepoint_make} {mdl} ({sys_name}).pdf"
                            if link:
                                matched_files.append(
                                    SharepointExtractor.SharepointEntry(
                                        name=forced,
                                        heirarchy=self.__get_entry_heirarchy__(no_doc_row),
                                        link=link,
                                        type=SharepointExtractor.EntryTypes.FILE_ENTRY
                                    )
                                )
                                resolved.add((yr, mdl, sys_name))
                                print(f"ℹ️ No real {sys_name} doc — using NO-doc: {orig_name}")
                                print(f"   ↳ Renaming for placement as: {forced}")
    
        # Dedupe by name
        seen = set()
        unique_matches = []
        for entry in matched_files:
            if entry.entry_name not in seen:
                seen.add(entry.entry_name)
                unique_matches.append(entry)
    
        print(f"📥 Matched {len(unique_matches)} files for repair across all links.")
        return [], unique_matches

       
    def extract_contents(self) -> tuple[list, list]:
            """
            Extracts the file and folder links from the defined sharepoint location for the current extractor object.
            Returns a tuple of lists. The first list holds all of our SharepointEntry objects for the folders in the sharepoint,
            and the second list holds all of our SharepointEntry objects for the files in the sharepoint.
            """
    
            time.sleep(2.0)
    
            if self.cleanup_mode:
                print("🔍 Clean up Mode: Navigating per broken link...")
    
                # If we were passed multiple links, use the batched resolver
                if hasattr(self, "sharepoint_links") and len(self.sharepoint_links) > 1:
                    return self.__cleanup_across_all_links__()
    
                matched_files = []
    
                for _, (yr, mk, mdl, sys) in self.broken_entries:
                    # Reverse map if Excel gave us a normalized string like "GFORCE"
                    for desc, acronym in self.REPAIR_SYNONYMS.items():
                        normalized = acronym.replace(" ", "").replace("&", "").replace("-", "").upper()
                        if normalized == sys.strip().upper():
                            sys = acronym
                            break
    
                    print(f"🔎 Seeking: {yr} ➝ {mk} ➝ {mdl} ➝ {sys}")
    
                    # STEP 1: reset to root folder
                    self.selenium_driver.get(self.sharepoint_link)
                    time.sleep(2.0)
    
                    # STEP 2: find year folder
                    year_folders, _ = self.__get_folder_rows__()
                    target_year = next((f for f in year_folders if yr.strip() == f.entry_name.strip()), None)
                    if not target_year:
                        print(f"❌ Year folder '{yr}' not found.")
                        continue
                    self.selenium_driver.get(target_year.entry_link)
                    time.sleep(1.5)
    
                    # STEP 3: find model folder
                    model_folders, _ = self.__get_folder_rows__()
                    target_model = next((f for f in model_folders if mdl.strip().upper() == f.entry_name.strip().upper()), None)
                    if not target_model:
                        print(f"❌ Model folder '{mdl}' not found under year '{yr}'.")
                        continue
                    self.selenium_driver.get(target_model.entry_link)
                    time.sleep(1.5)
    
                    # ── STEP 4: look for the file matching the system ──
                    try:
                        table = WebDriverWait(self.selenium_driver, 15).until(
                            EC.presence_of_element_located((By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__))
                        )
                        rows = table.find_elements(By.XPATH, self.__ONEDRIVE_TABLE_ROW_LOCATOR__)
    
                        no_doc_row = None
                        # strip any trailing “(n)” from sys to get the pure acronym
                        base_sys = re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", sys).strip()
    
                        # one regex to catch "(ACC)", "(ACC 1)", "ACC 1", "ACC", etc.
                        regex = re.compile(
                            rf"(?<![A-Za-z0-9])"       # no alnum just before
                            rf"\(?"                    # optional "("
                            rf"{re.escape(base_sys)}"  # your acronym
                            rf"(?:\s*\d+)?"            # optional digits
                            rf"\)?"                    # optional ")"
                            rf"(?![A-Za-z0-9])",       # no alnum just after
                            re.IGNORECASE
                        )
    
                        for row in rows:
                            name = self.__get_row_name__(row)
    
                            # store NO-doc for fallback if it mentions the same system
                            if name.lower().startswith("no ") and regex.search(name):
                                no_doc_row = row
    
                            # skip any row that doesn't match our system‐regex
                            if not regex.search(name):
                                continue
    
                            # ── Year check ──
                            ym = re.search(r"(20\d{2})", name)
                            if not ym or ym.group(1).strip() != yr.strip():
                                continue
    
                            # ── Model check ──
                            clean = re.sub(r"\(.*?\)", "", name)
                            clean = re.sub(r"(20\d{2})", "", clean).replace(".pdf", "").strip().upper()
                            if mdl.strip().upper() not in clean:
                                continue
    
                            # ✅ direct match!
                            link = self.__get_encrypted_link__(row)
                            matched_files.append(
                                SharepointExtractor.SharepointEntry(
                                    name=name,
                                    heirarchy=self.__get_entry_heirarchy__(row),
                                    link=link,
                                    type=SharepointExtractor.EntryTypes.FILE_ENTRY
                                )
                            )
                            print(f"✅ Direct match: {name}")
                            break
                        else:
                            # no direct hit → fall back on NO-doc row
                            if no_doc_row:
                                orig   = self.__get_row_name__(no_doc_row)
                                link   = self.__get_encrypted_link__(no_doc_row)
                                forced = f"{yr} {self.sharepoint_make} {mdl} ({sys}).pdf"
                                matched_files.append(
                                    SharepointExtractor.SharepointEntry(
                                        name=forced,
                                        heirarchy=self.__get_entry_heirarchy__(no_doc_row),
                                        link=link,
                                        type=SharepointExtractor.EntryTypes.FILE_ENTRY
                                    )
                                )
                                print(f"ℹ️ No real {sys} doc found — using NO-doc: {orig}")
                                print(f"   ↳ Renaming for placement as: {forced}")
                                # only log fallback once
                                print(f"Processed (fallback) hyperlink for {yr}, {mk}, {mdl}, {sys}")
    
                    except Exception as e:
                        print(f"⚠️ Failed to locate system file row: {e}")
    
                # ── DEDUPE matched_files by entry_name ──
                seen = set()
                unique_matches = []
                for entry in matched_files:
                    if entry.entry_name not in seen:
                        seen.add(entry.entry_name)
                        unique_matches.append(entry)
                return [], unique_matches
    
            # ── FULL MODE (cleanup_mode == False) ──
            sharepoint_folders, sharepoint_files = self.__get_folder_rows__()
    
            # Compile regex patterns for the selected ADAS systems if any are selected
            if self.repair_mode:
                adas_patterns = [re.compile(re.escape(rs), re.IGNORECASE) for rs in self.selected_adas] if self.selected_adas else None
            else:
                adas_patterns = (
                    [re.compile(rf"\({re.escape(adas)}\s*\d*\)", re.IGNORECASE) for adas in self.selected_adas]
                    if self.selected_adas else None
                )
    
            # … rest of your full‐indexing logic unchanged …
            filtered_files = []
            start_time = time.time()
            while sharepoint_folders:
                folder_link = sharepoint_folders.pop(0).entry_link
                child_folders, child_files = self.__get_folder_rows__(folder_link)
                time.sleep(3.0)
                sharepoint_folders.extend(child_folders)
    
                if self.repair_mode and self.selected_adas:
                    for file_entry in child_files:
                        # … existing repair filtering …
                        filtered_files.append(file_entry)
                elif adas_patterns:
                    for file_entry in child_files:
                        if any(p.search(file_entry.entry_name) for p in adas_patterns):
                            filtered_files.append(file_entry)
                else:
                    filtered_files.extend(child_files)
    
                print(f'{len(sharepoint_folders)} Folders Remain | {len(filtered_files)} Files Indexed')
    
            # ── DEDUPE filtered_files by entry_name ──
            seen = set()
            unique_files = []
            for entry in filtered_files:
                if entry.entry_name not in seen:
                    seen.add(entry.entry_name)
                    unique_files.append(entry)
    
            elapsed_time = time.time() - start_time
            print(f"Indexing routine took {elapsed_time:.2f} seconds.")
            return sharepoint_folders, unique_files

    def __simulate_entry_from_no_entry__(self, entry_name: str, real_link: str, heirarchy: str, sibling_files: list) -> 'SharepointExtractor.SharepointEntry':
        """
        Replaces 'No XYZ...' file with a simulated one using year/make/model from known good entries.
        Also tags the returned entry so downstream writers can style it dark yellow.
        """
        # Try to extract acronym (inside parentheses or inferred from known acronyms)
        acronym_match = re.search(r'\((.*?)\)', entry_name)
        if acronym_match:
            acronym = acronym_match.group(1)
        else:
            acronym = next((key for key in self.__DEFINED_MODULE_NAMES__ if key in entry_name.upper()), None)
        if not acronym:
            return None
    
        # Look for a similar file to simulate from
        for sibling in sibling_files:
            sibling_name = sibling.entry_name
            year_match = re.search(r'(20\d{2})', sibling_name)
            if not year_match:
                continue
            year = year_match.group(1)
    
            # Strip down to extract model like original logic
            base_name = re.sub(r'(20\d{2})', '', sibling_name)
            base_name = base_name.replace(".pdf", "").strip()
            base_name = re.sub(re.escape(self.sharepoint_make), "", base_name, flags=re.IGNORECASE).strip()
    
            tokens = []
            mod_names = {m.upper() for m in self.__DEFINED_MODULE_NAMES__}
            for token in base_name.split():
                if token.startswith("("):
                    content = token.strip("()")
                    if content.upper() in mod_names:
                        break
                    tokens.append(content)
                elif token.upper().strip("()[]") in mod_names:
                    break
                else:
                    tokens.append(token)
    
            model = " ".join(tokens)
            if model:
                new_name = f"{year} {self.sharepoint_make} {model} ({acronym})"
                se = SharepointExtractor.SharepointEntry(
                    name=new_name,
                    heirarchy=heirarchy,
                    link=real_link,
                    type=SharepointExtractor.EntryTypes.FILE_ENTRY
                )
                # 🔖 Mark this simulated entry as originating from a "No ..." document
                setattr(se, "is_no_doc", True)
                # (Optional) keep the original "No ..." filename for debugging/auditing
                setattr(se, "original_no_doc_name", entry_name)
                return se
    
        return None


    def populate_excel_file(self, file_entries: list) -> None:
        """
        Populates the excel file for the current make and stores all hyperlinks built in correct 
        locations.
        """
        import time, re, openpyxl
        start_time = time.time()
    
        # Load the Excel file
        model_workbook = openpyxl.load_workbook(self.excel_file_path)
        
        sheet_name = 'Model Version'
        if sheet_name not in model_workbook.sheetnames:
            print(f"WARNING: Sheet '{sheet_name}' not found. Defaulting to first sheet.")
            model_worksheet = model_workbook.active
        else:
            model_worksheet = model_workbook[sheet_name]
    
        print(f"Workbook loaded successfully: {self.excel_file_path}")
    
        # ── Fix stale hyperlink objects (copied OneDrive/SharePoint links) ──
        for row in model_worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.hyperlink and isinstance(cell.value, str) and cell.value.startswith("http"):
                    if cell.hyperlink.target != cell.value.strip():
                        print(f"🔧 Fixing hyperlink at {cell.coordinate}")
                        cell.hyperlink = cell.value.strip()
            
        # 🧭 Header-only: detect column indices from the sheet
        self.colmap = self._header_colmap_(model_worksheet)
        
        # Ensure the hyperlink column exists (create "Service Information" if missing)
        self._ensure_hyperlink_column(model_worksheet, "Service Information Hyperlink")
                

        # Index rows once per call
        self.row_index = self.__build_row_index__(model_worksheet, self.repair_mode)
    
        # ─────────────────────────────────────────────────────────────────────────────
        # KEY FIX: Only scan (Phase 1) when we are NOT applying matched files.
        # If file_entries is non-empty, we are in Phase 2 and must SKIP the rescan.
        # ─────────────────────────────────────────────────────────────────────────────
        do_phase1_scan = bool(self.cleanup_mode and not file_entries)
    
        if do_phase1_scan:
            print("🧹 Clean up Mode: Scanning for broken hyperlinks (Phase 1)...")

            # init once per scan
            self.broken_entries = []

            # ─────────────────────────────────────────────
            # Header-only column detection for cleanup
            # Ignore OG/NEW; always use Service Information column
            # ─────────────────────────────────────────────
            try:
                # Reuse existing header map if present, otherwise build it
                colmap = getattr(self, "colmap", None)
                if not colmap:
                    colmap = self._header_colmap(model_worksheet)
                    self.colmap = colmap

                # REQUIRED: system column (SME/Protech Generic System Name etc.)
                system_col = colmap["system"]

                # Hyperlink / Service Information column
                hyperlink_col = colmap.get("hyperlink")
                if not hyperlink_col:
                    # Fallback: ensure/create a Service Information column
                    # (uses your existing helper if you added it)
                    hyperlink_col = self._ensure_hyperlink_column(
                        model_worksheet,
                        "Service Information",
                        "Service Information Hyperlink"
                    )
            except Exception as e:
                print(f"⚠️ Failed to detect Service Information/System columns in cleanup mode: {e}")
                model_workbook.save(self.excel_file_path)
                model_workbook.close()
                return

            filename_col = 1  # Adjust if your file names are stored elsewhere

    
            # total rows for progress bar
            self.total_rows_to_check = sum(
                1 for key in self.row_index.values()
                if model_worksheet.cell(row=key, column=system_col).value
            )
            self.rows_checked = 0
            self.update_current_manufacturer_progress()
    
            for key, row in self.row_index.items():
                year, make, model, system_from_index = key
                print(f"🔎 Checking row {row}: Year={year}, Make={make}, Model={model}, System={system_from_index}")
    
                cell = model_worksheet.cell(row=row, column=hyperlink_col)
                if cell.hyperlink:
                    url = cell.hyperlink.target
                else:
                    url = str(cell.value).strip() if cell.value else None
    
                if not url:
                    self.rows_checked += 1
                    self.update_current_manufacturer_progress()
                    continue
    
                file_name_cell = model_worksheet.cell(row=row, column=filename_col)
                file_name = str(file_name_cell.value).strip() if file_name_cell.value else None
    
                system_cell = model_worksheet.cell(row=row, column=system_col)
                raw_value = system_cell.value
                system_name = str(raw_value).strip() if raw_value else "UNKNOWN"
    
                if url.lower() == "hyperlink not available":
                    print(f"⏩ Skipping 'Hyperlink Not Available' placeholder at row {row}")
                    self.rows_checked += 1
                    self.update_current_manufacturer_progress()
                    continue
                if not (url and url.lower().startswith("http")):
                    print(f"⏩ Skipping non-URL text at row {row}: {url}")
                    self.rows_checked += 1
                    self.update_current_manufacturer_progress()
                    continue
    
                if self.is_broken_sharepoint_link(url, file_name=file_name):
                    yr, mk, mdl, _ = key
                    print(f"🔧 Broken link found → Year: {yr}, Make: {mk}, Model: {mdl}, System: {system_name}")
                    self.broken_entries.append((row, (yr, mk, mdl, system_name)))
    
                self.rows_checked += 1
                self.update_current_manufacturer_progress()
    
            print(f"🔍 Found {len(self.broken_entries)} broken links. Handing off to Phase 2...")
    
            # Phase marker (optional, if you use it elsewhere)
            self._cleanup_phase = "apply"
    
            # Save after scan; no files to apply in this call
            print(f"Saving updated changes to {self.sharepoint_make} sheet now...")
            model_workbook.save(self.excel_file_path)
            model_workbook.close()
            elapsed_time = time.time() - start_time
            print(f"Sheet population routine took {elapsed_time:.2f} seconds.")
            return  # <── IMPORTANT: finish Phase 1 call here.
    
        else:
            if self.cleanup_mode:
                print("🧹 Phase 2: Applying fixes only (skipping rescan).")
    
        # ─────────────────────────────────────────────────────────────────────────────
        # Phase 2 (or normal mode): apply file_entries into the sheet
        # ─────────────────────────────────────────────────────────────────────────────
        current_model = ""
        adas_last_row = {}
    
        for file_entry in file_entries:
            #print(f"Processing file: {file_entry.entry_name}")
            file_name = file_entry.entry_name
    
            # Cleanup mode: force NO-docs into correct system row when needed
            if self.cleanup_mode and file_name.lower().startswith("no "):
                original_no_doc_name = file_name
                for _, (yr, mk, mdl, sys) in self.broken_entries:
                    if (yr in file_name or yr == "Unknown") and mdl.replace(" ", "").lower() in file_name.replace(" ", "").lower():
                        print(f"🔄 Forcing NO-doc {file_name} into system row: {sys}")
                        file_name = f"{yr} {self.sharepoint_make} {mdl} ({sys})"
                        file_entry.entry_name = file_name
                        print(f"   ↳ Renaming NO-doc for proper placement: {file_name}")
            
                        # 🔖 mark the item so __update_excel__ knows to color it yellow
                        setattr(file_entry, "is_no_doc", True)
                        self._last_is_no_doc = True
            
                        if hasattr(self, "__add_red_text_marker"):
                            self.__add_red_text_marker(
                                model_worksheet, yr, self.sharepoint_make, mdl, sys, original_no_doc_name
                            )
                        break
            
    
            # Synonym normalization
            for desc, acr in self.REPAIR_SYNONYMS.items():
                pattern = f"({desc})"
                if pattern in file_name:
                    file_name = file_name.replace(pattern, f"({acr})")
                    file_entry.entry_name = file_name
                    break
    
            # Year
            year_match = re.search(r'(20\d{2})', file_name)
            file_year = year_match.group(1) if year_match else "Unknown"
    
            # Model
            base_name = re.sub(r'(20\d{2})', '', file_name)
            base_name = base_name.replace(".pdf", "").strip()
            base_name = re.sub(re.escape(self.sharepoint_make), "", base_name, flags=re.IGNORECASE).strip()
    
            model_tokens = []
            mod_names = {m.upper() for m in self.__DEFINED_MODULE_NAMES__}
    
            for token in base_name.split():
                if token.startswith("("):
                    content = token.strip("()")
                    if content.strip().upper() in mod_names:
                        break
                    else:
                        model_tokens.append(content)
                elif token.upper().strip("()[]") in mod_names:
                    break
                else:
                    model_tokens.append(token)
    
            file_model = " ".join(model_tokens).strip() if model_tokens else "Unknown"
            if file_model == "Unknown":
                segments = file_entry.entry_heirarchy.split("\\")
                if len(segments) > 1:
                    file_model = segments[-2]
    
            if file_model != current_model:
                current_model = file_model
                adas_last_row = {}
    
            if file_entry.entry_link is None:
                print(f"❌ Could not retrieve link for: {file_name}")
                error_text = f"{file_name} - Hyperlink Error, Check SharePoint"
                self.__update_excel__(
                    model_worksheet,
                    file_year,
                    file_model,
                    error_text,
                    "",
                    adas_last_row,
                    None
                )
                continue
    
            #if self.__update_excel_with_whitelist__(model_worksheet, file_name, file_entry.entry_link):
                #if self.cleanup_mode:
                    #print(f"Fixed hyperlink for: {file_entry.entry_name}")
                #continue
            self._last_is_no_doc = bool(getattr(file_entry, "is_no_doc", False)) or file_name.lower().startswith("no ")

            self.__update_excel__(
                model_worksheet,
                file_year,
                file_model,
                file_name,
                file_entry.entry_link,
                adas_last_row,
                None
            )
    
            #if self.cleanup_mode:
                #print(f"Fixed hyperlink for: {file_entry.entry_name}")
    
        print(f"Saving updated changes to {self.sharepoint_make} sheet now...")
        model_workbook.save(self.excel_file_path)
        model_workbook.close()
    
        elapsed_time = time.time() - start_time
        print(f"Sheet population routine took {elapsed_time:.2f} seconds.")
    

    def update_current_manufacturer_progress(self, *, checked=None, total=None):
        """
        Emit machine-readable progress for the Current Manufacturer bar.
        The GUI (Hyper.py) should listen for lines starting with 'CM_PROGRESS'.
        """
        try:
            if checked is not None:
                self.rows_checked = checked
            if total is not None:
                self.total_rows_to_check = total
    
            total = getattr(self, 'total_rows_to_check', 0) or 0
            done  = getattr(self, 'rows_checked', 0) or 0
            if total <= 0:
                return
    
            percent = min(100, int((done / total) * 100))
            print(f"CM_PROGRESS {done}/{total} ({percent}%)")
            sys.stdout.flush()
        except Exception as e:
            print(f"CM_PROGRESS_ERROR {e}")
            sys.stdout.flush()

    def __generate_chrome_options__(self) -> Options:
        """
        Configures Chrome to use a valid custom user profile.
        Chrome requires --user-data-dir to be a non-default location for automation.
        """
    
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-browser-side-navigation")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    
        # Define a dedicated automation user data directory
        home_dir = os.path.expanduser("~")
        automation_profile_base = os.path.join(home_dir, "ChromeAutomationProfiles")
    
        # Try Default first, then Profile 1
        profile_name = "Default"
        if not os.path.exists(os.path.join(home_dir, "AppData", "Local", "Google", "Chrome", "User Data", "Default")):
            profile_name = "Profile 1"
    
        # Copy the real profile to a custom folder if not already copied
        original_profile = os.path.join(home_dir, "AppData", "Local", "Google", "Chrome", "User Data", profile_name)
        target_profile = os.path.join(automation_profile_base, profile_name)
    
        if not os.path.exists(target_profile):
            import shutil
            shutil.copytree(original_profile, target_profile)
    
        chrome_options.add_argument(f"--user-data-dir={automation_profile_base}")
        chrome_options.add_argument(f"--profile-directory={profile_name}")
    
        print(f"[Chrome Profile] Using copied profile: {profile_name}")
        return chrome_options
  
    def __is_row_folder__(self, row_element: WebElement) -> bool:
        # grab just the real name
        name = self.__get_row_name__(row_element).splitlines()[0]
        # if it’s got a .pdf/.docx/.xlsx/.pptx *anywhere* in that name, it’s a file
        if re.search(r'\.(pdf|docx?|xlsx?|pptx?)\b', name, re.IGNORECASE):
            return False
        return True
  
    def __get_row_name__(self, row_element: WebElement) -> str:
        """
        Read only the *first line* of the aria-label or innerText,
        so we never pull in date/author on subsequent prints.
        """
        raw = row_element.get_attribute("aria-label")
        if raw and raw.strip():
            # only the actual name, before any newline/date/author lines
            return raw.strip().splitlines()[0]

        # Fallback to the visible text, again only first line
        text = row_element.text.strip()
        return text.splitlines()[0]

    def __get_unencrypted_link__(self, row_element: WebElement) -> str:
        """
        Build the “AllItems.aspx?id=…” URL for this row, but:
          • Only use the first line of the aria-label (the real name)
          • Strip *exactly* one trailing “%2F” if present (no more)
          • Preserve everything after the first '&' (viewid, ga, noAuthRedirect…)
        """
        # 1) grab only the real name (no date/author)
        full_label = self.__get_row_name__(row_element)        # e.g. "2012\nJune 15…"
        item_name  = full_label.splitlines()[0]                # “2012”

        # 2) split off base & query
        current = self.selenium_driver.current_url
        try:
            base_url, query = current.split('?', 1)
        except ValueError:
            raise RuntimeError(f"Unexpected URL format: {current!r}")

        # 3) split into the id= piece and the rest
        id_part, rest = query.split('&', 1)
        key, old_val = id_part.split('=', 1)

        # 4) remove at most one trailing "%2F"
        if old_val.endswith('%2F'):
            base_id = old_val[:-3]
        else:
            base_id = old_val

        # 5) URL-encode only the item_name, tack it on
        encoded_name = urllib.parse.quote(item_name, safe='')
        new_val      = f"{base_id}%2F{encoded_name}"

        # 6) rebuild
        return f"{base_url}?{key}={new_val}&{rest}"
      
    def __get_encrypted_link__(self, row_element: WebElement) -> str:
        """
        Tries to generate a SharePoint share link for the given row.
        Retries up to 5 times. Returns None if it fails every time or hits a 120s timeout.
        """
    
        if self.__DEBUG_RUN__:
            return f"Link For: {self.__get_row_name__(row_element)}"
    
        starting_clipboard_content = self.__get_clipboard_content__()
        selector_locator = ".//div[@role='gridcell' and contains(@data-automationid, 'row-selection-')]"
        selector_element = WebDriverWait(row_element, self.__MAX_WAIT_TIME__) \
            .until(EC.presence_of_element_located((By.XPATH, selector_locator)))
    
        # scroll into view & click
        self.selenium_driver.execute_script("arguments[0].scrollIntoView(true);", selector_element)
        try:
            selector_element.click()
        except ElementClickInterceptedException:
            self.selenium_driver.execute_script("arguments[0].click();", selector_element)
    
        time.sleep(1.0)
    
        # Start timer for 120-second max timeout
        start_time = time.time()
    
        # 🔁 Retry up to 10 times
        for retry_count in range(10):
            try:
                # click Share button
                row_element.find_element(By.XPATH, ".//button[@data-automationid='shareHeroId']").click()
                time.sleep(1.0)
    
                # keyboard navigation to copy link
                ActionChains(self.selenium_driver).send_keys(
                    Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.ENTER
                ).perform()
                time.sleep(1.25)
                ActionChains(self.selenium_driver).send_keys(
                    Keys.TAB, Keys.ARROW_DOWN, Keys.TAB, Keys.TAB, Keys.ENTER
                ).perform()
                time.sleep(1.25)
                ActionChains(self.selenium_driver).send_keys(Keys.ENTER).perform()
                time.sleep(1.25)
                ActionChains(self.selenium_driver).send_keys(Keys.ESCAPE).perform()
    
                # unselect the row
                time.sleep(1.0)
                selector_element.click()
    
                # check clipboard
                encrypted_file_link = self.__get_clipboard_content__()
                if encrypted_file_link != starting_clipboard_content:
                    return encrypted_file_link  # ✅ SUCCESS → return link
    
                print(f"⚠️ Did not Successfully Gather link on attempt {retry_count + 1}. Retrying…")
    
            except Exception as e:
                print(f"⚠️ Attempt {retry_count + 1} failed: {e}")
                time.sleep(2.0)
    
            # check hard timeout
            if time.time() - start_time > 120:
                print("⏳ Timeout: Could not get link in 120 seconds. Moving on.")
                return None  # ❌ Fail after timeout
    
        print("❌ Failed to get SharePoint link after 10 retries.")
        return None  # ❌ Fail after 10 attempts
      
    def __get_clipboard_content__(self) -> str:
            """
            Local helper method used to pull clipboard content for generated links
            Returns the link generated by onedrive
            """
            
            # Pull the clipboard content and store it, then dump the link contents out of it
            for retry_count in range(10):
            
                # Open the clipboard and pull our file link   
                try:    
                    win32clipboard.OpenClipboard()
                    encrypted_file_link = win32clipboard.GetClipboardData()
                    win32clipboard.CloseClipboard()

                    # Return the link generated here
                    return encrypted_file_link

                # On failures, retry opening the clipboard if possible
                except:
                
                    # Check if we can retry or not
                    if retry_count == 10:
                        raise Exception("ERROR! Failed to open the clipboard!")
                
                    # Wait a moment before retrying to open the clipboard 
                    win32clipboard.CloseClipboard()
                    time.sleep(1.0)  
                        
    def is_broken_sharepoint_link(self, url: str, file_name: str = None) -> bool:
        try:
            original_tab = self.selenium_driver.current_window_handle
    
            # Open new tab
            self.selenium_driver.execute_script("window.open('');")
            WebDriverWait(self.selenium_driver, 5).until(
                lambda d: len(d.window_handles) > 1
            )
            new_tab = [h for h in self.selenium_driver.window_handles if h != original_tab][0]
            self.selenium_driver.switch_to.window(new_tab)
    
            # Load URL & wait for body
            self.selenium_driver.get(url)
            WebDriverWait(self.selenium_driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
    
            # ── NEW “Part”-filename check ───────────────────────────
            # First try the standard viewer span
            try:
                title_span = WebDriverWait(self.selenium_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "span.fui-Text"))
                )
                loaded_name = title_span.text or ""
                if "part" in loaded_name.lower():
                    print(f"ℹ️ Loaded filename contains 'Part': {loaded_name} → treating link as good.")
                    return False
            except:
                # Span didn’t appear — fall back to our heroField XPath
                xpath_fallback = (
                    "//span[@data-id='heroField'"
                    " and contains("
                      "translate(text(),"
                                 " 'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                                 " 'abcdefghijklmnopqrstuvwxyz'"
                      "),"
                      " 'part'"
                    ")]"
                )
                try:
                    WebDriverWait(self.selenium_driver, 3).until(
                        EC.presence_of_element_located((By.XPATH, xpath_fallback))
                    )
                    print(f"✅ XPath fallback found for Part file → link considered good.")
                    return False
                except:
                    # no part indicator, carry on to normal broken-link checks
                    pass
    
            # ── Your existing error checks below ────────────────────
    
            # Check for SharePoint error panel
            error_element = self.selenium_driver.find_elements(
                By.ID, "ctl00_PlaceHolderPageTitleInTitleArea_ErrorPageTitlePanel"
            )
            if error_element and "something went wrong" in error_element[0].text.lower():
                return True
    
            body_text = self.selenium_driver.find_element(By.TAG_NAME, "body").text.lower()
            if "sorry, something went wrong" in body_text:
                return True
    
            # Check that the PDF viewer loaded the filename span (again)
            try:
                WebDriverWait(self.selenium_driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "span.fui-Text"))
                )
            except:
                print("❌ File viewer didn't load filename — treating as broken.")
                return True
    
            return False
    
        except Exception as e:
            print(f"⚠️ Error checking link: {url} → {e}")
            return True
    
        finally:
            # Close only the new tab, then switch back
            if self.selenium_driver.current_window_handle != original_tab:
                if len(self.selenium_driver.window_handles) > 1:
                    self.selenium_driver.close()
                    self.selenium_driver.switch_to.window(original_tab)

    def __get_entry_heirarchy__(self, row_element: WebElement) -> str:
        # Find all breadcrumb elements
        title_elements = self.selenium_driver.find_elements(By.XPATH, self.__ONEDRIVE_PAGE_NAME_LOCATOR__)
        try:
            # Use a case-insensitive substring match rather than exact equality
            matching_element = next(title_element for title_element in title_elements
                                    if self.sharepoint_make.lower() in title_element.get_attribute("innerText").lower())
            title_index = title_elements.index(matching_element)
        except StopIteration:
            # If no match is found, fall back to the first element
            title_index = 0
    
        child_elements = title_elements[title_index:]
        entry_heirarchy = ""
        for child_element in child_elements:
            folder_name = child_element.get_attribute("innerText").strip()
            # (Keep your renaming logic here...)
            if folder_name == "RS3":
                folder_name = "RS 3"
            elif folder_name == "RS5":
                folder_name = "RS 5"
            elif folder_name == "RS6":
                folder_name = "RS 6"
            elif folder_name == "RS7":
                folder_name = "RS 7"
            #elif folder_name == "SQ3":
             #   folder_name = "SQ 3"
            #elif folder_name == "SQ5":
             #   folder_name = "SQ 5"
            #elif folder_name == "SQ7":
             #   folder_name = "SQ 7"
            #elif folder_name == "SQ8":
             #   folder_name = "SQ 8"
            elif folder_name == "VERANO":
                folder_name = "Verano"   
            elif folder_name == "Trailbalzer":
                folder_name = "Trailblazer"  
            elif folder_name == "Savanna":
                folder_name = "Savana"   
            elif folder_name == "Clarity":
                folder_name = "CLARITY ELECTRIC"  
            elif folder_name == "Clarity Plug In":
                folder_name = "CLARITY PLUG-IN"   
            elif folder_name == "EX35":
                folder_name = "EX"  
            elif folder_name == "G37 Convertible":
                folder_name = "G Convertible"  
            elif folder_name == "G37 Coupe":
                folder_name = "G Coupe"
            elif folder_name == "G37 Sedan":
                folder_name = "G Sedan"
            elif folder_name == "QX56":
                folder_name = "QX"   
            elif folder_name == "Grand Cherokee WL":          
                folder_name = "Grand Cherokee"                
            elif folder_name == "Wrangler (JL)":            
                folder_name = "Wrangler"   
            elif folder_name == "Wrangler JL":
                folder_name = "Wrangler"   
            elif folder_name == "K5 [Optima]":
                folder_name = "K5"
            elif folder_name == "K7 [Cadenza]":
                folder_name = "K7"   
            elif folder_name == "New Range Rover":
                folder_name = "Range Rover" 
            elif folder_name == "New Range Rover Evoque":
                folder_name = "Evoque" 
            elif folder_name == "New Range Rover Sport":
                folder_name = "Sport"    
            elif folder_name == "Range Rover Sport":
                folder_name = "Sport"     
            elif folder_name == "Range Rover Velar":
                folder_name = "Velar"     
            elif folder_name == "RCF":
                folder_name = "RC F"  
            elif folder_name == "CX3":
                folder_name = "CX-3" 
            elif folder_name == "CX30":
                folder_name = "CX-30" 
            elif folder_name == "CX5":
                folder_name = "CX-5" 
            elif folder_name == "CX50":
                folder_name = "CX-50"
            elif folder_name == "CX9":
                folder_name = "CX-9" 
            elif folder_name == "MX30":
                folder_name = "MX-30"   
            elif folder_name == "MX5":
                folder_name = "MX-5" 
            elif folder_name == "Mazda 2":
                folder_name = "2"  
            elif folder_name == "Mazda 3":
                folder_name = "3"   
            elif folder_name == "Mazda 5":
                folder_name = "5" 
            elif folder_name == "Mazda 6":
                folder_name = "6"  
            elif folder_name == "F54 Clubman":
                folder_name = "Clubman"  
            elif folder_name == "F55 Hardtop 4 Door":
                folder_name = "Hardtop 4D"
            elif folder_name == "F56 Hardtop 2 Door":
                folder_name = "Hardtop 2D" 
            elif folder_name == "F57 Convertible":
                folder_name = "Convertible"  
            elif folder_name == "F60 Countryman":
                folder_name = "Countryman" 
            elif folder_name == "Panamera 971":
                folder_name = "Panamera"
            elif folder_name == "Culinan":
                folder_name = "Cullinan"       
            elif folder_name == "RAV 4":
                folder_name = "RAV4"                 
            entry_heirarchy += folder_name + "\\"
    
        entry_heirarchy += self.__get_row_name__(row_element)
        return entry_heirarchy
     
    def __get_folder_rows__(self, row_link: str = None) -> tuple[list, list]:
        if row_link is not None:
            self.selenium_driver.get(row_link)
    
        indexed_files = []
        indexed_folders = []
    
        # Compile ADAS/Repair‐mode regex patterns
        if self.repair_mode and self.selected_adas:
            adas_patterns = [re.compile(re.escape(rs), re.IGNORECASE) for rs in self.selected_adas]
        elif self.selected_adas:
            adas_patterns = [
                re.compile(rf"\({re.escape(adas)}\s*\d*\)", re.IGNORECASE)
                for adas in self.selected_adas
            ]
        else:
            adas_patterns = None
    
        # ─── ROBUST WAIT FOR FOLDER ROWS ───
        # 1) wait for the table container to appear
        WebDriverWait(self.selenium_driver, self.__MAX_WAIT_TIME__).until(
            EC.presence_of_element_located((By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__))
        )
        # 2) wait for any loading spinner to vanish
        try:
            WebDriverWait(self.selenium_driver, 5).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, ".loading-spinner"))
            )
        except TimeoutException:
            pass
    
        # 3) grab all table containers on the page
        page_elements = self.selenium_driver.find_elements(By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__)
    
        for page_element in page_elements:
            # Poll the number of rows until it stabilizes
            prev_count = -1
            stable = 0
            rows = []
            while stable < 3:
                rows = page_element.find_elements(By.XPATH, self.__ONEDRIVE_TABLE_ROW_LOCATOR__)
                count = len(rows)
                if count == prev_count:
                    stable += 1
                else:
                    stable = 0
                prev_count = count
                time.sleep(0.5)
    
            if not rows:
                print("No table rows found in folder; skipping...")
                continue
    
            # Get the folder title from the page header
            page_title = (
                self.selenium_driver
                .find_elements(By.XPATH, self.__ONEDRIVE_PAGE_NAME_LOCATOR__)[-1]
                .get_attribute("innerText")
                .strip()
            )
    
            # Now iterate the fully−loaded rows
            for row_element in rows:
                entry_name = self.__get_row_name__(row_element)
                entry_hierarchy = self.__get_entry_heirarchy__(row_element)
    
                    # — SPECIAL: if Repair mode & SAS is selected, grab the SAS folder itself —
                if self.repair_mode and 'SAS' in [s.upper() for s in self.selected_adas] \
                   and entry_name.strip().upper() == 'SAS':
                    # row_link is the folder URL we came in on
                    folder_url = row_link or self.selenium_driver.current_url
                    indexed_files.append(
                        SharepointExtractor.SharepointEntry(
                            entry_name,
                            entry_hierarchy,
                            folder_url,
                            SharepointExtractor.EntryTypes.FILE_ENTRY
                        )
                    )
                    continue

                # Special handling for "No ..." entries
                if entry_name.lower().startswith("no"):
                    simulated_entry = self.__simulate_entry_from_no_entry__(
                        entry_name,
                        self.__get_encrypted_link__(row_element),   # Get real SharePoint link
                        self.__get_entry_heirarchy__(row_element),
                        indexed_files
                    )
                    if simulated_entry:
                        indexed_files.append(simulated_entry)
                    continue  # Do not add the original "No ..." item
                                
                # skip unwanted terms
                ignore_terms = ["old", "part", "replacement", "data", "statement", "stament"]
                if any(term in entry_name.lower() for term in ignore_terms):
                    continue
    
                if self.__is_row_folder__(row_element):
                    if page_title == self.sharepoint_make and not re.search(r"\d{4}", entry_name):
                        continue
    
                    folder_link = self.__get_unencrypted_link__(row_element)
                    # special deep-folder check
                    if re.search("|".join(self.__DEFINED_MODULE_NAMES__), entry_name):
                        self.selenium_driver.switch_to.new_window(WindowTypes.TAB)
                        self.selenium_driver.get(folder_link)
                        try:
                            sub_table = WebDriverWait(self.selenium_driver, 25).until(
                                EC.presence_of_element_located((By.XPATH, self.__ONEDRIVE_TABLE_LOCATOR__))
                            )
                            sub_rows = sub_table.find_elements(By.XPATH, self.__ONEDRIVE_TABLE_ROW_LOCATOR__)
                            sub_names = [self.__get_row_name__(r) for r in sub_rows]
                        except:
                            sub_names = []
                        self.selenium_driver.close()
                        self.selenium_driver.switch_to.window(self.selenium_driver.window_handles[0])
    
                        if any(re.search(r"([PpAaRrTt]{4})|(\d+\s*\.[^\s]+)", name) for name in sub_names):
                            folder_link = self.__get_encrypted_link__(row_element)
                            indexed_files.append(
                                SharepointExtractor.SharepointEntry(
                                    entry_name,
                                    entry_hierarchy,
                                    folder_link,
                                    SharepointExtractor.EntryTypes.FOLDER_ENTRTY
                                )
                            )
                            continue
    
                    indexed_folders.append(
                        SharepointExtractor.SharepointEntry(
                            entry_name,
                            entry_hierarchy,
                            folder_link,
                            SharepointExtractor.EntryTypes.FOLDER_ENTRTY
                        )
                    )
                    continue
    
                # === 🔍 FILTERING STARTS HERE ===
                if self.selected_adas:
                    if self.repair_mode:
                        module_matches = re.findall(r'\((.*?)\)', entry_name)
                        found_match = False
                    
                        if module_matches:
                            for module in module_matches:
                                module = module.strip().upper()
                                if module in [s.upper() for s in self.selected_adas]:
                                    found_match = True
                                    break
                        else:
                            name_without_ext = os.path.splitext(entry_name)[0]
                            last_word = name_without_ext.split()[-1].strip().upper()
                            if last_word in [s.upper() for s in self.selected_adas]:
                                found_match = True
                    
                        if not found_match:
                            print(f"Skipping {entry_name} — No matching system found in {self.selected_adas}")
                            continue

                    else:
                        if not any(p.search(entry_name) for p in adas_patterns):
                            continue
                # === 🔍 FILTERING ENDS HERE ===
    
                    # — SPECIAL: if file mentions MDPS, hyperlink the parent folder instead —
                if not self.__is_row_folder__(row_element) \
                   and any(phrase in entry_name.upper() for phrase in ['EXCEPT MDPS', 'MDPS ONLY']):
                    folder_url = row_link or self.selenium_driver.current_url
                    indexed_files.append(
                        SharepointExtractor.SharepointEntry(
                            entry_name,
                            entry_hierarchy,
                            folder_url,
                            SharepointExtractor.EntryTypes.FILE_ENTRY
                        )
                    )
                    continue


                file_link = self.__get_encrypted_link__(row_element)
                indexed_files.append(
                    SharepointExtractor.SharepointEntry(
                        entry_name,
                        entry_hierarchy,
                        file_link,
                        SharepointExtractor.EntryTypes.FILE_ENTRY
                    )
                )
    
        return [indexed_folders, indexed_files]
     
    def __update_excel_with_whitelist__(self, ws, entry_name, document_url):
        normalized_entry_name = entry_name.replace("(", "").replace(")", "").replace("-", "/").replace("[", "").replace("]", "").replace("WL", "").replace("Multipurpose", "Multipurpose Camera").replace("-PL-PW072NLB", " Side Blind Zone Alert").replace("forward Collision Warning/Lane Departure Warning (FCW/LDW)", "FCW/LDW").strip().upper()
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            cell_value = str(row[0].value).strip().upper()
    
            # Restrict based on selected ADAS acronyms
            if self.selected_adas and cell_value not in self.selected_adas:
                continue
    
            if cell_value in self.__ADAS_SYSTEMS_WHITELIST__:
                if cell_value in normalized_entry_name:
                    cell = ws.cell(row=row[0].row, column=self.HYPERLINK_COLUMN_INDEX)
                    cell.hyperlink = document_url
                    cell.value = document_url
                    cell.font = Font(color="0000FF", underline='single')
                    # Color logic
                    force_red = bool(getattr(self, "debug_mode", False)) and bool(getattr(self, "write_in_debug", True))
                    if getattr(self, "_last_match_approx", False):
                        cell.font = Font(color="FF0000", underline='single')   # red hyperlink for regex/fuzzy placement
                    else:
                        cell.font = Font(color="0000FF", underline='single')   # blue for perfect match
                    print(f"Hyperlink for {entry_name} added at {cell.coordinate}")
                    return True
        return False
 

    # ── Header utils (header-only; no fallbacks) ─────────────────────────────────
    def _norm_hdr(self, s: str) -> str:
        # Normalize header text: trim, uppercase, collapse inner spaces
        import re
        return re.sub(r"\s+", " ", str(s).strip().upper())

    # ── Header-only helpers ─────────────────────────────────────────────
    def _header_colmap_(self, ws):
        """Return 1-based indices for required headers; no fallbacks."""
        import re
        def norm(s): return re.sub(r"\s+", " ", str(s).strip().upper())
    
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = {}
        for i, c in enumerate(header_cells):
            if c.value is None: 
                continue
            headers[norm(c.value)] = i + 1  # 1-based
    
        def pick(*names):
            for n in names:
                key = norm(n)
                if key in headers:
                    return headers[key]
            return None
    
        colmap = {
            "year":   pick("Year"),
            "make":   pick("Make"),
            "model":  pick("Model"),
            "system": pick(  # ADAS + Repair
                "SME Generic System Name",
                "Protech Generic System Name",
                "Generic System Name",
                "System Name",
                "System",
            ),
            "hyperlink": pick(  # includes “Service Information”
                "Hyperlink", "Link", "URL",
                "Service Information", "Service Information Hyperlink",
                "SI", "SI Link", "SI URL",
            ),
        }
        missing = [k for k in ("year","make","model","system") if not colmap.get(k)]
        if missing:
            raise ValueError(f"Missing required header(s): {', '.join(missing)}")
        return colmap
    
    def _cell_val_upper(self, row_tuple, one_based_idx):
        if not one_based_idx:
            return ""
        i = one_based_idx - 1
        if i < 0 or i >= len(row_tuple):
            return ""
        v = row_tuple[i].value
        return (str(v).strip().upper() if v is not None else "")

    # ★ Add this helper once inside your class (above the two methods below)
    def _system_val_for_row(self, row, repair_mode: bool):
        """
        Return (system_text, system_norm_for_index) for a given openpyxl 'row' (tuple of cells).
        Uses the correct columns for OG vs NEW and Repair vs ADAS.
        NOTE: 'row[i]' here is 0-based indexing (row[0] == Column A).
        """
        if repair_mode:
            # Repair SI
            if self.excel_mode == "new":
                # NEW Repair: Column T (0-based 19)
                sys_cell = row[19] if len(row) > 19 and row[19].value else None
            elif str(self.sharepoint_make).lower() == "toyota":
                # OG Repair for Toyota: Column E (0-based 4)
                sys_cell = row[4] if len(row) > 4 and row[4].value else None
            else:
                # OG Repair default: Column D (0-based 3)
                sys_cell = row[3] if len(row) > 3 and row[3].value else None
        else:
            # ADAS SI
            if self.excel_mode == "new":
                # NEW ADAS: Column U (0-based 20) after S→U move
                sys_cell = row[20] if len(row) > 20 and row[20].value else None
            else:
                # OG ADAS: Column E (0-based 4)
                sys_cell = row[18] if len(row) > 18 and row[18].value else None
    
        sys_text = (str(sys_cell.value).strip().upper() if sys_cell else "")
        sys_norm = re.sub(r"[^A-Z0-9]", "", sys_text)  # EXACT match with your __build_row_index__
        return sys_text, sys_norm
    
    # ★ REPLACE your __update_excel__ with this (adds header-only hyperlink column ensure,
    #   overwrites OG 'Placeholder', keeps acronym verifier & NO-doc color logic)
    def __update_excel__(self, ws, year, model, doc_name, document_url, adas_last_row, cell_address=None):
        # Skip filtering if in Repair mode
        if not self.repair_mode:
            if self.selected_adas and not any(adas in doc_name.upper() for adas in self.selected_adas):
                return
    
        # Ensure we have the correct hyperlink column by HEADER ONLY
        try:
            if not getattr(self, "HYPERLINK_COLUMN_INDEX", None):
                # accept both names; create "Service Information" if none exist
                self._ensure_hyperlink_column(ws, "Service Information", "Service Information Hyperlink")
        except Exception:
            # best effort: still try to proceed — downstream write will raise if missing
            pass
    
        # Try to find the correct Excel row for this system
        if doc_name in self.SPECIFIC_HYPERLINKS:
            cell = ws[self.SPECIFIC_HYPERLINKS[doc_name]]
            error_message = None
            # ensure exact (no red) when SPECIFIC_HYPERLINKS used
            self._last_match_approx = False
        else:
            cell, error_message = self.__find_row_in_excel__(
                ws, year, self.sharepoint_make, model, doc_name,
                repair_mode=self.repair_mode, row_index=getattr(self, "row_index", None)
            )
    
        # --- VERIFY the picked row actually matches (Y, M, Model, System). If not, fix it. ---
        try:
            Y  = (year or "").strip().upper()
            M  = (self.sharepoint_make or "").strip().upper()
            MR = (model or "").strip().upper()
            sys_raw     = _extract_system_from_filename(doc_name)
            sys_norm_ix = re.sub(r"[^A-Z0-9]", "", (sys_raw or "").upper())
    
            # Ensure we have the latest index
            self.row_index = getattr(self, "row_index", None) or self.__build_row_index__(ws, repair_mode=self.repair_mode)
    
            exact_key = (Y, M, MR, sys_norm_ix)
            expected_row = self.row_index.get(exact_key)
    
            if expected_row and cell and cell.row != expected_row:
                print(f"🔁 Row verifier: correcting from row {cell.row} to expected row {expected_row} for {doc_name}")
                cell = ws.cell(row=expected_row, column=self.HYPERLINK_COLUMN_INDEX)
                self._last_match_approx = False
            elif not expected_row and cell is None:
                # nothing indexed for this exact key and no cell chosen → will create new row below
                pass
            elif not expected_row:
                # we matched via regex/fuzzy to some row; mark approx so it goes red
                self._last_match_approx = True
        except Exception as _e:
            print(f"⚠️ Row verifier error for {doc_name}: {_e}")
    
        # --- ★ ADAS acronym verifier: retarget to the correct per-system row to avoid overwrites ---
        try:
            if not self.repair_mode and cell:
                # Build the target acronym from the filename’s system
                sys_raw2 = _extract_system_from_filename(doc_name)
                if sys_raw2:
                    sn_index = re.sub(r"[^A-Z0-9]", "", (sys_raw2 or "").upper())   # e.g., 'ACC (4)' → 'ACC4'
                    sn_letters = re.sub(r"[^A-Z]", "", sn_index)                    # 'ACC4' → 'ACC'
    
                    name_idx = _adas_name_col_index(self.repair_mode, getattr(self, "excel_mode", "og"))
                    ok = False
    
                    if name_idx is not None:
                        # Does the current row already have the right acronym?
                        try:
                            # openpyxl is 1-based; name_idx is 0-based
                            cur_name_cell = ws.cell(row=cell.row, column=name_idx + 1)
                            cur_letters, cur_alnum = _adas_name_norms(cur_name_cell.value if cur_name_cell else "")
                            if (cur_alnum == sn_index) or (cur_letters == sn_letters):
                                ok = True
                        except Exception:
                            ok = False
    
                        if not ok:
                            # Search among same Year/Make/Model for the row whose 'Name' matches the target acronym
                            Y2  = (year or "").strip().upper()
                            M2  = (self.sharepoint_make or "").strip().upper()
                            MR2 = (model or "").strip().upper()
    
                            for row in ws.iter_rows(min_row=2, max_col=max(22, name_idx + 1)):
                                if not any(c.value for c in row):
                                    continue
                                yr  = (str(row[0].value) or "").strip().upper()
                                mk  = (str(row[1].value) or "").strip().upper()
                                mdl = (str(row[2].value) or "").strip().upper()
                                if yr != Y2 or mk != M2 or mdl != MR2:
                                    continue
    
                                name_cell2 = row[name_idx] if len(row) > name_idx else None  # 0-based index
                                letters2, alnum2 = _adas_name_norms(name_cell2.value if name_cell2 else "")
                                if (alnum2 == sn_index) or (letters2 == sn_letters):
                                    print(f"🔁 Acronym verifier: retarget {doc_name} from row {cell.row} → {row[0].row}")
                                    cell = ws.cell(row=row[0].row, column=self.HYPERLINK_COLUMN_INDEX)
                                    break
        except Exception as _e:
            print(f"⚠️ Acronym verifier error for {doc_name}: {_e}")
        # --- ★ END acronym verifier ---
    
        # Create a unique key ...
        if self.repair_mode:
            module_matches = re.findall(r'\((.*?)\)', doc_name)
            system_name = None
            for mod in module_matches:
                if self.selected_adas and mod.strip().upper() in [s.upper() for s in self.selected_adas]:
                    system_name = mod.strip().upper()
                    break
            if not system_name:
                if module_matches:
                    system_name = module_matches[0].strip().upper()
                else:
                    system_name = os.path.splitext(doc_name)[0].split()[-1].strip().upper()
            key = (year, self.sharepoint_make, model, system_name)
        else:
            key = (year, self.sharepoint_make, model, doc_name)
    
        # If we didn’t find a matching cell, create one at the bottom
        if not cell:
            if cell_address:
                cell = ws[cell_address]
            else:
                row = ws.max_row + 1
                if key in adas_last_row:
                    row = adas_last_row[key] + 1
                else:
                    adas_last_row[key] = row
                cell = ws.cell(row=row, column=self.HYPERLINK_COLUMN_INDEX)
    
            # force "approx" on fallback placements so they go red
            self._last_match_approx = True
    
            # Place RED NAME text in the proper error column
            if self.repair_mode:
                error_column = 7    # G
            elif self.excel_mode == "new":
                error_column = 10   # J
            else:
                error_column = 11   # K
            error_cell = ws.cell(row=cell.row, column=error_column)
            error_cell.value = doc_name.splitlines()[0]
            error_cell.font = Font(color="FF0000")
    
        # ─────────────────────────────────────────────────────────────────────────
        # ✅ Always set visible text + unified color/underline rules
        # Precedence: NO-doc (yellow) > approx/debug (red) > exact (blue)
        # Also: overwrite OG 'Placeholder' or empty text with friendly link text
        # ─────────────────────────────────────────────────────────────────────────
        if document_url:
            # Friendly display text (used only in debug mode)
            def _mk_link_text(y, mk, mdl, dn):
                try:
                    # keep your preferred phrasing; avoids double-parens if dn has them
                    return f"Link For: {str(y).strip()} {str(mk).title().strip()} {str(mdl).title().strip()} ({dn}).pdf"
                except Exception:
                    return f"Link For: {dn}.pdf"
    
            # Neutralize Excel default Hyperlink style first
            try:
                cell.style = "Normal"
            except Exception:
                pass
    
            # Always write the hyperlink target
            cell.hyperlink = document_url
    
            approx = bool(getattr(self, "_last_match_approx", False))
            debug_writing = bool(getattr(self, "debug_mode", False)) and bool(getattr(self, "write_in_debug", True))
    
            # Detect NO-doc: combine filename check and the flag your other code sets
            try:
                is_no_doc_name = doc_name.strip().lower().startswith("no ")
            except Exception:
                is_no_doc_name = False
            is_no_doc_flag = bool(getattr(self, "_last_is_no_doc", False))
            is_no_doc = is_no_doc_flag or is_no_doc_name
    
            import os
            base_name = os.path.splitext(doc_name)[0].strip() if doc_name else ""
    
            # Decide what TEXT to show in the hyperlink cell
            # ------------------------------------------------
            #  🔹 DEBUG OFF  → always show the URL itself
            #  🔹 DEBUG ON   → show pretty "Link For: ..." text (or filename)
            # ------------------------------------------------
            if not debug_writing:
                # DEBUG MODE OFF → ALWAYS URL TEXT
                cell.value = document_url
            else:
                # DEBUG MODE ON → keep your previous smart behavior
                cur_text = (str(cell.value).strip() if cell.value is not None else "")
                if (not cur_text) or (cur_text.lower() == "placeholder") or cur_text.lower().startswith("link for:") or cur_text.lower().startswith("http"):
                    cell.value = _mk_link_text(
                        year,
                        self.sharepoint_make,
                        model,
                        _extract_system_from_filename(doc_name) or doc_name
                    )
                else:
                    # If you want to always enforce the friendly text, uncomment:
                    # cell.value = _mk_link_text(year, self.sharepoint_make, model, _extract_system_from_filename(doc_name) or doc_name)
                    pass
    
            # 🔶 NO-doc (yellow) > approx/debug (red) > exact (blue)
            if is_no_doc:
                cell.font = Font(color="9B870C", underline='single')   # dark yellow
            elif approx or debug_writing:
                cell.font = Font(color="FF0000", underline='single')   # red
            else:
                cell.font = Font(color="0000FF", underline='single')   # blue
    
            # ────────────────────────────────────────────────────────────────
            # ★ LEFT PLACEHOLDER: place filename in the cell LEFT of the
            #   hyperlink if that cell is empty or 'Placeholder'
            #   (left cell only; does NOT touch hyperlink text)
            # ────────────────────────────────────────────────────────────────
            try:
                hyperlink_col = getattr(self, "HYPERLINK_COLUMN_INDEX", None) or cell.column
                left_col = hyperlink_col - 1
                if left_col >= 1:
                    left_cell = ws.cell(row=cell.row, column=left_col)
                    left_val = (str(left_cell.value).strip().lower() if left_cell.value else "")
                    if left_val == "" or left_val == "placeholder":
                        if base_name:
                            left_cell.value = base_name
                            # match your error-style red for placeholders on the left
                            left_cell.font = Font(color="FF0000")
            except Exception as _e:
                print(f"⚠️ Placeholder-left error for {doc_name}: {_e}")
    
        else:
            cell.hyperlink = None
            cell.value = f"{doc_name} "
            cell.font = Font(color="FF0000")
    
            if not hasattr(self, "mismatched_files"):
                self.mismatched_files = []
            self.mismatched_files.append(doc_name)
            print(f"⚠️ No hyperlink for {doc_name} → adding to proper location as placeholder")
    
        adas_last_row[key] = cell.row
        print(f"Hyperlink for {doc_name} added at {cell.coordinate} "
              f"[{'approx' if getattr(self, '_last_match_approx', False) else 'exact'}]")
    
        # reset the NO-doc flag so it doesn’t bleed into the next write
        self._last_is_no_doc = False





    

    def __find_row_in_excel__(self, ws, year, make, model, file_name, repair_mode=False, row_index=None):
        """
        Strict on:  Year + Make + System
        Model:      exact(raw) → regex(raw) → fuzzy(raw)
        System:     if exact (with digits) fails, try a letters-only fallback
        Flags:      self._last_match_approx = True on any relaxed (system/model) match
        """
        self._last_match_approx = False
    
        if row_index is None:
            try:
                row_index = self.__build_row_index__(ws, repair_mode=repair_mode)
            except TypeError:
                row_index = self.__build_row_index__(ws)
    
        Y  = (year  or '').strip().upper()
        M  = (make  or '').strip().upper()
        MR = (model or '').strip().upper()  # RAW upper (your index uses this)
    
        # If the SharePoint filename says "4C Coupe", pretend the model is just "4C"
        if re.search(r'\b4C\s*COUPE\b', (file_name or '').upper()):
            MR = "4C"
        # If the SharePoint filename shows "CR-Z", simulate Excel model "CR-Z [HEV]"
        elif re.search(r'\bCR[-\s]?Z\b', (file_name or '').upper()) and "HEV" not in MR:
            MR = "CR-Z [HEV]"
    
        # ★ NEW: Jeep Grand Cherokee body code shim — pick up [WK]/[WL] even if it appears after the system
        # ★ Jeep Grand Cherokee body code shim (standalone token or in []/() anywhere in filename)
        # Actual:   2022 Jeep Grand Cherokee (ACC 2) WK
        # Simulate: 2022 Jeep Grand Cherokee WK (ACC 2)
        m = re.search(r'(?:\(|\[)?\s*(WK|WL)\s*(?:\)|\])?', (file_name or '').upper())
        if m:
            code = m.group(1)  # 'WK' or 'WL'
            # ensure MR ends with the body code as a plain suffix (no brackets)
            MR_core = re.sub(r'\s+\b(?:WK|WL)\b\s*$', '', MR)   # drop any existing WK/WL suffix
            MR = f"{MR_core} {code}".strip()
        
        sys_raw = _extract_system_from_filename(file_name)
        if not (Y and M and sys_raw):
            return None, file_name
    
        # system keys
        SN_index = _norm_system_index(sys_raw)  # keep digits; e.g., "APA 2" -> "APA2"
        SN_loose = _norm_system_loose(sys_raw)  # letters only; e.g., "APA2" -> "APA"
    
        # 1) EXACT by raw model + index-style system
        key = (Y, M, MR, SN_index)
        if key in row_index:
            return ws.cell(row=row_index[key], column=self.HYPERLINK_COLUMN_INDEX), None
    
        # 1b) EXACT with "loose" system (rare sheets that store letters-only)
        if SN_loose and SN_loose != SN_index:
            key_loose = (Y, M, MR, SN_loose)
            if key_loose in row_index:
                self._last_match_approx = True
                return ws.cell(row=row_index[key_loose], column=self.HYPERLINK_COLUMN_INDEX), None
    
        # ★ Guard rails (only-exact allowed for these)
        if _is_force_bottom_model(MR):          # Silverado 4500/5500/6500(+HD), F-450/F-550, E-450, etc.
            return None, file_name
        if _is_force_bottom_combo(Y, M, MR):    # your force-bottom combos (Evoque, RX450h, etc.)
            return None, file_name
    
        # 2) REGEX on RAW model (strict Y/M/System in {SN_index, SN_loose})
        rgx = _model_regex_from_excel(model)
        for (yr, mk, mdl_raw, sys_norm), r in row_index.items():
            if yr == Y and mk == M and sys_norm in (SN_index, SN_loose):
                # block NX300h ↔ ES300 style cross-family hops
                if _cross_family_conflict(mdl_raw, MR):
                    continue
                if rgx.search(_strip_qualifiers(mdl_raw)):
                    self._last_match_approx = True
                    return ws.cell(row=r, column=self.HYPERLINK_COLUMN_INDEX), None
    
        # 3) FUZZY on RAW model (strict Y/M/System in {SN_index, SN_loose})
        best_row, best_score = None, 0.0
        for (yr, mk, mdl_raw, sys_norm), r in row_index.items():
            if yr == Y and mk == M and sys_norm in (SN_index, SN_loose):
                # block cross-family hops
                if _cross_family_conflict(mdl_raw, MR):
                    continue
                sc = _similar(mdl_raw, MR)
                if sc > best_score:
                    best_score, best_row = sc, r
        if best_row and best_score >= 0.72:
            self._last_match_approx = True
            return ws.cell(row=best_row, column=self.HYPERLINK_COLUMN_INDEX), None

    
        # ======================= ★ Model-first rescue when sheet System is missing =======================

        def _system_missing(txt: str) -> bool:
            return _system_missing_text(txt)
        
        # Pre-compute normalized targets from the filename system
        sn_letters = re.sub(r"[^A-Z]", "", SN_index)   # e.g., 'ACC2' -> 'ACC'
        sn_alnum   = SN_index                          # e.g., 'ACC2'
        
        name_col_idx = _adas_name_col_index(self.repair_mode, getattr(self, "excel_mode", "og"))
        
        candidate_by_model = []
        for row in ws.iter_rows(min_row=2, max_col=22):
            if not any(c.value for c in row):
                continue
        
            yr  = (str(row[0].value) or "").strip().upper()
            mk  = (str(row[1].value) or "").strip().upper()
            mdl = (str(row[2].value) or "").strip().upper()
            if yr != Y or mk != M or mdl != MR:
                continue
        
            # pick the same System column your index uses
            if self.repair_mode:
                if self.excel_mode == "new":
                    sys_cell = row[19] if len(row) > 19 else None   # T
                elif str(self.sharepoint_make).lower() == "toyota":
                    sys_cell = row[4]  if len(row) > 4  else None   # E
                else:
                    sys_cell = row[3]  if len(row) > 3  else None   # D
            else:
                if self.excel_mode == "new":
                    sys_cell = row[20] if len(row) > 20 else None   # U
                else:
                    sys_cell = row[18]  if len(row) > 18  else None   # S
        
            sys_txt = str(sys_cell.value).strip().upper() if (sys_cell and sys_cell.value) else ""
            if not _system_missing(sys_txt):
                # If the System cell isn't missing, earlier logic should have matched; skip.
                continue
        
            # ★ Correct 'Name' column (Column S) for OG/NEW ADAS
            name_txt = ""
            if name_col_idx is not None and len(row) > name_col_idx and row[name_col_idx].value:
                name_txt = str(row[name_col_idx].value).strip().upper()
        
            name_letters, name_alnum = _adas_name_norms(name_txt)
        
            # Prefer rows whose 'Name' matches the file system
            # Examples:
            #   file 'ACC 2' → SN_index='ACC2' → match name_alnum == 'ACC2'
            #   file 'ACC'   → match name_letters == 'ACC'
            if name_alnum and name_alnum == sn_alnum:
                self._last_match_approx = True
                return ws.cell(row=row[0].row, column=self.HYPERLINK_COLUMN_INDEX), None
            if name_letters and name_letters == sn_letters:
                self._last_match_approx = True
                return ws.cell(row=row[0].row, column=self.HYPERLINK_COLUMN_INDEX), None
        
            # keep as fallback candidate (exact model, missing system but no acronym match)
            candidate_by_model.append(row[0].row)
        
        # If we found any exact-model candidates with missing System but no acronym match, use the first.
        if candidate_by_model:
            self._last_match_approx = True
            return ws.cell(row=candidate_by_model[0], column=self.HYPERLINK_COLUMN_INDEX), None
        # ===================== ★ END model-first rescue =======================
    
        # === letters-only SYSTEM fallback across all variants (APA1/APA2 → APA) ===
        if SN_loose:
            exact_row = None
            regex_row = None
            best_row2, best_score2 = None, 0.0
    
            for (yr, mk, mdl_raw, sys_norm), r in row_index.items():
                if yr != Y or mk != M:
                    continue
                # collapse the indexed system to letters-only and compare
                if _norm_system_loose(sys_norm) != SN_loose:
                    continue
    
                # ★ NEW: block cross-family hops
                if _cross_family_conflict(mdl_raw, MR):
                    continue
    
                # model exact?
                if mdl_raw == MR:
                    exact_row = r
                    break
    
                # model regex?
                if rgx.search(_strip_qualifiers(mdl_raw)):
                    if regex_row is None:
                        regex_row = r
                    continue
    
                # model fuzzy
                sc = _similar(mdl_raw, MR)
                if sc > best_score2:
                    best_score2, best_row2 = sc, r
    
            if exact_row is not None:
                self._last_match_approx = True   # relaxed by system
                return ws.cell(row=exact_row, column=self.HYPERLINK_COLUMN_INDEX), None
            if regex_row is not None:
                self._last_match_approx = True
                return ws.cell(row=regex_row, column=self.HYPERLINK_COLUMN_INDEX), None
            if best_row2 and best_score2 >= 0.72:
                self._last_match_approx = True
                return ws.cell(row=best_row2, column=self.HYPERLINK_COLUMN_INDEX), None
    
        # nothing found
        return None, file_name
    
    def _ensure_hyperlink_column(self, ws, *preferred_headers):
        """
        Ensure a hyperlink column exists by header only (no fallbacks).
        Tries all provided header names (case/space-insensitive). If none found,
        creates a new column using the FIRST preferred header name.
        Usage:
            self._ensure_hyperlink_column(ws, "Service Information", "Service Information Hyperlink")
        """
        import re
    
        def norm(s: str) -> str:
            return re.sub(r"\s+", " ", str(s).strip().upper())
    
        # Default preferred names if caller didn't pass any
        if not preferred_headers:
            preferred_headers = (
                "Service Information",
                "Service Information Hyperlink",
                "Service Information (URL)",
                "Hyperlink",
                "Link",
                "URL",
                "SI",
                "SI Link",
                "SI URL",
            )
    
        # Make sure we have a fresh header map
        self.colmap = getattr(self, "colmap", None) or self._header_colmap(ws)
    
        # 1) Try to find any of the preferred headers that already exist
        for name in preferred_headers:
            col = self.colmap.get("hyperlink")
            # If _header_colmap already recognized a generic "hyperlink" and it's present, use it
            # Otherwise, explicitly look up the exact normalized name in row 1:
            if col:
                self.HYPERLINK_COLUMN_INDEX = col
                return col
    
            # direct scan for this specific header text in row 1
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            for i, cell in enumerate(header_row, start=1):
                if cell.value and norm(cell.value) == norm(name):
                    self.colmap["hyperlink"] = i
                    self.HYPERLINK_COLUMN_INDEX = i
                    return i
    
        # 2) None found — create the FIRST preferred header
        create_name = preferred_headers[0]
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = create_name
        self.colmap["hyperlink"] = new_col
        self.HYPERLINK_COLUMN_INDEX = new_col
        return new_col

    

    def __build_row_index__(self, ws, repair_mode=False):
        """
        Header-only row index:
          key = (YEAR, MAKE, MODEL, SYSTEM_NORM) -> row_number
        Uses self.colmap; skips placeholder rows; no OG/NEW fallbacks.
        """
        # Ensure header map exists
        colmap = getattr(self, "colmap", None)
        if not colmap:
            colmap = self._header_colmap(ws)
            self.colmap = colmap
    
        Yc, Mc, Mdc, Sc = colmap["year"], colmap["make"], colmap["model"], colmap["system"]
        Hc = colmap.get("hyperlink")  # may be None
    
        index = {}
    
        # 🧹 CLEANUP MODE
        if getattr(self, "cleanup_mode", False):
            if not Hc:
                return index  # nothing to do if hyperlink col missing
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=Hc)
                val = str(cell.value).strip().lower() if cell.value else ""
                # skip placeholders or blank lines
                if not val or val == "placeholder":
                    continue
                looks_link = bool(cell.hyperlink) or val.startswith("http")
                if not looks_link:
                    continue
    
                row_cells = tuple(ws.iter_rows(min_row=r, max_row=r, max_col=ws.max_column))[0]
                year  = self._cell_val_upper(row_cells, Yc)
                make  = self._cell_val_upper(row_cells, Mc)
                model = self._cell_val_upper(row_cells, Mdc)
    
                sys_text = self._cell_val_upper(row_cells, Sc)
                import re
                system_norm = re.sub(r"[^A-Z0-9]", "", sys_text)
    
                key = (year, make, model, system_norm)
                index[key] = r
            return index
    
        # 🔹 NORMAL MODE
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            if not any(c.value for c in row):
                continue
    
            # skip placeholder rows if hyperlink col present
            if Hc:
                hv = str(row[Hc - 1].value).strip().lower() if row[Hc - 1].value else ""
                if hv == "placeholder":
                    continue
    
            year  = self._cell_val_upper(row, Yc)
            make  = self._cell_val_upper(row, Mc)
            model = self._cell_val_upper(row, Mdc)
    
            sys_text = self._cell_val_upper(row, Sc)
            if _system_missing_text(sys_text):
                continue
    
            import re
            system_norm = re.sub(r"[^A-Z0-9]", "", sys_text)
    
            key = (year, make, model, system_norm)
            index[key] = row[0].row
    
        return index

        
    
    

#####################################################################################################################################################

if __name__ == '__main__':   

   # (Individual File testing without GUI, take away the # to perform whichever is needed)) 
   # excel_file_path = r'C:\Users\dromero3\Desktop\Excel Documents\Toyota Pre-Qual Long Sheet v6.3.xlsx'
   # sharepoint_link = 'https://calibercollision.sharepoint.com/:f:/g/enterpriseprojects/VehicleServiceInformation/EiB53aPXartJhkxyWzL5AFABZQsY3x-XDWPXQCqgFIrvoQ?e=m4DrKQ'
   # debug_run = True


    sharepoint_link = sys.argv[1]
    excel_file_path = sys.argv[2]
    debug_run = False
    
    extractor = SharepointExtractor(sharepoint_link, excel_file_path, debug_run)

    print("=" * 68)

    if extractor.cleanup_mode:
        # Clean up mode: find broken links → re-index only those
        extractor.populate_excel_file([])

        if extractor.broken_entries:
            print("🔁 Re-indexing SharePoint to replace broken links...")
            _, filtered_files = extractor.extract_contents()
            print(f"📥 Matched {len(filtered_files)} files for repair.")
            extractor.populate_excel_file(filtered_files)

    else:
        # Normal mode: index entire folder and populate Excel
        folders, files = extractor.extract_contents()
        extractor.populate_excel_file(files)

    print("=" * 68)
    print(f"Extraction and population for {extractor.sharepoint_make} is complete!")
