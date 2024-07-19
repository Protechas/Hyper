import time
import psutil
import os
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common import actions
from selenium.webdriver.common import action_chains
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.webdriver.chrome.options import Options
import re
import win32clipboard
import openpyxl



def process_subdocuments(driver, wait, ws, subdocuments, year, model, adas_last_row, parent_xpath):
    for sub_doc_name, sub_doc_info in subdocuments.items():
        if isinstance(sub_doc_info, dict) and 'folder_xpath' in sub_doc_info:
            print(f"Accessing subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments'], year, model, adas_last_row, sub_doc_info['folder_xpath'])
            navigate_back_to_element(driver, wait, parent_xpath)
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)
            time.sleep(2) # Go back one more time to ensure stability
        elif isinstance(sub_doc_info, dict) and 'folder2_xpath' in sub_doc_info:
            print(f"Accessing nested subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder2_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder2_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments2'], year, model, adas_last_row, sub_doc_info['folder2_xpath'])
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)  # Go back one more time to ensure stability
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)
        else:
            print(f"Retrieving sub-document: {sub_doc_name}")
            document_url = navigate_and_extract(driver, wait, sub_doc_info['xpath'])
            time.sleep(2)
            update_excel(ws, year, model, sub_doc_name, document_url, adas_last_row, sub_doc_info.get('cell_address'))
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath) # Go back one more time to ensure stability

def process_documents(driver, wait, ws, model_data, year, model, adas_last_row):
    for doc_name, doc_info in model_data['documents'].items():
        if isinstance(doc_info, dict) and 'folder_xpath' in doc_info:
            print(f"Accessing folder: {doc_name}")
            double_click_element(driver, wait, doc_info['folder_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, doc_info['folder_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, doc_info['subdocuments'], year, model, adas_last_row, doc_info['folder_xpath'])
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])
            time.sleep(2)
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])  # Go back one more time to ensure stability
            time.sleep(2)
        else:
            print(f"Retrieving document: {doc_name}")
            document_url = navigate_and_extract(driver, wait, doc_info)
            time.sleep(2)
            update_excel(ws, year, model, doc_name, document_url, adas_last_row)
            time.sleep(2)
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])

def navigate_and_extract(driver, wait, xpath):
    
    win32clipboard.OpenClipboard()
    encrypted_file_link = win32clipboard.GetClipboardData()
    win32clipboard.CloseClipboard()

    # Double-click the element to open its context menu or relevant options
    double_click_element(driver, wait, xpath)
    
    # Wait until the share button is visible and click it
    share_button_xpath = ".//button[@data-automationid='shareHeroId']"
    
    # Scroll the share button into view
    share_button = wait.until(EC.visibility_of_element_located((By.XPATH, share_button_xpath)))
    driver.execute_script("arguments[0].scrollIntoView(true);", share_button)
    time.sleep(1)  # Ensure the scroll action is complete

    # Wait until the element is clickable and click it
    wait.until(EC.element_to_be_clickable((By.XPATH, share_button_xpath))).click()
    
    # Add a series of actions to navigate through the share menu and copy the link to the clipboard
    time.sleep(0.75)
    ActionChains(driver).send_keys(Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.ENTER).perform()
    time.sleep(0.50)
    ActionChains(driver).send_keys(Keys.ARROW_DOWN, Keys.TAB, Keys.TAB, Keys.ENTER).perform()
    time.sleep(0.50)
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    time.sleep(0.75)
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()

    # Extract the URL from the clipboard
    document_url = driver.execute_script("return navigator.clipboard.readText();")
    
    # Navigate back to the previous element
    driver.back()
    time.sleep(2)
    # Improved wait mechanism
    for _ in range(10):  # Retry up to 10 times
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
            break
        except TimeoutException:
            time.sleep(1)  # Wait a bit longer if necessary

    time.sleep(2)  # Allow extra time to ensure full navigation back

    # Check if the specified XPath is present, if not, navigate forward
    
    return document_url

def navigate_back_to_element(driver, wait, xpath):
    driver.back()
    time.sleep(3)  # Ensure the page fully loads

def check_and_navigate_forward(driver, wait, check_xpath):
    try:
        wait = WebDriverWait(driver, 5)
        wait.until(EC.visibility_of_element_located((By.XPATH, check_xpath)))
    except TimeoutException:
        driver.forward()
        time.sleep(2)  # Wait to ensure forward navigation completes
        print("Navigated forward as the specified XPath was not found")

def update_excel(ws, year, model, doc_name, document_url, adas_last_row, cell_address=None):
    if cell_address:
        cell = ws[cell_address]
    else:
        cell = find_row_in_excel(ws, year, "Acura", model, doc_name)
        if cell:
            row = cell.row
        else:
            row = ws.max_row + 1
        if doc_name in adas_last_row:
            row = adas_last_row[doc_name] + 1
        else:
            adas_last_row[doc_name] = row
        cell = ws.cell(row=row, column=12)

    cell.hyperlink = document_url
    cell.value = document_url
    cell.font = Font(color="0000FF", underline='single')
    adas_last_row[doc_name] = cell.row
    ws.parent.save(excel_file_path)
    print(f"Hyperlink for {doc_name} added at {cell.coordinate}")

def find_row_in_excel(ws, year, make, model, adas_system):
    for row in ws.iter_rows(min_row=2, max_col=8):
        year_cell, make_cell, model_cell, adas_cell = row[0], row[1], row[2], row[7]
        if (str(year_cell.value).strip() == str(year).strip() and
            str(make_cell.value).strip().lower() == make.lower().strip() and
            str(model_cell.value).strip().lower() == model.lower().strip() and
            adas_system.lower().strip() in str(adas_cell.value).lower().strip()):
            return ws.cell(row=year_cell.row, column=12)
    return None

def double_click_element(driver, wait, xpath):
    element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    ActionChains(driver).double_click(element).perform()

def check_if_chrome_running():
    """Check if any Chrome instances are running."""
    for process in psutil.process_iter(['name']):
        if process.info['name'] == 'chrome.exe':
            return True
    return False

def get_chrome_options(use_existing_profile):
    chrome_options = Options()
    if use_existing_profile:
        # Get the user's home directory dynamically
        home_dir = os.path.expanduser("~")
        
        # Construct the path to the Chrome user data directory
        user_data_dir = os.path.join(home_dir, "AppData", "Local", "Google", "Chrome", "User Data")
        
        # Add the user data directory to Chrome options
        chrome_options.add_argument(f"user-data-dir={user_data_dir}")
        
        # Optionally, specify the profile directory (e.g., "Default" for the default profile)
        profile_dir = "Default"  # Change to the specific profile if needed
        chrome_options.add_argument(f"profile-directory={profile_dir}")
    return chrome_options

def ask_user_choice():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    use_existing_profile = messagebox.askyesno(
        "Chrome Instance",
        "Would you like to use the currently installed version of Chrome?\n"
        "Click 'Yes' for the currently installed version or 'No' for a new instance."
    ) 
    root.destroy()  # Destroy the main window
    return use_existing_profile

def run_acura_script(excel_path):
    if check_if_chrome_running():
        raise Exception("The program has detected an instance of Google Chrome running on your system. Please ensure that all Chrome instances are closed before proceeding.")
    
    use_existing_profile = ask_user_choice()
    chrome_options = get_chrome_options(use_existing_profile)
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-extensions")  # Disable extensions to avoid conflicts
    chrome_options.add_argument("--disable-infobars")  # Disable infobars
    chrome_options.add_argument("--disable-browser-side-navigation")  # Disable side navigation issues
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Avoid detection as bot


    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 10)
    
    # Your structured data
    years_models_documents = {
        
        '2012': {
            'year_page_xpath': '//*[@data-grid-row="1"]/div[3]',
            'models': {
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="1"]/div[3]',
                        'BSW': '//*[@data-grid-row="2"]/div[3]',
                        'BUC': '//*[@data-grid-row="3"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'RL': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]'
                    }
                },
                'TL': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'BSW': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'TSX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'ZDX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]'
                    }
                }
            }
        },
        '2013': {
            'year_page_xpath': '//*[@data-grid-row="2"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'TL': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'BSW/RCTW': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'TSX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'ZDX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                }
            }
        },
        '2014': {
            'year_page_xpath': '//*[@data-grid-row="3"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA 1': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L126'  # Specify the exact cell for the hyperlink as a fall back
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'cell_address': 'L127'  # Specify the exact cell for the hyperlink as a fall back
                                }
                            }
                        },
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA 1': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L145'  # Specify the exact cell for the hyperlink as a fall back
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'cell_address': 'L146'  # Specify the exact cell for the hyperlink as a fall back
                                }
                            }
                        }
                    }
                },
                'TL': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'BSW/RCTW': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'TSX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                }
            }
        },
        '2015': {
            'year_page_xpath': '//*[@data-grid-row="4"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA 1': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L182'  # Specify the exact cell for the hyperlink as a fall back
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'cell_address': 'L183'  # Specify the exact cell for the hyperlink as a fall back
                                }
                            }
                        },
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'BUC': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA 1': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L201'  # Specify the exact cell for the hyperlink as a fall back
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'cell_address': 'L202'  # Specify the exact cell for the hyperlink as a fall back
                                }
                            }
                        }
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2016': {
            'year_page_xpath': '//*[@data-grid-row="5"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA Folder 2': {
                                    'folder2_xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'subdocuments2': {
                                        'LKA 1': {
                                            'xpath': '//*[@data-grid-row="0"]/div[3]',
                                            'cell_address': 'L231'  # Specify the exact cell for the hyperlink as a fallback
                                        }
                                    }
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L230'  # Specify the exact cell for the hyperlink as a fallback
                                },
                                'LKA 3': {
                                    'xpath': '//*[@data-grid-row="2"]/div[3]',
                                    'cell_address': 'L229'  # Specify the exact cell for the hyperlink as a fallback
                                }
                            }
                        },
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA Folder': {
                            'folder_xpath': '//*[@data-grid-row="0"]/div[3]',
                            'subdocuments': {
                                'LKA 1': {
                                    'xpath': '//*[@data-grid-row="0"]/div[3]',
                                    'cell_address': 'L249'  # Specify the exact cell for the hyperlink as a fall back
                                },
                                'LKA 2': {
                                    'xpath': '//*[@data-grid-row="1"]/div[3]',
                                    'cell_address': 'L250'  # Specify the exact cell for the hyperlink as a fall back
                                }
                            }
                        },
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2017': {
            'year_page_xpath': '//*[@data-grid-row="6"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'NSX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="1"]/div[3]',
                        'APA': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]',
                        'SVC': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2018': {
            'year_page_xpath': '//*[@data-grid-row="7"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'NSX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2019': {
            'year_page_xpath': '//*[@data-grid-row="8"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'NSX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2020': {
            'year_page_xpath': '//*[@data-grid-row="9"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'NSX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="5"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA': '//*[@data-grid-row="7"]/div[3]',
                        'SVC': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2021': {
            'year_page_xpath': '//*[@data-grid-row="10"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'AHL': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'AHL': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA': '//*[@data-grid-row="7"]/div[3]',
                        'SVC': '//*[@data-grid-row="8"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2022': {
            'year_page_xpath': '//*[@data-grid-row="11"]/div[3]',
            'models': {
                'ILX': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'NSX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]',
                        'BUC': '//*[@data-grid-row="2"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="2"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'BUC': '//*[@data-grid-row="6"]/div[3]',
                        'LKA': '//*[@data-grid-row="7"]/div[3]',
                        'SVC': '//*[@data-grid-row="8"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="4"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2023': {
            'year_page_xpath': '//*[@data-grid-row="12"]/div[3]',
            'models': {
                'Integra': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        },
        '2024': {
            'year_page_xpath': '//*[@data-grid-row="13"]/div[3]',
            'models': {
                'Integra': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'MDX': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'LKA': '//*[@data-grid-row="5"]/div[3]',
                        'SVC': '//*[@data-grid-row="6"]/div[3]'
                    }
                },
                'RDX': {
                    'model_page_xpath': '//*[@data-grid-row="2"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="1"]/div[3]',
                        'AEB': '//*[@data-grid-row="2"]/div[3]',
                        'AHL': '//*[@data-grid-row="3"]/div[3]',
                        'APA': '//*[@data-grid-row="4"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                },
                'TLX': {
                    'model_page_xpath': '//*[@data-grid-row="3"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="1"]/div[3]',
                        'APA': '//*[@data-grid-row="3"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="4"]/div[3]',
                        'BUC': '//*[@data-grid-row="5"]/div[3]',
                        'LKA': '//*[@data-grid-row="6"]/div[3]',
                        'SVC': '//*[@data-grid-row="7"]/div[3]'
                    }
                }
            }
        }
    }

    try:
        # Navigate to the main SharePoint page for Acura
        print("Navigating to Acura's main SharePoint page...")
        driver.get('https://calibercollision-my.sharepoint.com/:f:/g/personal/mark_klingenhofer_protechdfw_com/EjIo8sg9qXNEt6CDCKpeRGkBj8fLo4MSiJe7w0h7hZ30rQ?e=hFT1RF')
 
        # Give the user up to 120 seconds to log in
        max_wait_time = 120
        start_time = time.time()

        try:
            # Wait until the element with the specified XPath is found, or until 60 seconds have passed
            element = WebDriverWait(driver, max_wait_time).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-grid-row="1"]/div[3]'))
            )
        except:
            # If the element is not found within 120 seconds, print a message
            print("The element was not found within 120 seconds.")

        # Calculate the elapsed time
        elapsed_time = time.time() - start_time

        # If less than 60 seconds have passed and the element is found, continue with the rest of the code
        if elapsed_time < max_wait_time:
            # Continue with the code after successful login
            print("Element found, continuing with the code...")
        else:
            # Handle the situation where the element was not found within the allotted time
            print("Proceeding without finding the element...")
        
        # Setup some locators for finding table elements
        onedrive_page_name_locator = "//li[contains(@data-automationid, 'breadcrumb-listitem')]"
        onedrive_table_locator = "//div[@data-automationid='list-pages']/div[contains(@id, 'virtualized-list')]"  
        onedrive_table_row_locator = "./div[contains(@data-automationid, 'row') and contains(@id, 'virtualized-list')]"
        onedrive_table_row_column_locator = "./div[@role='gridcell' and contains(@data-automationid, '$FIELD_NAME')]"
       
        # Define some local helper functions for finding row information
        def is_row_folder(row_element: WebElement) -> bool:
            
            # Find the icon element and check if it's a folder or file
            icon_element_locator = onedrive_table_row_column_locator.replace("$FIELD_NAME", "field-DocIcon")
            icon_element = WebDriverWait(row_element, max_wait_time)\
                .until(EC.presence_of_element_located((By.XPATH, icon_element_locator)))
            
            # Return true if this folder is in the name, false if it is not
            return "folder" in icon_element.accessible_name
        def get_row_name(row_element: WebElement) -> str:
            
            # Find the name column element and return the name for the row in use
            return row_element.get_attribute("aria-label").strip()
        def get_folder_link(row_element: WebElement) -> str:
            
            # Build and return a new URL for this row entry
            base_url = driver.current_url.replace("&ga=1", "")    # Base URL for the current page
            row_name = get_row_name(row_element)                  # The name we're looking to open
            row_link = base_url + "%2F" + row_name                # Relative folder URL based on drive layout

            # Return the built URL here
            return row_link
        def get_file_link(row_element: WebElement) -> str:

            # Find the selector element and try to click it here
            selector_element_locator = onedrive_table_locator.replace("$FIELD_NAME", "row-selection")
            selector_element = WebDriverWait(row_element, max_wait_time)\
                .until(EC.presence_of_element_located((By.XPATH, selector_element_locator)))            
            
            # Pull the name element from the row and find child buttons for it
            name_element_locator = onedrive_table_row_column_locator.replace("$FIELD_NAME", "field-LinkFilename")
            name_element = row_element.find_element(By.XPATH, name_element_locator) 
            ActionChains(driver).move_to_element_with_offset(name_element, 50, 0).perform()
            
            # Find the share button element and click it here. Setup share settings and copy the link to the clipboard
            name_element.find_element(By.XPATH, ".//button[@data-automationid='shareHeroId']").click()
            time.sleep(0.75)
            ActionChains(driver).send_keys(Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.TAB, Keys.ENTER).perform()
            time.sleep(0.50)
            ActionChains(driver).send_keys(Keys.ARROW_DOWN, Keys.TAB, Keys.TAB, Keys.ENTER).perform()           
            time.sleep(0.50)
            ActionChains(driver).send_keys(Keys.ENTER).perform()  
            time.sleep(0.75)
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()             
                        
            # Pull the clipboard content and store it, then dump the link contents out of it
            win32clipboard.OpenClipboard()
            encrypted_file_link = win32clipboard.GetClipboardData()
            win32clipboard.CloseClipboard()

            # Return the stored link from the clipboard
            return encrypted_file_link
        def get_folder_rows(row_link: str = None) -> tuple[list, list]:
            
            # Navigate to the next link if needed and find the title of the page
            if row_link != None: driver.get(row_link)
            
            # Find the parent table element and find all child rows in it
            table_element = WebDriverWait(driver, max_wait_time)\
                .until(EC.presence_of_element_located((By.XPATH, onedrive_table_locator)))
            table_elements = table_element.find_elements(By.XPATH, onedrive_table_row_locator)

            # Find our page title once the table content has appeard and see if this is a year or model page
            page_title = driver.find_elements(By.XPATH, onedrive_page_name_locator)[-1].get_attribute("innerText").strip()
            is_year_folder = re.search("\\d{4}", page_title) != None        

            # Setup lists for the output files and folders
            indexed_files = [ ] 
            indexed_folders = [ ]       

            # Iterate all the rows and get link URLs for each one
            for row_element in table_elements:

                # Pull our row name before testing to filter rows we don't want
                row_name = get_row_name(row_element)
                if "no" in row_name.lower(): continue
                if "old" in row_name.lower(): continue
                
                # Check if this is a folder entry or not and make sure the name of the folder is a four digit year
                if not is_row_folder(row_element):
                    indexed_files.append(get_file_link(row_element))
                    continue

                # Before pulling a folder link, make sure it's either a Model or Year folder
                # Some models have a space in them to see if this is a year page or not first
                if not is_year_folder and ' ' in row_name: continue
                if re.search("\\d{4}|[^ \\n]+", row_name) == None: continue
            
                # Store the URL for the row entry on our list and move on
                row_link = get_folder_link(row_element)
                indexed_folders.append(row_link)

            # Return our built list of indexed rows and elements here
            return [indexed_folders, indexed_files]    

        # Index and store base folders and files here then iterate them all
        get_folder_results = get_folder_rows()
        base_files = get_folder_results[1]
        base_folders = get_folder_results[0]

        # Iterate the contents of the base folders list as long as it has contents
        while len(base_folders) > 0:

            # Store the current folder value and navigate to it for indexing
            folder_link = base_folders.pop(0)
            get_child_results = get_folder_rows(folder_link)
            
            # Add all of our links to the files and folders to our base lists
            for file_link in get_child_results[1]: base_files.append(file_link)   
            for folder_link in get_child_results[0]: base_folders.append(folder_link)
            
            # Log out how many child links and folders exist now
            print(f'{len(base_folders)} Folders Remain | {len(base_files)} Files Indexed')



        # Clicks Acura
        print("Locating Acura link and clicking...")
        double_click_element(driver, wait, '//*[@data-grid-row="1"]/div[3]')
        time.sleep(1)                       
        adas_last_row = {}
        
        # Ensure the Excel file is valid and can be opened
        try:
            wb = load_workbook(excel_path)
            ws = wb['Model Version']  # Correctly referencing the worksheet
            print(f"Workbook loaded successfully: {excel_path}")
        except Exception as e:
            print(f"Failed to open the Excel file: {e}")
            return

        for year, data in years_models_documents.items():
            print(f"Processing year: {year}")
            year_page_xpath = data['year_page_xpath']
            double_click_element(driver, wait, year_page_xpath)
            time.sleep(2)
            wait.until(EC.visibility_of_element_located((By.XPATH, year_page_xpath)))

            for model, model_data in data['models'].items():
                print(f"Accessing model: {model}")
                model_page_xpath = model_data['model_page_xpath']
                double_click_element(driver, wait, model_page_xpath)
                time.sleep(2)
                wait.until(EC.visibility_of_element_located((By.XPATH, model_page_xpath)))
                adas_last_row = {}  # Reset ADAS last row tracker for each model
                process_documents(driver, wait, ws, model_data, year, model, adas_last_row)
                navigate_back_to_element(driver, wait, year_page_xpath)  # Go back one more time to ensure stability
            navigate_back_to_element(driver, wait, '//*[@data-grid-row="1"]/div[3]')
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()
        print("Script completed, you may exit now.")
        
if __name__ == "__main__":
    excel_file_path = sys.argv[1]  # The Excel file path is expected as the first argument
    run_acura_script(excel_file_path)