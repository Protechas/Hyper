"""
PDF Annotation Detector and Extractor
Core functionality for detecting annotated pages and extracting them to separate files.
"""

import fitz  # PyMuPDF
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import os
import re
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class PDFAnnotationExtractor:
    """
    Handles detection and extraction of annotated pages from PDF files.
    """
    
    def __init__(self):
        self.supported_extensions = ['.pdf', '.PDF']
        # Regex patterns to detect multi-part documents
        # Matches: "part 1", "Part 2", "part-3", "Part_4", "part-1", etc.
        self.part_patterns = [
            r'\s+part[\s\-_]+(\d+)',  # "part 1", "part-1", "part_1"
            r'\s+Part[\s\-_]+(\d+)',  # "Part 1", "Part-1", "Part_1"
            r'[\s\-_]part[\s\-_](\d+)',  # variations with separators
            r'[\s\-_]Part[\s\-_](\d+)',  # variations with separators
        ]
        # Maximum file size in KB before compression (1400 KB = 1.4 MB)
        self.max_file_size_kb = 1400
        
        # Standard manufacturer names (approved list)
        self.standard_manufacturers = [
            'Acura', 'Alfa Romeo', 'Audi', 'BMW', 'BrightDrop', 'Buick', 
            'Cadillac', 'Chevrolet', 'Chrysler', 'Dodge', 'Fiat', 'Ford', 
            'Genesis', 'GMC', 'Honda', 'Hyundai', 'Infiniti', 'Jaguar', 
            'Jeep', 'Kia', 'Land Rover', 'Lexus', 'Lincoln', 'Mazda', 
            'Mercedes Benz', 'MINI', 'Mitsubishi', 'Nissan', 'Porsche', 
            'Ram', 'Rolls Royce', 'Subaru', 'Tesla', 'Toyota', 'Volkswagen', 
            'Volvo'
        ]
        
        # Common misspellings and variations mapping
        self.manufacturer_aliases = {
            'alfa romero': 'Alfa Romeo',
            'cadilac': 'Cadillac',
            'caddilac': 'Cadillac',
            'caddillac': 'Cadillac',
            'mercedes': 'Mercedes Benz',
            'mercedes-benz': 'Mercedes Benz',
            'vw': 'Volkswagen',
            'range rover': 'Land Rover',
            'rollsroyce': 'Rolls Royce',
            'rolls-royce': 'Rolls Royce',
            'suabru': 'Subaru',
            'nissian': 'Nissan',
            'porcshe': 'Porsche',
            'porsche': 'Porsche',
            'huyndai': 'Hyundai',
            'hyundia': 'Hyundai',
        }
    
    def is_glass_statement(self, filename: str) -> bool:
        """
        Check if a filename indicates a Glass Statement document.
        
        Args:
            filename: The filename to check
            
        Returns:
            True if it's a Glass Statement document, False otherwise
        """
        filename_lower = filename.lower()
        glass_keywords = ['glass statement', 'glass_statement', 'oem glass']
        return any(keyword in filename_lower for keyword in glass_keywords)
    
    def is_no_feature_document(self, filename: str) -> bool:
        """
        Check if a filename starts with 'No' (e.g., "No ACC For This Vehicle").
        These are documents indicating a feature is not available.
        
        Args:
            filename: The filename to check
            
        Returns:
            True if it's a "No feature" document, False otherwise
        """
        return filename.strip().startswith('No ')
    
    def is_unwanted_document(self, filename: str) -> bool:
        """
        Check if a filename matches patterns for unwanted support documents.
        These include Job Aids, Statements, generic guides, etc.
        
        Args:
            filename: The filename to check
            
        Returns:
            True if it's an unwanted document, False otherwise
        """
        filename_lower = filename.lower()
        
        # Keywords that indicate unwanted documents
        unwanted_keywords = [
            'swinnerton',
            'yewtree',
            'mopar',
            'job aid',
            'job-aid',
            'restraints warning',
            'restraint warning',
            'service aid',
            'servicing aid',
            'position statement',
            'post scan',
            'pre scan',
            'pre and post scan',
            'pre & post',
            'adas aid',
            'adas calibration aid',
            'general statement',
            'bumper repair',
            'collision repair',
            'ocs statement',
            'srs general',
            'oem parts statement',
            'alignment statement',
            'wheel alignment',
            'aiming guide',
            'aiming support',
            'adas configurations',
            'seat belt inspection',
            'sas position',
            'collision inspection',
            'windshield statement',
            'windshield guidelines',
            'windshield r&i'
        ]
        
        return any(keyword in filename_lower for keyword in unwanted_keywords)
    
    def normalize_manufacturer(self, manufacturer: str) -> Optional[str]:
        """
        Normalize a manufacturer name to match the standard list.
        Handles misspellings, variations, and partial matches.
        
        Args:
            manufacturer: The manufacturer name to normalize
            
        Returns:
            Normalized manufacturer name, or None if no match found
        """
        if not manufacturer:
            return None
        
        manufacturer_lower = manufacturer.lower().strip()
        
        # Check if it's in the aliases map
        if manufacturer_lower in self.manufacturer_aliases:
            return self.manufacturer_aliases[manufacturer_lower]
        
        # Check for exact match (case-insensitive)
        for standard in self.standard_manufacturers:
            if manufacturer_lower == standard.lower():
                return standard
        
        # Check if the parsed manufacturer starts with a known manufacturer
        # This handles cases like "Audi RS" -> "Audi" or "BMW I" -> "BMW"
        for standard in self.standard_manufacturers:
            if manufacturer_lower.startswith(standard.lower() + ' '):
                return standard
        
        # Check if a known manufacturer is contained in the parsed name
        # This helps with multi-word manufacturers
        for standard in self.standard_manufacturers:
            standard_words = standard.lower().split()
            manufacturer_words = manufacturer_lower.split()
            
            # If all words of standard manufacturer are in the parsed name
            if all(word in manufacturer_words for word in standard_words):
                return standard
        
        # No match found
        return None
    
    def parse_filename_structure(self, filename: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """
        Parse a filename to extract Year, Manufacturer, and Model.
        Expected formats:
          - "YYYY Manufacturer Model (...).pdf" → ("YYYY", "Manufacturer", "Model")
          - "YYYY-YYYY Manufacturer Model (...).pdf" → ("YYYY-YYYY", "Manufacturer", "Model")
          - "YYYY+ Manufacturer Model (...).pdf" → ("YYYY+", "Manufacturer", "Model")
        
        Args:
            filename: The filename to parse (without extension)
            
        Returns:
            Tuple of (year, manufacturer, model) or (None, None, None) if parsing fails
        """
        # Remove ALL occurrences of "_annotated_pages" suffix (handles re-processed files)
        clean_filename = filename
        while '_annotated_pages' in clean_filename:
            clean_filename = clean_filename.replace('_annotated_pages', '')
        clean_filename = clean_filename.strip()
        
        # Extract year first (could be range or single)
        year_pattern = r'^(\d{4}[\-\+]\d{0,4}|\d{4})'
        year_match = re.match(year_pattern, clean_filename)
        
        if not year_match:
            return (None, None, None)
        
        year = year_match.group(1)
        
        # Validate single year is reasonable
        if '-' not in year and '+' not in year:
            if not (1900 <= int(year) <= 2100):
                return (None, None, None)
        
        # Get the rest of the filename after the year
        after_year = clean_filename[year_match.end():].strip()
        
        # Handle BMW chassis codes (e.g., "F23 2 Series" -> should be "BMW 2 Series")
        # Common BMW chassis prefixes: F, G, E (followed by 2-3 digits)
        bmw_chassis_pattern = r'^([FGE]\d{2,3})\s+'
        chassis_match = re.match(bmw_chassis_pattern, after_year)
        if chassis_match:
            # This is likely a BMW with chassis code - prepend BMW
            after_year = 'BMW ' + after_year[chassis_match.end():]
        
        # Split by first parenthesis or closing parenthesis (handles malformed names)
        # This gives us the "Manufacturer Model" part
        paren_split = re.split(r'\s*[\(\)]', after_year, maxsplit=1)
        if not paren_split or not paren_split[0]:
            return (None, None, None)
        
        manufacturer_and_model = paren_split[0].strip()
        
        # Try to split manufacturer and model intelligently
        # Try matching 1, 2, and 3 word manufacturers against known list
        words = manufacturer_and_model.split()
        
        if len(words) < 2:
            return (None, None, None)  # Need at least manufacturer + model
        
        # Try 3-word manufacturer first (for "Alfa Romeo", "Land Rover", "Mercedes Benz")
        if len(words) >= 4:
            manufacturer_candidate = ' '.join(words[:3])
            normalized = self.normalize_manufacturer(manufacturer_candidate)
            if normalized:
                model = ' '.join(words[3:])
                model = self.clean_model_name(model)
                return (year, normalized, model)
        
        # Try 2-word manufacturer (for "Alfa Romeo", "Land Rover", etc.)
        if len(words) >= 3:
            manufacturer_candidate = ' '.join(words[:2])
            normalized = self.normalize_manufacturer(manufacturer_candidate)
            if normalized:
                model = ' '.join(words[2:])
                model = self.clean_model_name(model)
                return (year, normalized, model)
        
        # Try 1-word manufacturer (most common)
        manufacturer_candidate = words[0]
        normalized = self.normalize_manufacturer(manufacturer_candidate)
        if normalized:
            model = ' '.join(words[1:])
            model = self.clean_model_name(model)
            return (year, normalized, model)
        
        return (None, None, None)
    
    def clean_model_name(self, model: str) -> str:
        """
        Clean up model name by extracting vehicle type indicators and appending them.
        Converts [EV], [HEV], [PHEV], [FCEV] from brackets to part of the model name.
        
        Example: "Volt [PHEV]" becomes "Volt PHEV"
        
        Args:
            model: The model name to clean
            
        Returns:
            Cleaned model name with EV type appended (if present)
        """
        if not model:
            return model
        
        # Extract vehicle type indicators: [EV], [HEV], [PHEV], [FCEV]
        # These indicate electric/hybrid/fuel cell versions - append to model name
        ev_type = None
        
        # Check for each type and extract it
        if re.search(r'\[EV\]', model, flags=re.IGNORECASE):
            ev_type = 'EV'
            model = re.sub(r'\s*\[EV\]\s*', ' ', model, flags=re.IGNORECASE)
        elif re.search(r'\[HEV\]', model, flags=re.IGNORECASE):
            ev_type = 'HEV'
            model = re.sub(r'\s*\[HEV\]\s*', ' ', model, flags=re.IGNORECASE)
        elif re.search(r'\[PHEV\]', model, flags=re.IGNORECASE):
            ev_type = 'PHEV'
            model = re.sub(r'\s*\[PHEV\]\s*', ' ', model, flags=re.IGNORECASE)
        elif re.search(r'\[FCEV\]', model, flags=re.IGNORECASE):
            ev_type = 'FCEV'
            model = re.sub(r'\s*\[FCEV\]\s*', ' ', model, flags=re.IGNORECASE)
        
        # Clean up extra whitespace
        model = ' '.join(model.split())
        
        # Append the EV type to the model name (without brackets)
        if ev_type:
            model = f"{model} {ev_type}"
        
        return model.strip()
    
    def get_organized_output_path(self, base_output_dir: Path, filename: str, 
                                   use_organization: bool = True, 
                                   source_path: Path = None, input_dir: Path = None,
                                   preserve_structure: bool = False) -> Path:
        """
        Get the organized output directory path based on filename structure.
        Preserves nested subfolder structure within the Model folder if present.
        
        Args:
            base_output_dir: The base output directory
            filename: The filename to parse (without extension)
            use_organization: Whether to use organized folder structure
            source_path: The original source file path (for preserving subfolder structure)
            input_dir: The input directory root (for calculating relative paths)
            preserve_structure: If True, preserve exact input structure without reorganization
            
        Returns:
            Path to the organized output directory
        """
        # If preserve_structure is True, maintain exact relative path from input
        if preserve_structure and source_path and input_dir:
            try:
                rel_path = source_path.parent.relative_to(input_dir)
                organized_dir = base_output_dir / rel_path
                organized_dir.mkdir(parents=True, exist_ok=True)
                return organized_dir
            except (ValueError, RuntimeError):
                # If relative_to fails, fall back to base directory
                pass
        
        if not use_organization:
            return base_output_dir
        
        year, manufacturer, model = self.parse_filename_structure(filename)
        
        if year and manufacturer and model:
            # Create base nested structure: Manufacturer/Year/Model/
            organized_dir = base_output_dir / manufacturer / year / model
            
            # If source_path and input_dir are provided, preserve additional subfolder structure
            if source_path and input_dir:
                # Get the relative path from input_dir to the file's parent directory
                try:
                    rel_path = source_path.parent.relative_to(input_dir)
                    
                    # Check if there are additional subdirectories beyond the expected structure
                    # Expected structure: Make/Year/Model or just files at various levels
                    # We want to preserve any subdirectories within the Model folder
                    parts = rel_path.parts
                    
                    # Try to find where the Model folder is in the path
                    # Look for a path component that matches the model name
                    model_index = -1
                    for i, part in enumerate(parts):
                        # Check if this part contains the model name
                        if model.lower() in part.lower():
                            model_index = i
                            break
                    
                    # If we found the model folder and there are subdirectories after it,
                    # preserve those subdirectories
                    if model_index >= 0 and model_index + 1 < len(parts):
                        # Get the subdirectory path after the model folder
                        subfolder_parts = parts[model_index + 1:]
                        subfolder_path = Path(*subfolder_parts)
                        organized_dir = organized_dir / subfolder_path
                except (ValueError, RuntimeError):
                    # If relative_to fails, just use the base organized directory
                    pass
        else:
            # Place in Unknown folder if parsing fails
            organized_dir = base_output_dir / "Unknown"
        
        # Create the directory if it doesn't exist
        organized_dir.mkdir(parents=True, exist_ok=True)
        
        return organized_dir
    
    def compress_pdf_smart(self, pdf_path: Path, max_size_kb: int = None) -> Tuple[bool, int, int]:
        """
        Smart compression: Compress embedded images while preserving text as vectors.
        Text remains selectable, searchable, and perfect quality at all compression levels.
        
        Args:
            pdf_path: Path to the PDF file to compress
            max_size_kb: Maximum size in KB (uses self.max_file_size_kb if not specified)
            
        Returns:
            Tuple of (was_compressed, original_size_kb, new_size_kb)
        """
        if max_size_kb is None:
            max_size_kb = self.max_file_size_kb
        
        # Get original file size
        original_size = pdf_path.stat().st_size
        original_size_kb = original_size / 1024
        
        # Check if compression is needed
        if original_size_kb <= max_size_kb:
            return (False, int(original_size_kb), int(original_size_kb))
        
        print(f"  🔧 Compressing {pdf_path.name} ({int(original_size_kb)}KB)...")
        
        try:
            # Progressive image compression levels
            # Format: (max_image_dimension, jpeg_quality)
            # More aggressive progression to handle stubborn files
            quality_levels = [
                (2048, 85),  # Excellent - large images, high quality
                (1600, 75),  # High quality
                (1200, 70),  # Good quality
                (1024, 65),  # Standard quality
                (800, 60),   # Lower quality
                (640, 55),   # Aggressive
                (512, 50),   # More aggressive
                (400, 45),   # Very aggressive
                (320, 40),   # Extreme
                (256, 35),   # Ultra extreme
                (200, 30),   # Maximum
                (150, 25),   # Beyond maximum
                (100, 20),   # Extreme compression
            ]
            
            temp_path = pdf_path.with_suffix('.tmp.pdf')
            
            for max_dim, jpeg_quality in quality_levels:
                # Open original document
                doc = fitz.open(pdf_path)
                
                # Process each page - compress only images, keep text as vectors
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    
                    # Get all images on this page
                    image_list = page.get_images()
                    
                    for img_info in image_list:
                        xref = img_info[0]
                        
                        try:
                            # Extract the image
                            base_image = doc.extract_image(xref)
                            image_bytes = base_image["image"]
                            
                            # Create pixmap from image
                            pix = fitz.Pixmap(image_bytes)
                            
                            # Convert CMYK to RGB (saves space)
                            if pix.n >= 5:  # CMYK or more channels
                                pix = fitz.Pixmap(fitz.csRGB, pix)
                            
                            # Downsample if image is larger than max dimension
                            if pix.width > max_dim or pix.height > max_dim:
                                scale = min(max_dim / pix.width, max_dim / pix.height)
                                mat = fitz.Matrix(scale, scale)
                                pix = fitz.Pixmap(pix, mat)
                            
                            # Compress to JPEG at specified quality
                            compressed_image = pix.tobytes("jpeg", jpeg_quality)
                            
                            # Replace the image stream in the PDF
                            doc.update_stream(xref, compressed_image)
                            
                        except Exception:
                            # Skip problematic images
                            continue
                
                # Save with PDF-level compression
                doc.save(
                    temp_path,
                    garbage=4,
                    deflate=True,
                    clean=True,
                    deflate_images=True,
                    deflate_fonts=True
                )
                doc.close()
                
                # Check size
                new_size_kb = temp_path.stat().st_size / 1024
                
                print(f"    Attempt: Max {max_dim}px images, {jpeg_quality}% quality → {int(new_size_kb)}KB")
                
                # If size is acceptable, replace original and we're done
                if new_size_kb <= max_size_kb:
                    temp_path.replace(pdf_path)
                    print(f"  ✅ Compressed: {int(original_size_kb)}KB → {int(new_size_kb)}KB (text preserved)")
                    return (True, int(original_size_kb), int(new_size_kb))
                
                # Clean up for next attempt
                if temp_path.exists():
                    temp_path.unlink()
            
            # Use most compressed version even if still over limit
            final_size_kb = temp_path.stat().st_size / 1024 if temp_path.exists() else original_size_kb
            
            if temp_path.exists():
                temp_path.replace(pdf_path)
            
            if final_size_kb > max_size_kb:
                print(f"  ⚠️ Warning: {pdf_path.name} compressed to {int(final_size_kb)}KB (target: {max_size_kb}KB)")
            else:
                print(f"  ✅ Compressed: {int(original_size_kb)}KB → {int(final_size_kb)}KB (text preserved)")
            
            return (True, int(original_size_kb), int(final_size_kb))
            
        except Exception as e:
            print(f"  ❌ Error compressing {pdf_path.name}: {str(e)}")
            import traceback
            traceback.print_exc()
            # Clean up temp files
            temp_path = pdf_path.with_suffix('.tmp.pdf')
            if temp_path.exists():
                temp_path.unlink()
            return (False, int(original_size_kb), int(original_size_kb))
    
    def compress_pdf(self, pdf_path: Path, max_size_kb: int = None, max_attempts: int = 12) -> Tuple[bool, int, int]:
        """
        Compress a PDF file aggressively to get below the maximum size.
        Uses iterative compression and image downsampling if needed.
        
        Args:
            pdf_path: Path to the PDF file to compress
            max_size_kb: Maximum size in KB (uses self.max_file_size_kb if not specified)
            max_attempts: Maximum number of compression attempts
            
        Returns:
            Tuple of (was_compressed, original_size_kb, new_size_kb)
        """
        if max_size_kb is None:
            max_size_kb = self.max_file_size_kb
        
        # Get original file size
        original_size = pdf_path.stat().st_size
        original_size_kb = original_size / 1024
        
        # Check if compression is needed
        if original_size_kb <= max_size_kb:
            return (False, int(original_size_kb), int(original_size_kb))
        
        try:
            # Attempt 1: Standard aggressive compression
            doc = fitz.open(pdf_path)
            temp_path = pdf_path.with_suffix('.tmp.pdf')
            
            doc.save(
                temp_path,
                garbage=4,           # Maximum garbage collection
                deflate=True,        # Compress streams
                clean=True,          # Clean up unused objects
                linear=False,        # Don't linearize (saves space)
                pretty=False,        # Don't pretty-print (saves space)
                deflate_images=True, # Compress images
                deflate_fonts=True   # Compress fonts
            )
            doc.close()
            
            # Check size after first compression
            new_size_kb = temp_path.stat().st_size / 1024
            
            # If still too large, try more aggressive techniques
            attempt = 1
            while new_size_kb > max_size_kb and attempt < max_attempts:
                attempt += 1
                
                # Re-open the compressed document
                doc = fitz.open(temp_path)
                temp_path2 = pdf_path.with_suffix('.tmp2.pdf')
                
                # Calculate image quality reduction factor based on how much we need to reduce
                reduction_factor = max_size_kb / new_size_kb
                
                # Progressively more aggressive scaling (12 passes total)
                if attempt == 2:
                    max_dimension = 1024
                    base_quality = 60
                    convert_grayscale = False
                elif attempt == 3:
                    max_dimension = 800
                    base_quality = 50
                    convert_grayscale = False
                elif attempt == 4:
                    max_dimension = 640
                    base_quality = 42
                    convert_grayscale = False
                elif attempt == 5:
                    max_dimension = 512
                    base_quality = 36
                    convert_grayscale = False
                elif attempt == 6:
                    max_dimension = 400
                    base_quality = 30
                    convert_grayscale = False
                elif attempt == 7:
                    max_dimension = 320
                    base_quality = 26
                    convert_grayscale = False
                elif attempt == 8:
                    max_dimension = 256
                    base_quality = 22
                    convert_grayscale = False
                elif attempt == 9:
                    max_dimension = 200
                    base_quality = 20
                    convert_grayscale = True  # Start converting to grayscale
                elif attempt == 10:
                    max_dimension = 160
                    base_quality = 18
                    convert_grayscale = True
                elif attempt == 11:
                    max_dimension = 128
                    base_quality = 16
                    convert_grayscale = True
                else:  # attempt 12
                    max_dimension = 96
                    base_quality = 15
                    convert_grayscale = True
                
                # Process each page and downsample images
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    
                    # Get all images on the page
                    image_list = page.get_images()
                    
                    for img_index, img in enumerate(image_list):
                        xref = img[0]  # xref number
                        
                        try:
                            # Extract image
                            base_image = doc.extract_image(xref)
                            image_bytes = base_image["image"]
                            
                            # Convert to pixmap and downsample
                            pix = fitz.Pixmap(image_bytes)
                            
                            # Convert CMYK to RGB first (saves space)
                            if pix.n > 4:  # CMYK or other color space
                                pix = fitz.Pixmap(fitz.csRGB, pix)
                            
                            # Convert to grayscale for extreme compression (passes 9-12)
                            if convert_grayscale and pix.n >= 3:  # RGB or RGBA
                                pix = fitz.Pixmap(fitz.csGRAY, pix)
                            
                            # Downsample: reduce dimensions aggressively
                            if pix.width > max_dimension or pix.height > max_dimension:
                                # Scale down images
                                scale = min(max_dimension / pix.width, max_dimension / pix.height)
                                mat = fitz.Matrix(scale, scale)
                                pix = fitz.Pixmap(pix, mat)
                            
                            # Save with JPEG compression (progressively lower quality)
                            quality = max(15, int(base_quality * reduction_factor))
                            
                            # Use JPEG for color/grayscale
                            if pix.n >= 3:  # Color
                                img_bytes = pix.tobytes("jpeg", quality)
                            else:  # Grayscale - JPEG works better than PNG for compression
                                img_bytes = pix.tobytes("jpeg", quality)
                            
                            # Replace image in document
                            doc.update_stream(xref, img_bytes)
                            
                        except Exception as img_error:
                            # Skip problematic images
                            continue
                
                # Save with maximum compression again
                doc.save(
                    temp_path2,
                    garbage=4,
                    deflate=True,
                    clean=True,
                    linear=False,
                    pretty=False,
                    deflate_images=True,
                    deflate_fonts=True
                )
                doc.close()
                
                # Update temp file
                if temp_path.exists():
                    temp_path.unlink()
                temp_path2.rename(temp_path)
                
                # Check new size
                new_size_kb = temp_path.stat().st_size / 1024
            
            # Get final size
            final_size_kb = temp_path.stat().st_size / 1024
            
            # Replace original with compressed version
            temp_path.replace(pdf_path)
            
            # Log if we couldn't get below target
            if final_size_kb > max_size_kb:
                print(f"  ⚠️ Warning: {pdf_path.name} compressed to {int(final_size_kb)}KB (target: {max_size_kb}KB)")
            
            return (True, int(original_size_kb), int(final_size_kb))
            
        except Exception as e:
            print(f"  ⚠️ Warning: Could not compress {pdf_path.name}: {str(e)}")
            # Clean up temp files if they exist
            for temp_name in ['.tmp.pdf', '.tmp2.pdf']:
                temp_path = pdf_path.with_suffix(temp_name)
                if temp_path.exists():
                    temp_path.unlink()
            return (False, int(original_size_kb), int(original_size_kb))
    
    def extract_part_info(self, filename: str) -> Tuple[Optional[str], Optional[int], Optional[str]]:
        """
        Extract part information from a filename.
        
        Args:
            filename: The filename to parse (without extension)
            
        Returns:
            Tuple of (base_name, part_number, matched_pattern) or (None, None, None) if no part found
            base_name: The filename without the part suffix
            part_number: The part number as integer
            matched_pattern: The actual matched text
        """
        for pattern in self.part_patterns:
            match = re.search(pattern, filename, re.IGNORECASE)
            if match:
                part_num = int(match.group(1))
                base_name = filename[:match.start()]
                matched_text = match.group(0)
                return (base_name.strip(), part_num, matched_text)
        
        return (None, None, None)
    
    def group_multipart_documents(self, pdf_files: List[Path]) -> Dict[str, List[Path]]:
        """
        Group multi-part documents together.
        
        Args:
            pdf_files: List of PDF file paths
            
        Returns:
            Dictionary mapping base names to lists of file paths
            Single files will have their full path as the key with a single-item list
            Multi-part files will be grouped under their base name
        """
        groups = defaultdict(list)
        
        for pdf_path in pdf_files:
            filename = pdf_path.stem  # filename without extension
            base_name, part_num, _ = self.extract_part_info(filename)
            
            if base_name and part_num:
                # This is a multi-part document
                # Use the base name as the group key
                groups[base_name].append((part_num, pdf_path))
            else:
                # This is a single document
                # Use the full path as the key to keep duplicates separate
                # This prevents duplicate filenames from different directories being grouped
                groups[str(pdf_path)].append((0, pdf_path))
        
        # Sort the grouped files by part number and convert to just paths
        result = {}
        for base_name, files in groups.items():
            sorted_files = sorted(files, key=lambda x: x[0])  # Sort by part number
            # Remove any duplicate paths (in case the same file was added twice)
            unique_paths = []
            seen = set()
            for _, path in sorted_files:
                if str(path) not in seen:
                    unique_paths.append(path)
                    seen.add(str(path))
            result[base_name] = unique_paths
        
        return result
        
    def detect_annotated_pages(self, pdf_path: Path) -> List[int]:
        """
        Detect which pages in a PDF contain annotations.
        
        Args:
            pdf_path: Path to the PDF file
            
        Returns:
            List of page numbers (0-indexed) that contain annotations
        """
        annotated_pages = []
        
        try:
            doc = fitz.open(pdf_path)
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # Check for annotations
                annots = page.annots()
                if annots:
                    has_annotation = False
                    for annot in annots:
                        # Filter out link annotations (internal PDF links)
                        # We want actual human annotations
                        if annot.type[0] not in [fitz.PDF_ANNOT_LINK]:
                            has_annotation = True
                            break
                    
                    if has_annotation:
                        annotated_pages.append(page_num)
            
            doc.close()
            
        except Exception as e:
            print(f"Error processing {pdf_path}: {str(e)}")
            return []
        
        return annotated_pages
    
    def extract_pages(self, pdf_path: Path, page_numbers: List[int], output_path: Path, 
                     combine: bool = True) -> bool:
        """
        Extract specific pages from a PDF and save to a new file.
        
        Args:
            pdf_path: Path to source PDF
            page_numbers: List of page numbers to extract (0-indexed)
            output_path: Path where extracted PDF(s) should be saved
            combine: If True, combine all pages into one PDF. If False, save each page separately.
            
        Returns:
            True if successful, False otherwise
        """
        if not page_numbers:
            return False
        
        try:
            doc = fitz.open(pdf_path)
            
            if combine:
                # Create a new PDF with all annotated pages
                new_doc = fitz.open()
                for page_num in page_numbers:
                    new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
                
                new_doc.save(output_path)
                new_doc.close()
                print(f"Saved combined annotated pages to: {output_path}")
                
            else:
                # Save each page as a separate file
                output_dir = output_path.parent
                base_name = output_path.stem
                
                for idx, page_num in enumerate(page_numbers, 1):
                    new_doc = fitz.open()
                    new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
                    
                    page_output = output_dir / f"{base_name}_page_{page_num + 1}.pdf"
                    new_doc.save(page_output)
                    new_doc.close()
                    print(f"Saved page {page_num + 1} to: {page_output}")
            
            doc.close()
            return True
            
        except Exception as e:
            print(f"Error extracting pages from {pdf_path}: {str(e)}")
            return False
    
    def process_single_pdf(self, pdf_path: Path, output_dir: Path, 
                          combine: bool = True, organize_folders: bool = False,
                          keep_parts_separate: bool = False, base_name: str = None,
                          copy_all_files: bool = False, input_dir: Path = None) -> Dict:
        """
        Process a single PDF: detect annotations and extract annotated pages.
        
        Args:
            pdf_path: Path to the PDF file
            output_dir: Directory where output should be saved
            combine: Whether to combine pages into one file or save separately
            organize_folders: Whether to organize output into Manufacturer/Year/Model folders
            keep_parts_separate: Whether to keep multi-part docs in separate files (in subfolder)
            base_name: Base name for subfolder if this is part of a multi-part document
            copy_all_files: Whether to copy files even if they have no annotations
            input_dir: The input directory root (for preserving relative subfolder structure)
            
        Returns:
            Dictionary with processing results
        """
        result = {
            'file': str(pdf_path),
            'success': False,
            'annotated_pages': [],
            'output_file': None,
            'error': None,
            'compressed': False,
            'original_size_kb': 0,
            'final_size_kb': 0,
            'copied_without_annotations': False
        }
        
        # Detect annotated pages
        annotated_pages = self.detect_annotated_pages(pdf_path)
        result['annotated_pages'] = annotated_pages
        
        if not annotated_pages:
            # If no annotations and copy_all_files is enabled, copy the entire file
            if copy_all_files:
                try:
                    # Prepare output path - preserve exact structure if input_dir provided
                    # Otherwise use organization mode
                    preserve_exact = (input_dir is not None)
                    organized_dir = self.get_organized_output_path(
                        output_dir, pdf_path.stem, organize_folders, 
                        source_path=pdf_path, input_dir=input_dir,
                        preserve_structure=preserve_exact
                    )
                    
                    # If this is part of a multi-part doc and we're keeping parts separate, create subfolder
                    if keep_parts_separate and base_name:
                        organized_dir = organized_dir / base_name
                        organized_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Output filename
                    output_filename = pdf_path.name
                    output_path = organized_dir / output_filename
                    
                    # Copy the entire file
                    import shutil
                    shutil.copy2(pdf_path, output_path)
                    
                    result['success'] = True
                    result['output_file'] = str(output_path)
                    result['error'] = 'No annotations - full file copied'
                    result['copied_without_annotations'] = True
                    return result
                except Exception as e:
                    result['success'] = False
                    result['error'] = f'Failed to copy file: {str(e)}'
                    return result
            else:
                result['success'] = True
                result['error'] = 'No annotated pages found'
                return result
        
        # Prepare output path - preserve exact structure if input_dir provided
        # Otherwise use organization mode
        preserve_exact = (input_dir is not None)
        organized_dir = self.get_organized_output_path(
            output_dir, pdf_path.stem, organize_folders,
            source_path=pdf_path, input_dir=input_dir,
            preserve_structure=preserve_exact
        )
        
        # If this is part of a multi-part doc and we're keeping parts separate, create subfolder
        if keep_parts_separate and base_name:
            organized_dir = organized_dir / base_name
            organized_dir.mkdir(parents=True, exist_ok=True)
        
        # Output filename - use original name, clean of any suffixes
        # Remove _part-1, _part-2 etc to avoid duplication when re-processing
        clean_name = pdf_path.stem
        # Remove _annotated_pages if present
        while '_annotated_pages' in clean_name:
            clean_name = clean_name.replace('_annotated_pages', '')
        
        output_filename = f"{clean_name}.pdf"
        output_path = organized_dir / output_filename
        
        # Extract pages
        success = self.extract_pages(pdf_path, annotated_pages, output_path, combine)
        
        if success:
            # Compress if needed using smart compression
            was_compressed, orig_size, new_size = self.compress_pdf_smart(output_path)
            
            result['success'] = True
            result['output_file'] = str(output_path)
            result['compressed'] = was_compressed
            result['original_size_kb'] = orig_size
            result['final_size_kb'] = new_size
        else:
            result['error'] = 'Failed to extract pages'
        
        return result
    
    def process_multipart_group(self, pdf_paths: List[Path], base_name: str, 
                               output_dir: Path, combine: bool = True, 
                               organize_folders: bool = False, 
                               keep_parts_separate: bool = False,
                               copy_all_files: bool = False,
                               input_dir: Path = None) -> Dict:
        """
        Process a group of multi-part PDFs: combine into one output OR keep separate in subfolder.
        
        Args:
            pdf_paths: List of PDF file paths (all parts)
            base_name: Base name for the output file or subfolder
            output_dir: Directory where output should be saved
            combine: Whether to combine pages into one file or save separately
            organize_folders: Whether to organize output into Manufacturer/Year/Model folders
            keep_parts_separate: Whether to keep parts as separate files in a subfolder
            copy_all_files: Whether to copy files even if they have no annotations
            input_dir: The input directory root (for preserving relative subfolder structure)
            
        Returns:
            Dictionary with processing results
        """
        result = {
            'file': base_name + ' (multi-part)',
            'parts': [str(p) for p in pdf_paths],
            'success': False,
            'annotated_pages': [],
            'output_file': None,
            'error': None,
            'compressed': False,
            'original_size_kb': 0,
            'final_size_kb': 0,
            'copied_without_annotations': False
        }
        
        # If keeping parts separate, process each part individually
        if keep_parts_separate:
            results = []
            all_annotated_pages = []
            files_copied = 0
            
            for pdf_path in pdf_paths:
                part_result = self.process_single_pdf(
                    pdf_path, output_dir, combine, organize_folders,
                    keep_parts_separate=True, base_name=base_name,
                    copy_all_files=copy_all_files, input_dir=input_dir
                )
                results.append(part_result)
                
                if part_result.get('annotated_pages'):
                    # Convert single file annotated pages (list of ints) to tuple format for consistency
                    for page_num in part_result['annotated_pages']:
                        all_annotated_pages.append((pdf_path.name, page_num + 1))
                elif part_result.get('copied_without_annotations'):
                    files_copied += 1
            
            # Combine results for reporting
            if all_annotated_pages or files_copied > 0:
                result['success'] = True
                result['annotated_pages'] = all_annotated_pages
                if all_annotated_pages:
                    result['output_file'] = f"Multiple files in {base_name}/ folder"
                elif files_copied > 0:
                    result['output_file'] = f"{files_copied} file(s) copied in {base_name}/ folder"
                    result['copied_without_annotations'] = True
                
                # Compression stats - sum from all parts
                any_compressed = any(r.get('compressed') for r in results)
                if any_compressed:
                    result['compressed'] = True
                    result['original_size_kb'] = sum(r.get('original_size_kb', 0) for r in results)
                    result['final_size_kb'] = sum(r.get('final_size_kb', 0) for r in results)
            else:
                result['success'] = True
                result['error'] = 'No annotated pages found in any part'
            
            return result
        
        # Original combined approach
        # Collect all annotated pages from all parts
        all_annotated_pages = []
        
        try:
            # Create a new document to hold all annotated pages
            combined_doc = fitz.open()
            
            for pdf_path in pdf_paths:
                # Detect annotated pages in this part
                annotated_pages = self.detect_annotated_pages(pdf_path)
                
                if annotated_pages:
                    # Track for reporting
                    all_annotated_pages.extend([(pdf_path.name, p+1) for p in annotated_pages])
                    
                    # Extract and add to combined document
                    doc = fitz.open(pdf_path)
                    for page_num in annotated_pages:
                        combined_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)
                    doc.close()
            
            if len(combined_doc) > 0:
                # Save the combined document with organization, preserving subfolder structure
                # Use the first PDF path as reference for subfolder structure
                preserve_exact = (input_dir is not None)
                organized_dir = self.get_organized_output_path(
                    output_dir, base_name, organize_folders,
                    source_path=pdf_paths[0], input_dir=input_dir,
                    preserve_structure=preserve_exact
                )
                
                # Clean base name of any suffixes to avoid duplication
                clean_base_name = base_name
                while '_annotated_pages' in clean_base_name:
                    clean_base_name = clean_base_name.replace('_annotated_pages', '')
                
                output_filename = f"{clean_base_name}.pdf"
                output_path = organized_dir / output_filename
                
                combined_doc.save(output_path)
                combined_doc.close()
                
                # Compress if needed using smart compression
                was_compressed, orig_size, new_size = self.compress_pdf_smart(output_path)
                
                result['success'] = True
                result['annotated_pages'] = all_annotated_pages
                result['output_file'] = str(output_path)
                result['compressed'] = was_compressed
                result['original_size_kb'] = orig_size
                result['final_size_kb'] = new_size
            else:
                combined_doc.close()
                
                # If no annotations but copy_all_files is enabled, copy all parts
                if copy_all_files:
                    try:
                        # Create a combined PDF from all parts
                        combined_full_doc = fitz.open()
                        for pdf_path in pdf_paths:
                            doc = fitz.open(pdf_path)
                            combined_full_doc.insert_pdf(doc)
                            doc.close()
                        
                        # Save the combined document, preserving subfolder structure
                        preserve_exact = (input_dir is not None)
                        organized_dir = self.get_organized_output_path(
                            output_dir, base_name, organize_folders,
                            source_path=pdf_paths[0], input_dir=input_dir,
                            preserve_structure=preserve_exact
                        )
                        clean_base_name = base_name
                        while '_annotated_pages' in clean_base_name:
                            clean_base_name = clean_base_name.replace('_annotated_pages', '')
                        
                        output_filename = f"{clean_base_name}.pdf"
                        output_path = organized_dir / output_filename
                        
                        combined_full_doc.save(output_path)
                        combined_full_doc.close()
                        
                        result['success'] = True
                        result['output_file'] = str(output_path)
                        result['error'] = 'No annotations - full combined file copied'
                        result['copied_without_annotations'] = True
                    except Exception as e:
                        result['success'] = False
                        result['error'] = f'Failed to copy multi-part files: {str(e)}'
                else:
                    result['success'] = True
                    result['error'] = 'No annotated pages found in any part'
        
        except Exception as e:
            result['error'] = f'Failed to process multi-part document: {str(e)}'
        
        return result
    
    def scan_folder(self, input_dir: Path, output_dir: Path, 
                   combine: bool = True, recursive: bool = True, 
                   combine_parts: bool = False, ignore_glass_statements: bool = False,
                   ignore_no_feature_docs: bool = True, ignore_support_docs: bool = True,
                   organize_folders: bool = False, keep_parts_separate: bool = False,
                   copy_all_files: bool = False) -> List[Dict]:
        """
        Scan a folder (and optionally subfolders) for PDFs and process them.
        
        Args:
            input_dir: Directory to scan for PDFs
            output_dir: Directory where outputs should be saved
            combine: Whether to combine pages into one file per PDF
            recursive: Whether to scan subfolders
            combine_parts: Whether to combine multi-part documents together
            ignore_glass_statements: Whether to skip Glass Statement documents
            ignore_no_feature_docs: Whether to skip "No [Feature] For This Vehicle" documents
            ignore_support_docs: Whether to skip Job Aids, Statements, and other support documents
            organize_folders: Whether to organize output into Manufacturer/Year/Model folders
            keep_parts_separate: Whether to keep multi-part docs in subfolders instead of combining
            copy_all_files: Whether to copy files even if they have no annotations
            
        Returns:
            List of result dictionaries for each PDF processed
        """
        results = []
        
        # Create output directory if it doesn't exist
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Find all PDF files
        if recursive:
            pdf_files = []
            for ext in self.supported_extensions:
                pdf_files.extend(input_dir.rglob(f"*{ext}"))
        else:
            pdf_files = []
            for ext in self.supported_extensions:
                pdf_files.extend(input_dir.glob(f"*{ext}"))
        
        # Remove duplicates (in case same file appears multiple times)
        pdf_files = list(set(pdf_files))
        
        # Track initial count and filtering
        initial_count = len(pdf_files)
        filtered_counts = {
            'Glass Statements': 0,
            'No Feature docs': 0,
            'Support docs': 0
        }
        
        # Filter out unwanted documents
        if ignore_glass_statements:
            pdf_files = [pdf for pdf in pdf_files if not self.is_glass_statement(pdf.stem)]
            filtered_counts['Glass Statements'] = initial_count - len(pdf_files)
            if filtered_counts['Glass Statements'] > 0:
                print(f"\nFiltered out {filtered_counts['Glass Statements']} Glass Statement document(s)")
        
        if ignore_no_feature_docs:
            before_no_filter = len(pdf_files)
            pdf_files = [pdf for pdf in pdf_files if not self.is_no_feature_document(pdf.stem)]
            filtered_counts['No Feature docs'] = before_no_filter - len(pdf_files)
            if filtered_counts['No Feature docs'] > 0:
                print(f"Filtered out {filtered_counts['No Feature docs']} 'No Feature' document(s)")
        
        if ignore_support_docs:
            before_support_filter = len(pdf_files)
            pdf_files = [pdf for pdf in pdf_files if not self.is_unwanted_document(pdf.stem)]
            filtered_counts['Support docs'] = before_support_filter - len(pdf_files)
            if filtered_counts['Support docs'] > 0:
                print(f"Filtered out {filtered_counts['Support docs']} support/generic document(s)")
        
        print(f"\nFound {len(pdf_files)} PDF file(s) to process\n")
        
        # Group multi-part documents if combine_parts is enabled
        if combine_parts:
            grouped_docs = self.group_multipart_documents(pdf_files)
            total_groups = len(grouped_docs)
            print(f"Grouped into {total_groups} document(s) (combining multi-part documents)\n")
            
            # Process each group
            idx = 1
            for base_name, pdf_paths in grouped_docs.items():
                if len(pdf_paths) > 1:
                    # Multi-part document
                    print(f"[{idx}/{total_groups}] Processing multi-part: {base_name} ({len(pdf_paths)} parts)")
                    result = self.process_multipart_group(
                        pdf_paths, base_name, output_dir, combine, organize_folders, 
                        keep_parts_separate, copy_all_files, input_dir=input_dir
                    )
                    results.append(result)
                    
                    if result['success'] and result['annotated_pages']:
                        print(f"  ✓ Found {len(result['annotated_pages'])} annotated page(s) across all parts")
                        for part_name, page_num in result['annotated_pages']:
                            print(f"    - {part_name}: page {page_num}")
                        if result.get('compressed', False):
                            print(f"  📦 Compressed: {result['original_size_kb']}KB → {result['final_size_kb']}KB")
                    elif result.get('copied_without_annotations'):
                        print(f"  - No annotations - files copied")
                    elif result['annotated_pages']:
                        print(f"  ✗ Error: {result['error']}")
                    else:
                        print(f"  - No annotations found in any part")
                else:
                    # Single document
                    pdf_path = pdf_paths[0]
                    print(f"[{idx}/{total_groups}] Processing: {pdf_path.name}")
                    result = self.process_single_pdf(
                        pdf_path, output_dir, combine, organize_folders, 
                        False, None, copy_all_files, input_dir=input_dir
                    )
                    results.append(result)
                    
                    if result['success'] and result['annotated_pages']:
                        print(f"  ✓ Found {len(result['annotated_pages'])} annotated page(s)")
                        if result.get('compressed', False):
                            print(f"  📦 Compressed: {result['original_size_kb']}KB → {result['final_size_kb']}KB")
                    elif result.get('copied_without_annotations'):
                        print(f"  - No annotations - file copied")
                    elif result['annotated_pages']:
                        print(f"  ✗ Error: {result['error']}")
                    else:
                        print(f"  - No annotations found")
                
                print()
                idx += 1
        else:
            # Process each PDF individually (original behavior)
            for idx, pdf_path in enumerate(pdf_files, 1):
                print(f"[{idx}/{len(pdf_files)}] Processing: {pdf_path.name}")
                result = self.process_single_pdf(
                    pdf_path, output_dir, combine, organize_folders, 
                    False, None, copy_all_files, input_dir=input_dir
                )
                results.append(result)
                
                if result['success'] and result['annotated_pages']:
                    print(f"  ✓ Found {len(result['annotated_pages'])} annotated page(s)")
                    if result.get('compressed', False):
                        print(f"  📦 Compressed: {result['original_size_kb']}KB → {result['final_size_kb']}KB")
                elif result.get('copied_without_annotations'):
                    print(f"  - No annotations - file copied")
                elif result['annotated_pages']:
                    print(f"  ✗ Error: {result['error']}")
                else:
                    print(f"  - No annotations found")
                print()
        
        # Print summary with discrepancy tracking
        self.print_summary(results, initial_count, filtered_counts)
        
        return results
    
    def verify_output_sizes(self, output_dir: Path, max_size_kb: int = None) -> List[Dict]:
        """
        Scan all output files and find any that exceed the size limit.
        
        Args:
            output_dir: Directory to scan for output files
            max_size_kb: Maximum acceptable size in KB
            
        Returns:
            List of dictionaries with oversized file information
        """
        if max_size_kb is None:
            max_size_kb = self.max_file_size_kb
        
        oversized_files = []
        
        # Recursively find all PDFs in output directory
        for pdf_path in output_dir.rglob("*.pdf"):
            file_size_kb = pdf_path.stat().st_size / 1024
            
            if file_size_kb > max_size_kb:
                oversized_files.append({
                    'file': str(pdf_path.name),
                    'full_path': str(pdf_path),
                    'size_kb': int(file_size_kb),
                    'over_limit_kb': int(file_size_kb - max_size_kb)
                })
        
        return oversized_files
    
    def generate_excel_report(self, results: List[Dict], output_dir: Path, 
                             initial_count: int, filtered_counts: Dict, 
                             runtime_seconds: float, settings: Dict) -> Path:
        """
        Generate a comprehensive Excel report of the processing results.
        
        Args:
            results: List of result dictionaries from processing
            output_dir: Directory where report should be saved
            initial_count: Initial number of PDFs found
            filtered_counts: Dictionary of filtered file counts by type
            runtime_seconds: Total processing time in seconds
            settings: Dictionary of processing settings used
            
        Returns:
            Path to the generated Excel report
        """
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # === SHEET 1: SUMMARY ===
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20
        
        summary_data = [
            ["PDF Annotation Extraction Report", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["STATISTICS", ""],
            ["Initial PDFs found", initial_count],
            ["", ""],
            ["Filtering Results", ""],
        ]
        
        # Add filtered counts
        total_filtered = sum(filtered_counts.values())
        for filter_type, count in filtered_counts.items():
            if count > 0:
                summary_data.append([f"  - {filter_type}", count])
        summary_data.append(["Total filtered", total_filtered])
        summary_data.append(["", ""])
        
        # Processing results
        pdfs_processed = len(results)
        pdfs_with_annotations = sum(1 for r in results if r['annotated_pages'])
        pdfs_without_annotations = pdfs_processed - pdfs_with_annotations
        total_pages = sum(len(r['annotated_pages']) for r in results)
        files_compressed = sum(1 for r in results if r.get('compressed', False))
        
        summary_data.extend([
            ["PDFs processed", pdfs_processed],
            ["PDFs with annotations", pdfs_with_annotations],
            ["PDFs without annotations", pdfs_without_annotations],
            ["Total annotated pages", total_pages],
            ["Output files created", pdfs_with_annotations],
            ["Files compressed", files_compressed],
            ["", ""],
            ["DISCREPANCY", ""],
            ["Input files", initial_count],
            ["Output files", pdfs_with_annotations],
            ["Difference", initial_count - pdfs_with_annotations],
            ["Reduction", f"{((initial_count - pdfs_with_annotations) / initial_count * 100):.1f}%"],
            ["", ""],
            ["RUNTIME", ""],
            ["Total time", f"{runtime_seconds:.1f} seconds"],
            ["Time per file", f"{runtime_seconds / pdfs_processed:.2f} sec/file" if pdfs_processed > 0 else "N/A"],
            ["", ""],
            ["SETTINGS", ""],
            ["Combine pages", settings.get('combine_pages', 'N/A')],
            ["Combine parts", settings.get('combine_parts', 'N/A')],
            ["Keep parts separate", settings.get('keep_parts_separate', 'N/A')],
            ["Recursive scan", settings.get('recursive', 'N/A')],
            ["Organize folders", settings.get('organize_folders', 'N/A')],
            ["Ignore Glass Statements", settings.get('ignore_glass', 'N/A')],
            ["Ignore No Feature docs", settings.get('ignore_no_docs', 'N/A')],
            ["Ignore Support docs", settings.get('ignore_support', 'N/A')],
            ["Copy all files", settings.get('copy_all_files', 'N/A')],
        ])
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Style headers
        ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
        for row_num in [4, 7, 15, 21, 24]:
            cell = ws_summary[f'A{row_num}']
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # === SHEET 2: FILES WITH ANNOTATIONS ===
        ws_annotated = wb.create_sheet("Files With Annotations")
        headers = ["File Name", "Year", "Manufacturer", "Model", "Annotated Pages", "Page Count", 
                   "Output File", "Compressed", "Original Size (KB)", "Final Size (KB)"]
        ws_annotated.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_annotated.cell(1, col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 2
        for result in results:
            if result.get('annotated_pages'):
                filename = Path(result['file']).name if 'file' in result else result.get('file', 'N/A')
                
                # Try to parse year/manufacturer/model
                year, manufacturer, model = self.parse_filename_structure(Path(filename).stem)
                
                # Handle multi-part documents
                if 'parts' in result:
                    annotated_info = f"{len(result['annotated_pages'])} pages (multi-part)"
                else:
                    page_nums = [p+1 for p in result['annotated_pages']] if isinstance(result['annotated_pages'][0], int) else []
                    annotated_info = str(page_nums) if page_nums else f"{len(result['annotated_pages'])} pages"
                
                ws_annotated.append([
                    filename,
                    year or "N/A",
                    manufacturer or "Unknown",
                    model or "N/A",
                    annotated_info,
                    len(result['annotated_pages']),
                    Path(result['output_file']).name if result.get('output_file') else "N/A",
                    "Yes" if result.get('compressed') else "No",
                    result.get('original_size_kb', 0),
                    result.get('final_size_kb', 0)
                ])
                row_num += 1
        
        # Auto-size columns
        for col in ws_annotated.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_annotated.column_dimensions[column].width = adjusted_width
        
        # === SHEET 3: FILES WITHOUT ANNOTATIONS ===
        ws_no_annotations = wb.create_sheet("Files Without Annotations")
        headers_no_annot = ["File Name", "Year", "Manufacturer", "Model"]
        ws_no_annotations.append(headers_no_annot)
        
        for col_num, header in enumerate(headers_no_annot, 1):
            cell = ws_no_annotations.cell(1, col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for result in results:
            if not result.get('annotated_pages'):
                filename = Path(result['file']).name if 'file' in result else result.get('file', 'N/A')
                year, manufacturer, model = self.parse_filename_structure(Path(filename).stem)
                
                ws_no_annotations.append([
                    filename,
                    year or "N/A",
                    manufacturer or "Unknown",
                    model or "N/A"
                ])
        
        for col in ws_no_annotations.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_no_annotations.column_dimensions[column].width = adjusted_width
        
        # === SHEET 4: STATISTICS BY MANUFACTURER ===
        ws_stats = wb.create_sheet("Statistics")
        
        # Collect statistics
        manufacturer_stats = defaultdict(lambda: {'files': 0, 'with_annotations': 0, 'pages': 0})
        year_stats = defaultdict(int)
        model_stats = defaultdict(int)  # Now uses "Manufacturer Model" format
        
        for result in results:
            filename = Path(result['file']).name if 'file' in result else result.get('file', 'N/A')
            year, manufacturer, model = self.parse_filename_structure(Path(filename).stem)
            
            if manufacturer:
                manufacturer_stats[manufacturer]['files'] += 1
                if result.get('annotated_pages'):
                    manufacturer_stats[manufacturer]['with_annotations'] += 1
                    manufacturer_stats[manufacturer]['pages'] += len(result['annotated_pages'])
            
            if year:
                year_stats[year] += 1
            
            # Combine manufacturer and model for better statistics
            if manufacturer and model:
                full_model_name = f"{manufacturer} {model}"
                model_stats[full_model_name] += 1
        
        # Manufacturer table
        ws_stats.append(["STATISTICS BY MANUFACTURER"])
        ws_stats.append([])
        headers_stats = ["Manufacturer", "Total Files", "With Annotations", "Without Annotations", "Total Pages"]
        ws_stats.append(headers_stats)
        
        for col_num, header in enumerate(headers_stats, 1):
            cell = ws_stats.cell(3, col_num)
            cell.font = header_font
            cell.fill = header_fill
        
        for manufacturer in sorted(manufacturer_stats.keys()):
            stats = manufacturer_stats[manufacturer]
            ws_stats.append([
                manufacturer,
                stats['files'],
                stats['with_annotations'],
                stats['files'] - stats['with_annotations'],
                stats['pages']
            ])
        
        # Year table
        current_row = ws_stats.max_row + 3
        ws_stats.cell(current_row, 1, "STATISTICS BY YEAR")
        ws_stats.cell(current_row, 1).font = Font(bold=True, size=11)
        current_row += 2
        
        ws_stats.cell(current_row, 1, "Year")
        ws_stats.cell(current_row, 2, "File Count")
        for col in [1, 2]:
            ws_stats.cell(current_row, col).font = header_font
            ws_stats.cell(current_row, col).fill = header_fill
        current_row += 1
        
        for year in sorted(year_stats.keys()):
            ws_stats.cell(current_row, 1, year)
            ws_stats.cell(current_row, 2, year_stats[year])
            current_row += 1
        
        # Model table (now shows Manufacturer + Model)
        current_row += 2
        ws_stats.cell(current_row, 1, "TOP 30 MODELS (Manufacturer + Model)")
        ws_stats.cell(current_row, 1).font = Font(bold=True, size=11)
        current_row += 2
        
        ws_stats.cell(current_row, 1, "Manufacturer + Model")
        ws_stats.cell(current_row, 2, "File Count")
        for col in [1, 2]:
            ws_stats.cell(current_row, col).font = header_font
            ws_stats.cell(current_row, col).fill = header_fill
        current_row += 1
        
        # Show top 30 models (increased from 20 for better visibility)
        sorted_models = sorted(model_stats.items(), key=lambda x: x[1], reverse=True)[:30]
        for full_model_name, count in sorted_models:
            ws_stats.cell(current_row, 1, full_model_name)
            ws_stats.cell(current_row, 2, count)
            current_row += 1
        
        # Auto-size columns
        ws_stats.column_dimensions['A'].width = 45  # Wider for "Manufacturer Model"
        ws_stats.column_dimensions['B'].width = 15
        ws_stats.column_dimensions['C'].width = 25
        ws_stats.column_dimensions['D'].width = 25
        ws_stats.column_dimensions['E'].width = 25
        
        # === SHEET 5: COMPRESSION REPORT ===
        ws_compression = wb.create_sheet("Compression Report")
        headers_comp = ["File Name", "Original Size (KB)", "Final Size (KB)", "Reduction (KB)", "Reduction (%)"]
        ws_compression.append(headers_comp)
        
        for col_num, header in enumerate(headers_comp, 1):
            cell = ws_compression.cell(1, col_num)
            cell.font = header_font
            cell.fill = header_fill
        
        for result in results:
            if result.get('compressed'):
                filename = Path(result['file']).name if 'file' in result else result.get('file', 'N/A')
                if result.get('output_file'):
                    filename = Path(result['output_file']).name
                
                orig = result.get('original_size_kb', 0)
                final = result.get('final_size_kb', 0)
                reduction_kb = orig - final
                reduction_pct = (reduction_kb / orig * 100) if orig > 0 else 0
                
                ws_compression.append([
                    filename,
                    orig,
                    final,
                    reduction_kb,
                    f"{reduction_pct:.1f}%"
                ])
        
        for col in ws_compression.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_compression.column_dimensions[column].width = adjusted_width
        
        # === SHEET 6: OVERSIZED FILES (POST-PROCESSING VERIFICATION) ===
        print("\n🔍 Verifying output file sizes...")
        oversized_files = self.verify_output_sizes(output_dir)
        
        ws_oversized = wb.create_sheet("Oversized Files")
        ws_oversized.append(["POST-PROCESSING VERIFICATION - FILES OVER 1400KB"])
        ws_oversized.append([])
        
        if oversized_files:
            headers_oversized = ["File Name", "Size (KB)", "Over Limit (KB)", "Full Path"]
            ws_oversized.append(headers_oversized)
            
            for col_num, header in enumerate(headers_oversized, 1):
                cell = ws_oversized.cell(3, col_num)
                cell.font = header_font
                cell.fill = header_fill
            
            # Sort by size descending
            oversized_files.sort(key=lambda x: x['size_kb'], reverse=True)
            
            for file_info in oversized_files:
                ws_oversized.append([
                    file_info['file'],
                    file_info['size_kb'],
                    file_info['over_limit_kb'],
                    file_info['full_path']
                ])
            
            # Style the header
            ws_oversized['A1'].font = Font(bold=True, size=12, color="FF0000")
            
            print(f"  ⚠️ Found {len(oversized_files)} file(s) still over 1400KB")
        else:
            ws_oversized.append(["✅ All output files are under 1400KB!"])
            ws_oversized['A3'].font = Font(bold=True, size=12, color="00AA00")
            print(f"  ✅ All output files are under 1400KB!")
        
        # Auto-size columns
        for col in ws_oversized.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_oversized.column_dimensions[column].width = adjusted_width
        
        # Save the report
        try:
            # Ensure output directory exists
            output_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = output_dir / f"Extraction_Report_{timestamp}.xlsx"
            
            # Save the workbook
            wb.save(str(report_path))
            
            # Verify file was created
            if report_path.exists():
                file_size_kb = report_path.stat().st_size / 1024
                print(f"\n📊 Excel report generated: {report_path.name} ({file_size_kb:.1f} KB)")
                return report_path
            else:
                print(f"\n⚠️ Warning: Excel report save completed but file not found at {report_path}")
                return None
                
        except PermissionError as e:
            print(f"\n❌ Permission denied when saving Excel report: {str(e)}")
            print(f"   Make sure the file isn't open in Excel and you have write permissions.")
            return None
        except Exception as e:
            print(f"\n❌ Error saving Excel report: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def print_summary(self, results: List[Dict], initial_count: int = 0, filtered_counts: Dict = None):
        """
        Print a summary of processing results with discrepancy tracking.
        
        Args:
            results: List of result dictionaries from processing
            initial_count: Initial number of PDF files found before filtering
            filtered_counts: Dictionary with counts of filtered files by type
        """
        total_processed = len(results)
        pdfs_with_annotations = sum(1 for r in results if r['annotated_pages'])
        total_annotated_pages = sum(len(r['annotated_pages']) for r in results)
        output_files_created = sum(1 for r in results if r['success'] and r['annotated_pages'])
        
        print("=" * 70)
        print("PROCESSING SUMMARY")
        print("=" * 70)
        
        if initial_count > 0:
            print(f"Initial PDFs found:           {initial_count}")
            
            if filtered_counts:
                total_filtered = sum(filtered_counts.values())
                print(f"\nFiltered out:")
                for filter_type, count in filtered_counts.items():
                    if count > 0:
                        print(f"  - {filter_type}: {count}")
                print(f"  Total filtered:             {total_filtered}")
            
            print(f"\nAfter filtering:              {total_processed}")
        else:
            print(f"Total PDFs processed:         {total_processed}")
        
        print(f"\nPDFs with annotations:        {pdfs_with_annotations}")
        print(f"PDFs without annotations:     {total_processed - pdfs_with_annotations}")
        print(f"Total annotated pages:        {total_annotated_pages}")
        print(f"\nOutput files created:         {output_files_created}")
        
        if initial_count > 0:
            discrepancy = initial_count - output_files_created
            reduction_pct = (discrepancy / initial_count * 100) if initial_count > 0 else 0
            print(f"\n{'='*70}")
            print(f"DISCREPANCY ANALYSIS")
            print(f"{'='*70}")
            print(f"Input files:                  {initial_count}")
            print(f"Output files:                 {output_files_created}")
            print(f"Difference:                   {discrepancy} ({reduction_pct:.1f}% reduction)")
        
        print("=" * 70)


def main():
    """
    Example usage / testing function
    """
    print("PDF Annotation Extractor - Core Module")
    print("=" * 60)
    
    # Example: Process a folder
    extractor = PDFAnnotationExtractor()
    
    # You can modify these paths for testing
    input_folder = Path(r"C:\Users\zgilk\Desktop Test and programs\HL try\test_pdfs")
    output_folder = Path(r"C:\Users\zgilk\Desktop Test and programs\HL try\output")
    
    if not input_folder.exists():
        print(f"\nTest folder not found: {input_folder}")
        print("Please create a 'test_pdfs' folder and add some PDFs with annotations to test.")
        print("\nOr modify the paths in the main() function to point to your PDF files.")
        return
    
    # Process all PDFs in the folder
    results = extractor.scan_folder(
        input_dir=input_folder,
        output_dir=output_folder,
        combine=True,  # Combine annotated pages into one PDF per source
        recursive=True  # Scan subfolders
    )
    
    # Print summary
    extractor.print_summary(results)


if __name__ == "__main__":
    main()

