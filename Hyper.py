import sys
from PyQt5.QtWidgets import (QApplication, QDialog, QPlainTextEdit, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLabel, QLineEdit, QPushButton,
                             QTreeWidget, QTreeWidgetItem, QMessageBox, QFileDialog, QCheckBox, QScrollArea, QListWidget, QProgressBar)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt,pyqtSignal,QThread, QTimer
from threading import Thread
import subprocess
import signal
import psutil
from time import sleep
import datetime
import time
import os
import logging
import re

######################################################################################     Username and Password    ######################################################################################

class LoginDialog(QDialog):
    def __init__(self, *, max_attempts=5, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sign in to Hyper")
        self.max_attempts = max_attempts
        self.attempts = 0

        # UI
         # ⬅ Increase window size
        self.setFixedSize(250, 150)  # was ~half this before
        v = QVBoxLayout(self)

        row_user = QHBoxLayout()
        row_user.addWidget(QLabel("Username:"))
        self.user_edit = QLineEdit()
        self.user_edit.setPlaceholderText("Enter username")
        row_user.addWidget(self.user_edit)
        v.addLayout(row_user)

        row_pass = QHBoxLayout()
        row_pass.addWidget(QLabel("Password:"))
        self.pass_edit = QLineEdit()
        self.pass_edit.setEchoMode(QLineEdit.Password)
        self.pass_edit.setPlaceholderText("Enter password")
        row_pass.addWidget(self.pass_edit)
        v.addLayout(row_pass)

        btns = QHBoxLayout()
        self.ok_btn = QPushButton("Sign In")
        self.cancel_btn = QPushButton("Cancel")
        btns.addWidget(self.ok_btn)
        btns.addWidget(self.cancel_btn)
        v.addLayout(btns)

        self.ok_btn.clicked.connect(self.try_login)
        self.cancel_btn.clicked.connect(self.reject)

        # optional: match your dark/light styling (simple, non-invasive)
        self.setStyleSheet("""
            QDialog { background: #2b2b2b; color: white; }
            QLabel  { color: white; }
            QLineEdit { background: #3a3a3a; color: white; border: 1px solid #555; border-radius: 4px; padding: 4px; }
            QPushButton { padding: 6px 12px; }
        """)

    # Replace this with your real auth check if needed
    def _validate(self, username: str, password: str) -> bool:
        # EXAMPLE ONLY — change to env vars, config, or your own logic.
        # import os; return (username == os.getenv("HYPER_USER") and password == os.getenv("HYPER_PASS"))
        return (username == "Dromero221" and password == "Hyperactive221")

    def try_login(self):
        u = self.user_edit.text().strip()
        p = self.pass_edit.text()
        if self._validate(u, p):
            self.accept()
            return

        self.attempts += 1
        remaining = self.max_attempts - self.attempts
        if remaining <= 0:
            QMessageBox.critical(self, "Access denied", "Too many failed attempts. Closing.")
            self.reject()  # caller will exit the app
            return

        QMessageBox.warning(self, "Invalid credentials",
                            f"Username or password is incorrect.\nAttempts left: {remaining}")
        self.pass_edit.clear()
        self.pass_edit.setFocus()

######################################################################################     Terminal & Certain GUI (Buttons and Switches) Code    ######################################################################################

def format_runtime(seconds):
    h, rem = divmod(int(seconds), 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}"

# ── configure a “Logs” folder in Documents ──
LOG_DIR = os.path.join(os.path.expanduser("~"), "Documents", "Hyper Logs")
os.makedirs(LOG_DIR, exist_ok=True)

# ── log filename with timestamp ──
now = datetime.datetime.now()   # ← module.datetime.now()
log_file = os.path.join(
    LOG_DIR,
    now.strftime("Hyper_Log_%m_%d_%Y_%H-%M-%S.log")
)

# ── basicConfig writes to both file and console ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)

#Adds Terminal infoormation
class WorkerThread(QThread):
    finished_signal = pyqtSignal(str, bool)
    output_signal   = pyqtSignal(str)
    progress_signal = pyqtSignal(int)  # ← added for live percentage updates

    def __init__(self, command: list[str], manufacturer: str, parent=None):
        super().__init__(parent)
        self.command      = command
        self.manufacturer = manufacturer
        self.process      = None

    def run(self):
        # ── Prepare env ──
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"

        if os.name == "nt":
            creationflags = subprocess.CREATE_NEW_PROCESS_GROUP
            preexec_fn    = None
        else:
            creationflags = 0
            preexec_fn    = os.setsid

        self.process = subprocess.Popen(
            self.command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
            universal_newlines=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            creationflags=creationflags,
            preexec_fn=preexec_fn
        )

        # Estimate total line count (for basic progress calculation)
        estimated_total_lines = 300  # ← adjust this if you want more accuracy
        processed_lines = 0

        # ── Stream stdout ──
        for line in iter(self.process.stdout.readline, ""):
            if not line:
                break
            self.output_signal.emit(line.rstrip("\n"))
            processed_lines += 1

            # Emit progress signal
            percent = min(100, int((processed_lines / estimated_total_lines) * 100))
            self.progress_signal.emit(percent)

        self.process.stdout.close()

        # ── Wait & then stream stderr if error ──
        ret = self.process.wait()
        success = (ret == 0)
        if not success:
            for e in iter(self.process.stderr.readline, ""):
                if not e:
                    break
                self.output_signal.emit(e.rstrip("\n"))
        self.process.stderr.close()

        # Ensure final progress is 100%
        self.progress_signal.emit(100)

        self.finished_signal.emit(self.manufacturer, success)

class TerminalDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Terminal Output")
        self.setGeometry(100, 100, 600, 400)

        self.layout = QVBoxLayout()
        self.terminal_output = QPlainTextEdit()
        self.terminal_output.setReadOnly(True)
        self.layout.addWidget(self.terminal_output)

        self.setLayout(self.layout)

    def append_output(self, text):
        self.terminal_output.appendPlainText(text)
        self.terminal_output.ensureCursorVisible()

class ModeSwitch(QCheckBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        # no visible text on the switch itself
        self.setText("")
        self.setFixedSize(50, 25)
        self.setStyleSheet("""
            QCheckBox {
                background-color: #888;
                border-radius: 12px;
            }
            QCheckBox::indicator {
                width: 21px; height: 21px;
                border-radius: 10px;
                background-color: white;
                margin: 2px;
            }
            QCheckBox::indicator:checked {
                margin-left: 27px;
            }
            QCheckBox:disabled {
                background-color: #555;
            }               
        """)

class CustomButton(QPushButton):
    def __init__(self, text, color, parent=None):
        super().__init__(text, parent)
        self.color = color
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 10px;
                font-size: 16px;
                border-radius: 5px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
        """)

    def darken_color(self, color):
        hex_color = color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        darkened_rgb = tuple(max(0, min(255, int(c * 0.85))) for c in rgb)
        return '#{:02x}{:02x}{:02x}'.format(*darkened_rgb)

class ToggleSwitch(QCheckBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setText("Dark Mode")
        self.setStyleSheet("""
            QCheckBox {
                font-size: 16px;
                color: white;
                background-color: #2e2e2e;
                border: 1px solid #555555;
                border-radius: 15px;
                padding: 10px;
            }
            QCheckBox::indicator {
                width: 0px;
                height: 0px;
            }
        """)
        self.setFixedSize(120, 40)
        self.setTristate(False)
        self.stateChanged.connect(self.updateAppearance)

    def updateAppearance(self, state):
        if self.isChecked():
            self.setText("Light Mode")
            self.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    color: black;
                    background-color: #f0f0f0;
                    border: 1px solid #cccccc;
                    border-radius: 15px;
                    padding: 10px;
                }
                QCheckBox::indicator {
                    width: 0px;
                    height: 0px;
                }
            """)
        else:
            self.setText("Dark Mode")
            self.setStyleSheet("""
                QCheckBox {
                    font-size: 16px;
                    color: white;
                    background-color: #2e2e2e;
                    border: 1px solid #555555;
                    border-radius: 15px;
                    padding: 10px;
                }
                QCheckBox::indicator {
                    width: 0px;
                    height: 0px;
                }
            """)
        self.parent().toggle_theme()
        
######################################################################################     Main Application Code    ######################################################################################

class SeleniumAutomationApp(QWidget):
    def __init__(self):
        super().__init__()
        
        # 🆕 hide CM progress lines in terminal output
        self.hide_cm_progress_in_terminal = True
        
        self.initUI()
        self.terminal = None           
        self.excel_paths = []
                                   ########################################################     ADAS SI Links     ########################################################
        self.manufacturer_links = {
            "Acura": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgO44QnAOBVHs5Xl_VaX_BcBfLxg-RRGdRqxjXSio4-ylA?e=dY0s2f (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EpSbfndQRGFLqDU7ir6VN2wBDoSKqqUeO8aNXVymXBa2iQ?e=67ytxd (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtpcOnBrMTtMmLUlEPAFrRABNdyonEKJPGZDlEv7XKzfPg?e=UeqTNv (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Alfa Romeo": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgXpZi9fqF9DvSbpl561hQgBWBpkhOV3FUjJg8spjJ_TIA?e=JbUro2 (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgZFflWuydBNqQkLkdkm5lABjtUKWQb-KO91IEZe77GtzQ?e=zjnMfl (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EqwwF1Cbp_FLtPCdbgBOj1oBTCpQ8tW5KLG47slxvMgS9g?e=OKddLK (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Audi": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ev6GIUc3RYJFsfApBMcIJ74BRSQctvC9dbIFbNAzMH1UGQ?e=aQ60uo (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvETjo6xweBImS9xQUSdsTYBfChMWtrp-ePOKxFSFxPlLA?e=KvAEhj (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EncDDXP7im1LhmraMyjVpcwBHr5wW-OU5P0KksmwdreqTw?e=hYcG4j (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "BMW": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtbdmoWEk3JPgm9fCnJds30Bz1ZBmSkLvGB4ycGspXnfrQ?e=fkT6We (2012 - 2016)",# Documents (2012 - 2016) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/El3B4OwoJ-BGlSMZxojt-L8BhnfT2YxtYMNoSsOiR_jHOA?e=O9ivgG (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgQoL-kCkbJHp-ZJ6Se8s_4BQ7hdonoPaoelYmv1kxUN7Q?e=pjsUVE (2022 - 2026)" # Documents (2022 - 2026) # Broken Still/ Not working / N/A
            ],
            "Brightdrop": [
                "https://sharepoint.com/.../Brightdrop (2012 - 2016)",# Documents (2012 - 2016)
                "https://sharepoint.com/.../Brightdrop (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ejtmfro-bUlGsYsRnftPTg8BjwJctxSV3MUYckWa2kvx2g?e=0H47Pr (2022 - 2026)" # Documents (2022 - 2026) 
            ],
            "Buick": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EukXmI2Ic7tDmOSkwaJm_egBSZNfRVZKhPyrlvUNxRmmDQ?e=wVspWx (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvHhbdsOkGNBkXrlCVqbxNgBcr5Edjle4n5bXduaPYVbzg?e=y0cWg1 (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EkLzgSZZYcRKpqwUoYs7RjsB1o_7MvCWHfjzqBwoB99TmA?e=9AjDir (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Cadillac": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjxE4DfO485BsvYkXcAqaqIB9chwARjfMtxL2LaKCBgLHQ?e=qJ8KsD (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmSzlsLEldVNnnBICuBXoA0Bidd3gZlai5xHri8-aUylXA?e=TKDNUF (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgU6G2rJ2ZVFl4YSWiXTMfIBBVKf4h9MPmHEbua24KhjPQ?e=zUwgO3 (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Chevrolet": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Evlw3FgdLyJDnX5rh-vwzbgBFzvkmTx0dPjqMhLfk3lnDg?e=AVBhEy (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EkJ92YFGDb5Fn4Pxqk6sDEsBUYNxfZtX8SsLqqBBMh9hRg?e=dHHULo (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EhHAJ0VuOqhDg2y1taiAyWgBwz_WsFn36AH86HZsh6geFw?e=LvFP2c (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Chrysler": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Esq1Q-DlT91KnlH-YSNr_xQBw4lTIUYsnrSqe5q9_yOtbg?e=9MG2IG (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EhkmJzElYeRMhTGeqNLhFIQBc7BmXae5nxlK0ag8oQ5YAQ?e=PRiu5f (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/El_qlDrwpmxPspw5MeRekmwBvw37wNf3TCHI97kZk82fXw?e=crb5Fi (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Dodge": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EhYhxdkjIP1EvM94_E9DikQBpmOqNmo49crObJefjtg3lw?e=ML7vkg (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EibUuEGr2qpLpxZ6pfRl-0EBxoSDgeqadOO9xOdQTPfRaw?e=lU4Whi (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Er2P0SGtxQhJnrVfbUGRcWEBfN0oAiCNs4XHaRU_CazDRA?e=QN6yEE (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Fiat": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvOVwi_fWRZNvffbECJ8GowBd8TI4pI4FvF-_5rPT5lqIw?e=Hc4T0m (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/ElrLCEUKvXlBhw97bY8jH88BDZU__O2hcWBPPlxviZPXyg?e=jyhySl (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/El1yZl0RD2RNgnQC0Y3Mo_0BJlcr5nt2IMMpHyGIRryxxg?e=gBCfeZ (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Ford": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Evlrka8XYfRDsyePMcgTPIQBdW2UfqnJZFDwpPCbHYDIoQ?e=Kc61El (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtQvr5IOAClHl71EvTUwxDEBjDs4kdyeRqI6zc-FyapkHw?e=eaZD9e (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjxutzqDXNhMsptbq4Lbkr8BKoHKaZ3ecGoDUEvnrMw_UA?e=2AiPNn (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Genesis": [
                "https://sharepoint.com/.../Genesis (2012 - 2016)",# Documents (2012 - 2016) # N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EqVqYvg1KghPtxkwu9F8XkMBOvYeuePjvh36wU-2etdxoA?e=EA3XNw (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eh92TYnyNkJKtgcnciCVA5AB-znIckE4I2REf614K5qkeA?e=DvYcwl (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "GMC": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EkjOrZETVzlGjZf137xhCMwBIfSvB8WOl-Z2QIK2NMHiAw?e=RRmkow (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmoSycYqjpxBtqvyIq3SMAABObwWOGjRPjrPz25YButiCQ?e=4QcT9S (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EoS9WZxD6aNDqYGqs91ctiwB2O5NG-oUS7E3qxF721q6Ww?e=767say (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Honda": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EitBF4UEMARKiR62xfVZysMB_nLxLcbGOV-2PneLZ_V8iA?e=QbVIhF (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EqoSh-ImVMNMsP3tX91XYisBGpiZvGYPtADxEYPpQtYAxQ?e=xr0pgS (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EqURKybJq41DohtvBstKxRQBXudDBeAO69H6rrPakChX0A?e=vrmHT5 (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Hyundai": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eg3ZzULtMIhIj5NNuGeOkGABilR76pGui806f65Pf1011Q?e=FpJZQn (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EoZcJNzkv05Dgt8pn4VJrMoBWPjbD7_Kn3noM__Js4642Q?e=X2PhwA (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjW3ucsaxHNDoR4lJLHZBKYBZj6o0sjJCqZr5AobXkQvJw?e=WVo0Ox (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Infiniti": [ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvosfHnnjlVLiZis0xJGVYgBjWTo-aKfmXhqFk5cqzawQg?e=ZLe6c4 (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Erigcso1ekZLk9gg_0UXW-oBcclm8HEw5nCc-P5sJ4RGJw?e=FLgVgy (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvtewQN1PAFLm82hoZT9xxABxs4wiTRt4XgOUR5OddmhJA?e=Q6gIVV (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Jaguar": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EoogbJM88UFDu4deUMN5si0B0gdGxYrmWOeRK3h-QD_6Ow?e=6wgNpU (2012 - 2016)",# Documents (2012 - 2016) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EuGXMEEqNAlOg0gPzivrS-oBYMLOAhQW283EgmGUAM4Yvg?e=3E9u8g (2017 - 2021)",# Documents (2017 - 2021) ✅ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eo7QEMg3xbNMmfKHuaLmWqYB_5DTu_sZz_0PPJbH0c9mrA?e=jsNVzT (2022 - 2026)" # Documents (2022 - 2026) ✅ 
            ],
            "Jeep": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EihFn2Un6D9PkOKJTUV3K7kBYX1pXoQ8o-VHaA9ak_ODdQ?e=GCvvR1 (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvhDLn4gv6RGt9vnOZdJPvwBXQ23V58HrY7z9wVa49WNAg?e=lcjhT2 (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Etm0G0lvmDBEu2wipvxWH94B7r3US49JMvEToHzYOGsV9g?e=Qxd2TI (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Kia": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EumE-fcIi7RPuGZ8sdsnH8MBXoEk7WI6u6u7OaOWx30g7A?e=eevS7d (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EsuP8lfXyDpJnkChBGpfpuEBj_OsAP1hhZlLREBadi8MZQ?e=7MN3N7 (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ev5VL24JpYpGlBjXJ_ihw28BAxPK5cMsWJiqgfFl7BxMMw?e=VwBuYn (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Land Rover": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/ElW0x5ldlb1Avj9TWwFDlF0BVA5-JtzIRxSY0QPWStrh3Q?e=S9evFx (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Evb37uS2Q4hKv0YxUYSquboBDh1nO0w0uD_BbmbhZpU1uQ?e=fKyMH3 (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvLaGHhPq3dLlKaDdlrHOZcBW7QVm97G1X-v6t4oUrsAIg?e=q2vjpX (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Lexus": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjnILgBMvUxGvU20USLPkTABX64H95a0lWHaiMmUCX3jog?e=rSrplp (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtdglpgetaNLnHIYr0SSXjEBTRWbe_LqlvR3v3qOw4vkog?e=6iUHht (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EhO2AfBnrE1MvNEpCsWIvOkBBTvrKzAuJvgF7sDDqZ-n0g?e=sCTPFn (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Lincoln": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EkbpJGSM-fJDlKLc1E1OvjMBF3wTKoxbJcuXEKTmv9BGOg?e=BUjcG4 (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/En0bZeut7NlGhASVRGGFdQ0BAhk0fFhuxIvSMAxtdnnzVA?e=YUowas (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eq-QDVfyLw1KpXtjOJF2S70BnMfhbj6D3wWHXxyVyH47zw?e=Rbm2P4 (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Mazda": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EpmyG0s2coNDrCioycqb5w8Bn1PY6DL7IYV6mjFlLFvqxw?e=TqSwhK (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EpbL3bu1Wa5HidZ_p3AyrCUBMO_hrWhTRjqeI-KgZF2c1w?e=zNLeIO (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EhtJ9dLTFjJLo85siCroJ7wBAoemHZwbIJlqPLVJca1Eug?e=enxPbd (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Mercedes": [
                "https://sharepoint.com/.../Mercedes (2012 - 2016)",# Documents (2012 - 2016)
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmxPr72nQWNAq8NgnEHMf9MBOC0RoxU_7IywgY8ytvRt7Q?e=PiesCv (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A ✅ Good
                "https://sharepoint.com/.../Mercedes (2022 - 2026)" # Documents (2022 - 2026)
            ],
            "Mini": [ 
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Etb0-TpXIMhCvSSE-OYz7xgBJ_zOQNbn1XPV0vzMEq63WQ?e=4BYfrx (2012 - 2016)",# Documents (2012 - 2016) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/ElbFUVyL5eRFrgza6_3YVk4B6jiFtnQnwYe3xY1XtGKRKw?e=TvbNx0 (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmNVWOLcHU5Nol_C8GXcL6QB6l3ltuaXHziyOFpTerXHzg?e=eciiYe (2022 - 2026)" # Documents (2022 - 2026) # Broken Still/ Not working / N/A
            ],
            "Mitsubishi": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/En2VOaIug3lKkVj--jyrGJABKv8lAi_Gq_SQk18yGPRrUA?e=gY43PS (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmcmQLiC-wFKhLdCpOFuZ00Bkzy-FFDZpbFeU1ZNM9D7Dg?e=axkbol (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ej-s4pi87ghDm15GIFu3FNQB96Cn1NLy2yhN1gvY5VKwYA?e=16FZbs (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Nissan": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eoj9oUiOqkBBn35lOK4CcmIBcMss_3rdvp7-juMdhv1rBw?e=U2lhyJ (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EnvxnVDdNrVNpgD1nvc09roBuYgnddQKo4eaO3XB6TakfQ?e=DTyUu6 (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjBOqNeUWC9Hj5tzNLDbGtwBSJAsXCU5NPebVURmaWKu_A?e=9dImx0 (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Porsche": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/ErbYi-naXahCuje8DUaVePEBIjDoKLw-AeL17DRlu0bC4w?e=2YwZJA (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EpldxYMij_NEoCY8LC6lu0cBrq0nTshjjK1TXmLYLl2z1Q?e=XgRNjI (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EgbYyvLqjcRJmBNBDE0pE70BdNmCbiXk8SQj7ssdDD2zUQ?e=kWmSJA (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Ram": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EjcqFA69GKROs7A1sL9stXwBrGcMe4mPWwtUV87-05jRhg?e=cDiNhH (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EmQO0oIVqfZMrA5C4wAQN28B7MnYNwAnMBG7St6VVIOlvg?e=PcxATl (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtkZMs1wwuNHhnN6zEAl8S8BbwV16Uf4Zfs-41eHTvdaDQ?e=MN4iPe (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Rolls Royce": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Es-YCmWbEzpBpWVgADAhYvgBXsbtsFK_8fe7hsbXyu0hfA?e=RgHl02 Royce (2012 - 2016)",# Documents (2012 - 2016) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EvkK0QMte9pDsnfXSDrhp-0BKCTP8S1FR6nfWLIZcJasVw?e=LWVPUR Royce (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eh4kFN7MzChHtKNfA_l8tSIBnIOznyqth45hBiQUAvxKew?e=H7GFGr Royce (2022 - 2026)" # Documents (2022 - 2026) # Broken Still/ Not working / N/A
            ],
            "Subaru": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EkLX0r88_ndCvsPURYQW4V4BLC2L11xJKFHTXigbg_aTBA?e=fkbTp7 (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EtILdNwpUVdMiU2HbHtwP78BrHmFObiNPetyZgOE82a0sg?e=clkNdu (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EsyhWqSCafFFgZ3rwqpv3H0BWVZXfyouqpgRSu7EhaT5oQ?e=roMgVf (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Tesla": [
                "https://sharepoint.com/.../Tesla (2012 - 2016)",# Documents (2012 - 2016) # Broken Still/ Not working / N/A
                "https://sharepoint.com/.../Tesla (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                "https://sharepoint.com/.../Tesla (2022 - 2026)" # Documents (2022 - 2026) # Broken Still/ Not working / N/A
            ],
            "Toyota": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EqAK8wENQ9xGq2Ko__kiMVABQz53OcEuNcopFxs5y1sVrg?e=WFaRdO (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EuaYS8YmVRhIppGqxWku_88BFKR2PA4j9XpK1TYYssXkKg?e=SWEEwK (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Eku5IrOt_8NFqCB2csrTdAkBS6zmu5d6wQuPPp986JnSlg?e=LpH6p7 (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Volkswagen": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Et3ZwE9XuT1Equj5v_YqDYQBZzsoIXaotEFT5X1PrahiNg?e=gBHl82 (2012 - 2016)",# Documents (2012 - 2016) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EuE6XexO-p9NmkHXm3tj0ZYBEnh6j_Efe5XTs-QdWtYj0A?e=DdU6PN (2017 - 2021)",# Documents (2017 - 2021) ✅
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EnKcjlMGgI1LnTHY8Edagi8BDxCWG8YCaiE-VoWMR9XkKg?e=ceaZaY (2022 - 2026)" # Documents (2022 - 2026) ✅
            ],
            "Volvo": [
                #"https://sharepoint.com/.../Volvo (2012 - 2016)",# Documents (2012 - 2016) # Broken Still/ Not working / N/A
                #"https://sharepoint.com/.../Volvo (2017 - 2021)",# Documents (2017 - 2021) # Broken Still/ Not working / N/A
                #"https://sharepoint.com/.../Volvo (2022 - 2026)" # Documents (2022 - 2026) # Broken Still/ Not working / N/A
            ]
        }
                             ########################################################     Repair SI Links     ########################################################
        self.repair_links = {
            "Acura": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EihKJ1TNSRdNpsQv8r32FFsB3jkwS6DfqW4Mcff4NrOr6A?e=qVTz2q (2012 - 2016)",# Documents (2012 - 2016)
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ek3lsuYY8cZEsJwsES_Q1KwBhX3TTKcKhB5C5mdXUNReDQ?e=GEE8Mv (2017 - 2021)",# Documents (2017 - 2021)
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Ei1SfDwSRlpKpSWMZ__bhdsB4VoeqkmzqFUDdb0anGPnbw?e=VJC52s (2022 - 2026)" # Documents (2022 - 2026)
            ],
            "Alfa Romeo": [
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EpvFWGXv3OxPjjcMd0R04yYBrTKxJGtCsLHzCtrEwtEDcg?e=x8pBMS (2012 - 2016)",# Documents (2012 - 2016)
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/EsJDtk8Y4ENCrsxhi5WIqGsBHmPOc3w6Nt_IBlZku_G8EQ?e=mm14mg (2017 - 2021)",# Documents (2017 - 2021)
                "https://calibercollision.sharepoint.com/:f:/s/O365-DepartmentofInformationSoloutions/Emy9X9R3Ug9AlQB0D8C50KgBcinQhqi6FHjNMScIm8ZVuw?e=8VnXv2 (2022 - 2026)" # Documents (2022 - 2026)
            ],
            "Audi": [
                "https://sharepoint.com/.../Audi (2012 - 2016)",
                "https://sharepoint.com/.../Audi (2017 - 2021)",
                "https://sharepoint.com/.../Audi (2022 - 2026)"
            ],
            "BMW": [
                "https://sharepoint.com/.../BMW (2012 - 2016)",
                "https://sharepoint.com/.../BMW (2017 - 2021)",
                "https://sharepoint.com/.../BMW (2022 - 2026)"
            ],
            "Brightdrop": [
                "https://sharepoint.com/.../Brightdrop (2012 - 2016)",
                "https://sharepoint.com/.../Brightdrop (2017 - 2021)",
                "https://sharepoint.com/.../Brightdrop (2022 - 2026)"
            ],
            "Buick": [
                "https://sharepoint.com/.../Buick (2012 - 2016)",
                "https://sharepoint.com/.../Buick (2017 - 2021)",
                "https://sharepoint.com/.../Buick (2022 - 2026)"
            ],
            "Cadillac": [
                "https://sharepoint.com/.../Cadillac (2012 - 2016)",
                "https://sharepoint.com/.../Cadillac (2017 - 2021)",
                "https://sharepoint.com/.../Cadillac (2022 - 2026)"
            ],
            "Chevrolet": [
                "https://sharepoint.com/.../Chevrolet (2012 - 2016)",
                "https://sharepoint.com/.../Chevrolet (2017 - 2021)",
                "https://sharepoint.com/.../Chevrolet (2022 - 2026)"
            ],
            "Chrysler": [
                "https://sharepoint.com/.../Chrysler (2012 - 2016)",
                "https://sharepoint.com/.../Chrysler (2017 - 2021)",
                "https://sharepoint.com/.../Chrysler (2022 - 2026)"
            ],
            "Dodge": [
                "https://sharepoint.com/.../Dodge (2012 - 2016)",
                "https://sharepoint.com/.../Dodge (2017 - 2021)",
                "https://sharepoint.com/.../Dodge (2022 - 2026)"
            ],
            "Fiat": [
                "https://sharepoint.com/.../Fiat (2012 - 2016)",
                "https://sharepoint.com/.../Fiat (2017 - 2021)",
                "https://sharepoint.com/.../Fiat (2022 - 2026)"
            ],
            "Ford": [
                "https://sharepoint.com/.../Ford (2012 - 2016)",
                "https://sharepoint.com/.../Ford (2017 - 2021)",
                "https://sharepoint.com/.../Ford (2022 - 2026)"
            ],
            "Genesis": [
                "https://sharepoint.com/.../Genesis (2012 - 2016)",
                "https://sharepoint.com/.../Genesis (2017 - 2021)",
                "https://sharepoint.com/.../Genesis (2022 - 2026)"
            ],
            "GMC": [
                "https://sharepoint.com/.../GMC (2012 - 2016)",
                "https://sharepoint.com/.../GMC (2017 - 2021)",
                "https://sharepoint.com/.../GMC (2022 - 2026)"
            ],
            "Honda": [
                "https://sharepoint.com/.../Honda (2012 - 2016)",
                "https://sharepoint.com/.../Honda (2017 - 2021)",
                "https://sharepoint.com/.../Honda (2022 - 2026)"
            ],
            "Hyundai": [
                "https://sharepoint.com/.../Hyundai (2012 - 2016)",
                "https://sharepoint.com/.../Hyundai (2017 - 2021)",
                "https://sharepoint.com/.../Hyundai (2022 - 2026)"
            ],
            "Infiniti": [
                "https://sharepoint.com/.../Infiniti (2012 - 2016)",
                "https://sharepoint.com/.../Infiniti (2017 - 2021)",
                "https://sharepoint.com/.../Infiniti (2022 - 2026)"
            ],
            "Jaguar": [
                "https://sharepoint.com/.../Jaguar (2012 - 2016)",
                "https://sharepoint.com/.../Jaguar (2017 - 2021)",
                "https://sharepoint.com/.../Jaguar (2022 - 2026)"
            ],
            "Jeep": [
                "https://sharepoint.com/.../Jeep (2012 - 2016)",
                "https://sharepoint.com/.../Jeep (2017 - 2021)",
                "https://sharepoint.com/.../Jeep (2022 - 2026)"
            ],
            "Kia": [
                "https://sharepoint.com/.../Kia (2012 - 2016)",
                "https://sharepoint.com/.../Kia (2017 - 2021)",
                "https://sharepoint.com/.../Kia (2022 - 2026)"
            ],
            "Land Rover": [
                "https://sharepoint.com/.../Land Rover (2012 - 2016)",
                "https://sharepoint.com/.../Land Rover (2017 - 2021)",
                "https://sharepoint.com/.../Land Rover (2022 - 2026)"
            ],
            "Lexus": [
                "https://sharepoint.com/.../Lexus (2012 - 2016)",
                "https://sharepoint.com/.../Lexus (2017 - 2021)",
                "https://sharepoint.com/.../Lexus (2022 - 2026)"
            ],
            "Lincoln": [
                "https://sharepoint.com/.../Lincoln (2012 - 2016)",
                "https://sharepoint.com/.../Lincoln (2017 - 2021)",
                "https://sharepoint.com/.../Lincoln (2022 - 2026)"
            ],
            "Mazda": [
                "https://sharepoint.com/.../Mazda (2012 - 2016)",
                "https://sharepoint.com/.../Mazda (2017 - 2021)",
                "https://sharepoint.com/.../Mazda (2022 - 2026)"
            ],
            "Mercedes": [
                "https://sharepoint.com/.../Mercedes (2012 - 2016)",
                "https://sharepoint.com/.../Mercedes (2017 - 2021)",
                "https://sharepoint.com/.../Mercedes (2022 - 2026)"
            ],
            "Mini": [
                "https://sharepoint.com/.../Mini (2012 - 2016)",
                "https://sharepoint.com/.../Mini (2017 - 2021)",
                "https://sharepoint.com/.../Mini (2022 - 2026)"
            ],
            "Mitsubishi": [
                "https://sharepoint.com/.../Mitsubishi (2012 - 2016)",
                "https://sharepoint.com/.../Mitsubishi (2017 - 2021)",
                "https://sharepoint.com/.../Mitsubishi (2022 - 2026)"
            ],
            "Nissan": [
                "https://sharepoint.com/.../Nissan (2012 - 2016)",
                "https://sharepoint.com/.../Nissan (2017 - 2021)",
                "https://sharepoint.com/.../Nissan (2022 - 2026)"
            ],
            "Porsche": [
                "https://sharepoint.com/.../Porsche (2012 - 2016)",
                "https://sharepoint.com/.../Porsche (2017 - 2021)",
                "https://sharepoint.com/.../Porsche (2022 - 2026)"
            ],
            "Ram": [
                "https://sharepoint.com/.../Ram (2012 - 2016)",
                "https://sharepoint.com/.../Ram (2017 - 2021)",
                "https://sharepoint.com/.../Ram (2022 - 2026)"
            ],
            "Rolls Royce": [
                "https://sharepoint.com/.../Rolls Royce (2012 - 2016)",
                "https://sharepoint.com/.../Rolls Royce (2017 - 2021)",
                "https://sharepoint.com/.../Rolls Royce (2022 - 2026)"
            ],
            "Subaru": [
                "https://sharepoint.com/.../Subaru (2012 - 2016)",
                "https://sharepoint.com/.../Subaru (2017 - 2021)",
                "https://sharepoint.com/.../Subaru (2022 - 2026)"
            ],
            "Tesla": [
                "https://sharepoint.com/.../Tesla (2012 - 2016)",
                "https://sharepoint.com/.../Tesla (2017 - 2021)",
                "https://sharepoint.com/.../Tesla (2022 - 2026)"
            ],
            "Toyota": [
                "https://sharepoint.com/.../Toyota (2012 - 2016)",
                "https://sharepoint.com/.../Toyota (2017 - 2021)",
                "https://sharepoint.com/.../Toyota (2022 - 2026)"
            ],
            "Volkswagen": [
                "https://sharepoint.com/.../Volkswagen (2012 - 2016)",
                "https://sharepoint.com/.../Volkswagen (2017 - 2021)",
                "https://sharepoint.com/.../Volkswagen (2022 - 2026)"
            ],
            "Volvo": [
                "https://sharepoint.com/.../Volvo (2012 - 2016)",
                "https://sharepoint.com/.../Volvo (2017 - 2021)",
                "https://sharepoint.com/.../Volvo (2022 - 2026)"
            ]
        }

        # how many times to try each manufacturer before giving up
        self.max_attempts = 10

        # track how many times we've tried each one
        self.attempts = {}

        # lists for status
        self.completed_manufacturers = []
        self.failed_manufacturers    = []
        self.failed_excels           = []
        self.given_up_manufacturers  = []
        
        self.thread         = None    # your singular thread slot, if you have one
        self.threads        = []      # ← now you can safely append to self.threads
        self.is_running     = False
        self.stop_requested = False
        self.pause_requested= False
        # ── Run-queue & reentrancy guards ─────────────────────────────────────
        self._finish_guard = False          # prevents double-entry of finish handler
        self.queue_active = False           # True while a multi-manufacturer batch is running
        self.current_index = 0              # index into self.excel_paths / manufacturers_to_run
        self._next_timer = None             # single reusable timer for “check again in 10s”
        # ──────────────────────────────────────────────────────────────────────
        
      
    def initUI(self):
        self.setWindowTitle('Hyper')
        self.setStyleSheet("background-color: #2e2e2e; color: white;")
        layout = QVBoxLayout()
    
        # Excel file selection layout
        file_selection_layout = QHBoxLayout()
        self.select_file_button = CustomButton('Select Excel Files', '#008000', self)
        self.select_file_button.clicked.connect(self.select_excel_files)
        file_selection_layout.addWidget(self.select_file_button)
    
        # Excel file path display
        self.excel_path_label = QLabel('No files selected')
        self.excel_path_label.setStyleSheet(
            "font-size: 14px; padding: 5px; "
            "border: 1px solid #555555; border-radius: 5px; "
            "background-color: #3e3e3e;"
        )
        
        # Excel file list (scrollable)
        self.excel_list = QListWidget(self)
        self.excel_list.setFixedHeight(100)   # tweak height as you like
        
        self.excel_list.setStyleSheet(
            "font-size: 14px; padding: 5px; "
            "background-color: #3e3e3e; color: white; "
            "border: 1px solid #555555; border-radius: 5px;"
        )
        self.excel_list.addItem('No files selected, please select files')
        file_selection_layout.addWidget(self.excel_list)
        layout.addLayout(file_selection_layout)
        self.si_mode_toggle = QCheckBox()
    
        # "Select All (Manufacturers)" and "Select All (ADAS Systems)" button layout
        select_all_buttons_layout = QHBoxLayout()
        self.select_all_manufacturers_button = CustomButton('Select All (Manufacturers)', '#e3b505', self)
        self.select_all_manufacturers_button.clicked.connect(self.select_all_manufacturers)
        select_all_buttons_layout.addWidget(self.select_all_manufacturers_button)
        
        # NEW: Select All (Year Ranges)
        self.select_all_years_button = CustomButton('Select All (Year Ranges)', '#e3b505', self)
        self.select_all_years_button.clicked.connect(self.select_all_year_ranges)
        select_all_buttons_layout.addWidget(self.select_all_years_button)
        
    
        self.select_all_adas_button = CustomButton('Select All (ADAS Systems)', '#e3b505', self)
        self.select_all_adas_button.clicked.connect(self.select_all_adas)
        select_all_buttons_layout.addWidget(self.select_all_adas_button)
        
        self.select_all_repair_button = CustomButton('Select All (Repair Systems)', '#e3b505', self)
        self.select_all_repair_button.clicked.connect(self.select_all_repair)
        select_all_buttons_layout.addWidget(self.select_all_repair_button)

        layout.addLayout(select_all_buttons_layout)
    
        # Manufacturer and ADAS selection layout
        manufacturer_selection_layout = QHBoxLayout()
    
        # Manufacturer tree widget with checkboxes
        manufacturer_list_layout = QVBoxLayout()
        
        # ▶ Label above the manufacturers list
        manufacturer_label = QLabel("Manufacturers")
        manufacturer_label.setAlignment(Qt.AlignHCenter)   # ⬅️ center text
        manufacturer_label.setStyleSheet("font-size: 14px; padding: 5px 6px;")
        manufacturer_list_layout.addWidget(manufacturer_label)
        
        self.manufacturer_tree = QTreeWidget(self)
        self.manufacturer_tree.setHeaderHidden(True)
        self.manufacturer_tree.setFixedWidth(200)  # 👈 Shift closer by narrowing it
        self.manufacturer_tree.setStyleSheet("""
            QTreeWidget {
                background-color: #3e3e3e;
                color: white;
                border: 1px solid #555555;
                border-radius: 5px;
                margin-left: 10px;  /* 👈 Fine-tune left shift */
            }
        """)
        
        # Manufacturer Check Boxes
        manufacturers = ["Acura", "Alfa Romeo", "Audi", "BMW", "Brightdrop", "Buick", "Cadillac", "Chevrolet", "Chrysler", "Dodge",
                         "Fiat", "Ford", "Genesis", "GMC", "Honda", "Hyundai", "Infiniti", "Jaguar", "Jeep", "Kia", "Land Rover", 
                         "Lexus", "Lincoln", "Mazda", "Mercedes", "Mini", "Mitsubishi", "Nissan", "Porsche", "Ram", 
                         "Rolls Royce", "Subaru", "Tesla", "Toyota", "Volkswagen", "Volvo"]
        for manufacturer in manufacturers:
            item = QTreeWidgetItem(self.manufacturer_tree)
            item.setText(0, manufacturer)
            item.setCheckState(0, Qt.Unchecked)
        
        manufacturer_list_layout.addWidget(self.manufacturer_tree)
        
        manufacturer_selection_layout.addLayout(manufacturer_list_layout)
        
        # === Year Ranges (match ADAS Systems visuals) ===
        years_selection_layout = QVBoxLayout()
        
        years_label = QLabel("Year Ranges")
        years_label.setAlignment(Qt.AlignHCenter)
        years_label.setStyleSheet("font-size: 14px; padding: 5px;")
        years_selection_layout.addWidget(years_label)
        
        # Mirror the ADAS pattern: a list + a place to store the created QCheckBox widgets
        year_items = ["2012–2016 Years", "2017–2021 Years", "2022–2026 Years", "2027–2031 Years"]
        self.year_checkboxes = []  # analogous to self.adas_checkboxes
        
        # Create the three year checkboxes just like ADAS does its acronyms
        for text in year_items:
            cb = QCheckBox(text, self)
            self.year_checkboxes.append(cb)
            years_selection_layout.addWidget(cb)
        
        # Keep direct handles for compatibility with existing logic
        # (old names used across your codebase)
        self.year_2012_2016 = self.year_checkboxes[0]
        self.year_2017_2021 = self.year_checkboxes[1]
        self.year_2022_2026 = self.year_checkboxes[2]
        self.year_2027_2031 = self.year_checkboxes[3] 
        
        # Also provide the *_checkbox aliases some code paths expect
        self.year_2012_2016_checkbox = self.year_2012_2016
        self.year_2017_2021_checkbox = self.year_2017_2021
        self.year_2022_2026_checkbox = self.year_2022_2026
        self.year_2027_2031_checkbox = self.year_2027_2031
        
        # Convenience list (kept for any existing loops)
        self._year_checkboxes = [self.year_2012_2016, self.year_2017_2021, self.year_2022_2026, self.year_2027_2031]
        
        # ⬆️ keep content pinned to the top
        years_selection_layout.addStretch(1)
        
        # Add the Year Ranges column into the same parent layout as ADAS Systems
        manufacturer_selection_layout.addLayout(years_selection_layout)

        
        
        # ADAS Acronyms section
        adas_selection_layout = QVBoxLayout()
        adas_label = QLabel("ADAS Systems")
        adas_label.setAlignment(Qt.AlignHCenter)    
        adas_label.setStyleSheet("font-size: 14px; padding: 5px;")
        adas_selection_layout.addWidget(adas_label)
    
        adas_acronyms = ["ACC", "AEB", "AHL", "APA", "BSW", "BUC", "LKA", "LW", "NV", "SVC", "WAMC"]
        self.adas_checkboxes = []
        repair_systems = [
            "SAS", "YAW", "G-Force", "SWS", "AHL", "NV", "HUD", "SRS", "SRA", 
            "ESC", "SRS D&E", "SCI", "SRR", "HLI", "TPMS", "SBI", "RC",
            "EBDE (1)", "EBDE (2)", "HDE (1)", "HDE (2)", "LGR", "PSI", "WRL",
            "PCM", "TRANS", "AIR", "ABS", "BCM","ODS","OCS","OCS2","OCS3","OCS4",
            "KEY", "FOB", "HVAC (1)", "HVAC (2)", "COOL", "HEAD (1)", "HEAD (2)",
        
            # Full Names
            "Steering Angle Sensor",
            "Yaw Rate Sensor",
            "G Force Sensor",
            "Seat Weight Sensor",
            "Adaptive Head Lamps",
            "Night Vision",
            "Heads Up Display",
            "Electronic Stability Control Relearn",
            "Airbag Disengagement/Engagement",
            "Steering Column Inspection",
            "Steering Rack Relearn",
            "Headlamp Initialization",
            "Tire Pressure Monitor Relearn",
            "Seat Belt Inspection",
            "Battery Disengagement",
            "Battery Engagement",
            "Hybrid Disengagement",
            "Hybrid Engagement",
            "Liftgate Relearn",
            "Power Seat Initialization",
            "Window Relearn",
            "Powertrain Control Module Program",
            "Transmission Control Module Program",
            "Airbag Control Module Program",
            "Antilock Brake Control Module Program",
            "Body Control Module Program",
            "Key Program",
            "Key FOB Relearn",
            "Heating, Air Conditioning, Ventilation EVAC",
            "Heating, Air Conditioning, Ventilation Recharge",
            "Coolant Services",
            "Headset Reset (Spring Style)",
            "Headset Reset (Squib Style)",
        ]

        self.repair_checkboxes = []

        for adas in adas_acronyms:
            checkbox = QCheckBox(adas, self)
            checkbox.setStyleSheet("font-size: 12px; padding: 5px;")
            self.adas_checkboxes.append(checkbox)
            adas_selection_layout.addWidget(checkbox)
    
        manufacturer_selection_layout.addLayout(adas_selection_layout)
        layout.addLayout(manufacturer_selection_layout)
                    
        # === Repair Systems Section (Label on top, Scrollable box underneath) ===
        
        # Vertical layout to hold both: label AND scrollable checkbox container
        repair_box_layout = QVBoxLayout()
        
        # Label (not scrollable)
        repair_label = QLabel("Repair Systems")
        repair_label.setAlignment(Qt.AlignHCenter)  
        repair_label.setFixedWidth(200)    
        repair_label.setStyleSheet("font-size: 14px; padding: 5px;")
        repair_box_layout.addWidget(repair_label)
        
        # Scrollable checkbox area (keep a reference so we can restyle it later)
        self.repair_scroll_area = QScrollArea()
        self.repair_scroll_area.setWidgetResizable(True)
        self.repair_scroll_area.setFixedWidth(180)
        self.repair_scroll_area.setStyleSheet(
            "background-color: #3e3e3e; border: 1px solid #555555; border-radius: 5px;"
        )
        
        repair_container = QWidget()
        repair_selection_layout = QVBoxLayout(repair_container)
        
        self.repair_checkboxes = []
        for system in repair_systems:
            checkbox = QCheckBox(system, self)
            checkbox.setStyleSheet("font-size: 12px; padding: 5px;")
            self.repair_checkboxes.append(checkbox)
            repair_selection_layout.addWidget(checkbox)
        
        self.repair_scroll_area.setWidget(repair_container)
        repair_box_layout.addWidget(self.repair_scroll_area)
        
        # Add the full repair module section to the right side
        manufacturer_selection_layout.addLayout(repair_box_layout)

        # Theme switch section
        theme_switch_section = QHBoxLayout()

        # ADAS / Repair SI Label and Toggle
        switch_layout = QHBoxLayout()
        switch_layout.setSpacing(8)
        
        self.label_adas   = QLabel("ADAS SI")
        self.label_repair = QLabel("Repair SI")
        for lbl in (self.label_adas, self.label_repair):
            lbl.setStyleSheet("font-size:14px; padding:5px;")
        
        self.mode_switch = ModeSwitch(self)
        # start unchecked => ADAS
        self.mode_switch.setChecked(False)
        self.mode_switch.stateChanged.connect(self.on_si_mode_toggled)
        
        switch_layout.addWidget(self.label_adas)
        switch_layout.addWidget(self.mode_switch)
        switch_layout.addWidget(self.label_repair)
        switch_layout.addStretch()
        
        layout.addLayout(switch_layout)
        
        # after creating self.si_mode_toggle …
        self.si_mode_toggle.stateChanged.connect(self.on_si_mode_toggled)

        # Excel Format Toggle Layout (OG / New)
        #excel_mode_layout = QHBoxLayout()
        #excel_mode_layout.setSpacing(8)
        #
        #label_og   = QLabel("SME")
        #label_new  = QLabel("New")
        #for lbl in (label_og, label_new):
        #    lbl.setStyleSheet("font-size:14px; padding:5px;")
        #
        #self.excel_mode_switch = ModeSwitch(self)
        #self.excel_mode_switch.setChecked(True)  # Start in New mode
        #
        #excel_mode_layout.addWidget(label_og)
        #excel_mode_layout.addWidget(self.excel_mode_switch)
        #excel_mode_layout.addWidget(label_new)
        #excel_mode_layout.addStretch()
        #
        #layout.addLayout(excel_mode_layout)

        # Dark mode toggle
        theme_switch_section.addStretch()
        self.theme_toggle = ToggleSwitch(self)
        theme_switch_section.addWidget(self.theme_toggle)
        layout.addLayout(theme_switch_section)
    
        # ── Clean up Mode checkbox ──
        self.cleanup_checkbox = QCheckBox("Broken Hyperlink Mode", self)
        self.cleanup_checkbox.setStyleSheet("font-size: 14px; padding: 5px;")
        layout.addWidget(self.cleanup_checkbox)    
    
        # ── Pause/Resume & Start/Stop Buttons ──
        self.pause_button = CustomButton('Pause Automation', '#e3a008', self)
        self.pause_button.clicked.connect(self.on_pause_resume)
        self.pause_button.setEnabled(False)
    
        self.start_button = CustomButton('Start Automation', '#008000', self)
        self.start_button.clicked.connect(self.on_start_stop)
    
        # Use a vertical layout here so Pause sits above Start
        self.button_layout = QVBoxLayout()  
        self.button_layout.addWidget(self.pause_button)
        self.button_layout.addWidget(self.start_button)
        layout.addLayout(self.button_layout)
    
        # ── Progress Bars ──
        self.current_manufacturer_label = QLabel("Current Manufacturer: None")
        self.current_manufacturer_label.setStyleSheet("font-size: 13px; padding: 5px;")
        self.current_manufacturer_progress = QProgressBar()
        self.current_manufacturer_progress.setMaximum(100)
        self.current_manufacturer_progress.setValue(0)
        self.current_manufacturer_progress.setFormat("%p%")
        self.current_manufacturer_progress.setStyleSheet("""
            QProgressBar {
                font-size: 12px;
                padding: 4px;
                color: black;           /* text color */
                text-align: center;     /* center the % */
            }
        """)
        
        # 🆕 Manufacturer Hyperlink Status
        self.manufacturer_hyperlink_label = QLabel("Manufacturer Hyperlinks : 0 / 0")
        self.manufacturer_hyperlink_label.setStyleSheet("font-size: 13px; padding: 5px;")
        self.manufacturer_hyperlink_bar = QProgressBar()
        self.manufacturer_hyperlink_bar.setMaximum(100)
        self.manufacturer_hyperlink_bar.setValue(0)
        self.manufacturer_hyperlink_bar.setFormat("%p%")
        self.manufacturer_hyperlink_bar.setStyleSheet("""
            QProgressBar {
                font-size: 12px;
                padding: 4px;
                color: black;
                text-align: center;
            }
        """)
        
        self.overall_progress_label = QLabel("Overall Progress: 0%")
        self.overall_progress_label.setStyleSheet("font-size: 13px; padding: 5px;")
        self.overall_progress_bar = QProgressBar()
        self.overall_progress_bar.setMaximum(100)
        self.overall_progress_bar.setValue(0)
        self.overall_progress_bar.setFormat("%p%")
        self.overall_progress_bar.setStyleSheet("""
            QProgressBar {
                font-size: 12px;
                padding: 4px;
                color: black;           /* text color */
                text-align: center;     /* center the % */
            }
        """)
        # Style used when manually stopped (red background + red chunk)
        self._bar_style_stopped = """
            QProgressBar {
                font-size: 12px;
                padding: 4px;
                color: white;                         /* % text */
                text-align: center;
                background-color: #4a0f0f;            /* bar background stays red even at 0% */
                border: 1px solid #aa4444;
                border-radius: 4px;
            }
            QProgressBar::chunk {
                background-color: #d32f2f;            /* the filled part */
                margin: 0px;
                border-radius: 2px;
            }
        """
        
        layout.addWidget(self.current_manufacturer_label)
        layout.addWidget(self.current_manufacturer_progress)
        
        # 🆕 Insert new hyperlink bar + label here
        layout.addWidget(self.manufacturer_hyperlink_label)
        layout.addWidget(self.manufacturer_hyperlink_bar)
        
        layout.addWidget(self.overall_progress_label)
        layout.addWidget(self.overall_progress_bar)      
        
        # after creating the bars
        self.current_manufacturer_progress.setObjectName("cmBar")
        self.manufacturer_hyperlink_bar.setObjectName("mhBar")
        self.overall_progress_bar.setObjectName("ovBar")
        
        # one stylesheet that preserves the grey groove; only the CHUNK turns red when stopped
        progress_css = """
        QProgressBar {
            font-size: 12px;
            padding: 4px;
            text-align: center;
            color: black;
            border: 1px solid #555555;
            border-radius: 4px;
            background: #e0e0e0;              /* normal groove */
        }
        QProgressBar::chunk {
            background-color: #19A602;         /* normal (green) fill */
        }
        
        /* when manually stopped, KEEP the same groove; only recolor the chunk.
           Use object-id selectors + !important to override any global red background. */
        QProgressBar#cmBar[stopped="true"],
        QProgressBar#mhBar[stopped="true"],
        QProgressBar#ovBar[stopped="true"] {
            background: #e0e0e0 !important;    /* cancel any red background rules */
        }
        
        QProgressBar#cmBar[stopped="true"]::chunk,
        QProgressBar#mhBar[stopped="true"]::chunk,
        QProgressBar#ovBar[stopped="true"]::chunk {
            background-color: #B30000 !important;  /* red fill only */
        }
        """
        
        # apply to each bar
        self.current_manufacturer_progress.setStyleSheet(progress_css)
        self.manufacturer_hyperlink_bar.setStyleSheet(progress_css)
        self.overall_progress_bar.setStyleSheet(progress_css)
        
        # after adding all widgets and layouts…
        self.si_mode_toggle.stateChanged.connect(self.on_si_mode_toggled)

        # set initial enabled/disabled state based on default toggle
        self.on_si_mode_toggled(self.mode_switch.checkState())

        self.setLayout(layout)
        self.resize(600, 400)

    def select_all_year_ranges(self):
        """Toggle all year range checkboxes on/off."""
        # Check if *all* are currently checked
        all_checked = all(cb.isChecked() for cb in self._year_checkboxes)
        
        # If all are checked, uncheck all; otherwise, check all
        for cb in self._year_checkboxes:
            cb.setChecked(not all_checked)

    
    def get_selected_year_ranges(self):
        ranges = []
        if getattr(self, "year_2012_2016", None) and self.year_2012_2016.isChecked():
            ranges.append((2012, 2016))
        if getattr(self, "year_2017_2021", None) and self.year_2017_2021.isChecked():
            ranges.append((2017, 2021))
        if getattr(self, "year_2022_2026", None) and self.year_2022_2026.isChecked():
            ranges.append((2022, 2026))
        if getattr(self, "year_2027_2031", None) and self.year_2027_2031.isChecked():   # ← NEW
            ranges.append((2027, 2031))
        return ranges
        
    
    def _filter_links_by_selected_years(self, links):
        """
        Filter manufacturer SharePoint links by the Year Ranges checkboxes.
        Priority: (1) regex match '(YYYY-YYYY)' in the link text; (2) index fallback:
          0 -> 2012–2016, 1 -> 2017–2021, 2 -> 2022–2026
        If nothing is selected, return all links.
        """
        import re
        selected = self.get_selected_year_ranges()  # [(2012,2016), (2017,2021), ...]
        if not selected:
            return links
    
        selected_set = set(selected)
        pattern = re.compile(r'(?:(?:\(|\[)?\s*)(20\d{2})\s*[–-]\s*(20\d{2})(?:\s*(?:\)|\]))?')
    
        # Try regex filtering first
        filtered = []
        unmatched_indices = []
        for i, link in enumerate(links):
            m = pattern.search(link)
            if m:
                s, e = int(m.group(1)), int(m.group(2))
                if (s, e) in selected_set:
                    filtered.append(link)
            else:
                unmatched_indices.append(i)
    
        if filtered:
            return filtered
    
        # Fallback: map checkboxes to 0/1/2 positions
        wanted_idx = []
        mapping = {(2012, 2016): 0, (2017, 2021): 1, (2022, 2026): 2, (2027, 2031): 3}
        for rng in selected:
            if rng in mapping:
                wanted_idx.append(mapping[rng])
    
        if not wanted_idx:
            return links  # nothing mapped → don't surprise the user
    
        return [links[i] for i in sorted(set(wanted_idx)) if 0 <= i < len(links)]

    

    def handle_extractor_output(self, line: str):
        """
        Consume stdout from SharepointExtractor. Update UI progress bars,
        parse report data, and manage terminal display.
        """
        import re
    
        # Ensure report buckets exist
        if not hasattr(self, "report_stats"):
            self.report_stats = {}
        if not hasattr(self, "_report_year_totals"):
            self._report_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0,  "2027–2031":0}
    
        # Build a canonical manufacturer map for quick detection
        manu_name_map = {}
        try:
            if hasattr(self, "selected_manufacturers"):
                manu_name_map = {m.lower(): m for m in self.selected_manufacturers}
            elif hasattr(self, "manufacturer_links"):
                manu_name_map = {m.lower(): m for m in self.manufacturer_links.keys()}
        except Exception:
            pass
    
        # ---------- helpers ----------
        def _norm_range(rng: str) -> str:
            # Normalize (2012–2016)/(2012-2016) → "2012–2016"
            rng = rng.strip().replace("—", "-").replace("–", "-")
            m = re.search(r"(\d{4})\s*-\s*(\d{4})", rng)
            if not m:
                return rng
            a, b = int(m.group(1)), int(m.group(2))
            return f"{a}–{b}"
    
        def _hms_to_seconds(s: str) -> int:
            # Accept "H:MM:SS" or "MM:SS"
            parts = [int(x) for x in s.split(":")]
            if len(parts) == 3:
                h, m, sec = parts
            elif len(parts) == 2:
                h, m, sec = 0, parts[0], parts[1]
            else:
                h, m, sec = 0, 0, parts[0]
            return h * 3600 + m * 60 + sec
    
        def _int(s: str) -> int:
            return int(s.replace(",", "").strip())
    
        # NEW: safe hint to current link (string), its index and total
        def _safe_link_hint():
            link_hint, idx, total = "", None, 0
            try:
                if hasattr(self, "_multi_links") and isinstance(self._multi_links, list):
                    total = len(self._multi_links)
                    idx = getattr(self, "_multi_link_index", 0)
                    if total and idx is not None and 0 <= idx < total:
                        link_hint = str(self._multi_links[idx])
            except Exception:
                pass
            return link_hint, idx, total
        
        def _ordinal_to_range_label(ord_num: int) -> str:
            """1→2012–2016, 2→2017–2021, 3→2022–2026, 4→2027–2031; else ''."""
            mapping = {
                1: "2012–2016",
                2: "2017–2021",
                3: "2022–2026",
                4: "2027–2031",  # NEW
            }
            return mapping.get(ord_num, "")

            
        # NEW: index→range fallback (matches your report buckets)
        def _index_range_label(idx: int | None, total: int) -> str:
            """
            Map a 0-based link index to a year-range label using the number of links available.
            - total >= 4 : 0→2012–2016, 1→2017–2021, 2→2022–2026, 3→2027–2031
            - total == 3 : 0→2012–2016, 1→2017–2021, 2→2022–2026
            - total == 2 : 0→2012–2020, 1→2021–2026 (existing special-case)
            """
            if idx is None or total <= 0:
                return ""
        
            if total >= 4:
                mapping = {0: "2012–2016", 1: "2017–2021", 2: "2022–2026", 3: "2027–2031"}  # NEW
                return mapping.get(idx, "")
        
            if total == 3:
                mapping = {0: "2012–2016", 1: "2017–2021", 2: "2022–2026"}
                return mapping.get(idx, "")
        
            if total == 2:
                # sensible fallback if a make only has 2 links
                mapping = {0: "2012–2020", 1: "2021–2026"}
                return mapping.get(idx, "")
        
            return ""
        
            
        # NOTE: assumes you added this class helper earlier
        # def _extract_year_range_from_link(self, link: str) -> str: ...
    
        # Tolerant regexes (compiled once per call)
        RE_CM   = re.compile(r"\s*CM_PROGRESS\s+(\d+)\s*/\s*(\d+)\s*\((\d+)%\)", re.IGNORECASE)
        RE_FR   = re.compile(r"(\d+)\s+Folders Remain", re.IGNORECASE)
    
        # Totals (several variants)
        RE_TOT_A = re.compile(r"(\d{1,2}:\d{2}(?::\d{2})?)\s+Total Time\s*\|\s*Total Files\s*:\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE)
        RE_TOT_B = re.compile(r"Total Files\s*:\s*(\d{1,3}(?:,\d{3})*)\s*\|\s*Total Time\s*:\s*(\d{1,2}:\d{2}(?::\d{2})?)", re.IGNORECASE)
        RE_TOT_TIME_ONLY  = re.compile(r"Total Time\s*:\s*(\d{1,2}:\d{2}(?::\d{2})?)", re.IGNORECASE)
        RE_TOT_FILES_ONLY = re.compile(r"(?:Total Files|Files Indexed|Files Found)\s*:\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE)
        RE_TOT_SECONDS    = re.compile(r"Indexing routine took\s*(\d+(?:\.\d+)?)\s*seconds", re.IGNORECASE)
    
        # Per-link (very tolerant): allow empty () in canonical form
        RE_LINK_CANON = re.compile(
            r"-{2,}\s*(\d+)(?:st|nd|rd|th)\s+([A-Za-z][A-Za-z &\-/]+?)\s+Link\s*\(([^)]*)\)\s*[:\-–]\s*"
            r"(\d{1,2}:\d{2}(?::\d{2})?)\s*\|\s*(?:Files(?:\s*(?:Found|Indexed))?|Files)\s*[:=]?\s*(\d{1,3}(?:,\d{3})*)",
            re.IGNORECASE
        )

        # Fallback: any line containing a year-range, a time token, and a files count token
        RE_YEAR   = re.compile(r"\((\s*\d{4}\s*[-–—]\s*\d{4}\s*)\)")
        RE_TIME   = re.compile(r"\b(\d{1,2}:\d{2}(?::\d{2})?)\b")
        RE_FILES1 = re.compile(r"(?:Files(?:\s*(?:Found|Indexed))?|Found)\s*[:=]?\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE)
        RE_FILES2 = re.compile(r"\b(\d{1,3}(?:,\d{3})*)\s+Files\b", re.IGNORECASE)
    
        # Manufacturer header lines (several styles)
        RE_MAKE1 = re.compile(r"^\s*Manufacturer\s*:\s*([A-Za-z][A-Za-z0-9 &\-/]+)\s*$", re.IGNORECASE)
        RE_MAKE2 = re.compile(r"^\s*Processing\s+(?:manufacturer|make)\s*[:\-]\s*([A-Za-z][A-Za-z0-9 &\-/]+)\s*$", re.IGNORECASE)
        RE_MAKE3 = re.compile(r"^\s*=+\s*([A-Za-z][A-Za-z0-9 &\-/]+)\s*=+\s*$")
    
        # We may receive multiple physical lines at once; handle each
        text = line if isinstance(line, str) else str(line)
        orig_ended_nl = text.endswith("\n")
        echo_lines = []
    
        for raw in text.splitlines():
            s = raw.strip()
            show_this_line = True
    
            # 0) Detect and lock current make from header-like lines or a bare manufacturer name
            mk_from_header = None
            m = RE_MAKE1.match(s) or RE_MAKE2.match(s) or RE_MAKE3.match(s)
            if m:
                mk_from_header = m.group(1).strip()
            elif s and s.lower() in manu_name_map:
                mk_from_header = manu_name_map[s.lower()]
    
            if mk_from_header:
                self._report_current_make = manu_name_map.get(mk_from_header.lower(), mk_from_header)
                self.report_stats.setdefault(self._report_current_make, {"total_time": 0, "total_files": 0, "links": []})
    
            # 1) CM progress: "CM_PROGRESS a/b (p%)"
            m_cm = RE_CM.match(s)
            if m_cm:
                done = int(m_cm.group(1))
                total = max(1, int(m_cm.group(2)))
                pct_from_text = int(m_cm.group(3))
                pct = max(0, min(100, pct_from_text if 0 <= pct_from_text <= 100 else int(done / total * 100)))
                self.current_manufacturer_progress.setValue(pct)
                if getattr(self, "hide_cm_progress_in_terminal", False):
                    show_this_line = False
    
            # 2) Cleanup-mode tickers
            if getattr(self, '_cleanup_mode', False):
                m_total = re.search(r'Total broken hyperlinks:\s*(\d+)', s)
                if m_total:
                    self._initial_broken = int(m_total.group(1))
                    self._fixed_count = 0
                if s.startswith(("Fixed hyperlink for", "✅", "❌")) and getattr(self, "_initial_broken", None):
                    self._fixed_count += 1
                    pct = int(self._fixed_count / max(1, self._initial_broken) * 100)
                    self.current_manufacturer_progress.setValue(pct)
    
            # 3) Normal mode progress using "N Folders Remain"
            m_fr = RE_FR.search(s)
            if m_fr:
                remaining = int(m_fr.group(1))
                if not hasattr(self, '_initial_folder_count') or self._initial_folder_count is None:
                    self._initial_folder_count = remaining
                else:
                    self._initial_folder_count = max(self._initial_folder_count, remaining)
                initial = max(1, self._initial_folder_count)
                pct = max(0, min(100, int((initial - remaining) / initial * 100)))
                self.current_manufacturer_progress.setValue(pct)
    
            # 4) Report parsing
            current_make = getattr(self, "_report_current_make", None)
            if not current_make and hasattr(self, "selected_manufacturers") and 0 <= self.current_index < len(self.selected_manufacturers):
                current_make = self.selected_manufacturers[self.current_index]
    
            # 4a) Per-link — canonical form first (now accepts empty '()')
            m_link = RE_LINK_CANON.search(s)
            if m_link:
                ord_str, make_name, yr_raw, rt_str, files_str = m_link.groups()
                mk = manu_name_map.get(make_name.strip().lower(), make_name.strip())
                printed_yr = _norm_range((yr_raw or "").strip())  # may be empty
                secs = _hms_to_seconds(rt_str)
                files = _int(files_str)
            
                # derive final range: printed → from link text → ordinal mapping → index mapping
                link_hint, idx, total = _safe_link_hint()
                fallback_range = ""
                try:
                    if hasattr(self, "_extract_year_range_from_link"):
                        fallback_range = self._extract_year_range_from_link(link_hint) or ""
                except Exception:
                    pass
                if not fallback_range:
                    try:
                        ord_num = int(ord_str)
                        fallback_range = _ordinal_to_range_label(ord_num)
                    except Exception:
                        fallback_range = ""
                if not fallback_range:
                    fallback_range = _index_range_label(idx, total)
            
                final_range = printed_yr or fallback_range
            
                self.report_stats.setdefault(mk, {"total_time": 0, "total_files": 0, "links": []})
                sig = (final_range, secs, files)
                existing = {(l["range"], l["time"], l["files"]) for l in self.report_stats[mk]["links"]}
                if sig not in existing:
                    self.report_stats[mk]["links"].append({
                        "range": final_range,
                        "time": secs,
                        "files": files
                    })
                    # increment GRAND TOTALS using the final label (not raw printed)
                    if final_range in self._report_year_totals:
                        self._report_year_totals[final_range] += files
            
                # keep your existing context line
                self._report_current_make = mk

    
            else:
                # 4b) Per-link — generic fallback (year-range + a time + a files number)
                m_year = RE_YEAR.search(s)
                if m_year and (RE_FILES1.search(s) or RE_FILES2.search(s)) and RE_TIME.search(s):
                    printed_yr = _norm_range(m_year.group(1) or "")
                    # decide which "make" to attach to: prefer explicit "for <make>" or context
                    mk = current_make
                    m_for = re.search(r"\bfor\s+([A-Za-z][A-Za-z0-9 &\-/]+)", s, re.IGNORECASE)
                    if m_for:
                        guess = m_for.group(1).strip()
                        mk = manu_name_map.get(guess.lower(), guess)
                    if mk:
                        # pick the last time token on the line
                        times = RE_TIME.findall(s)
                        rt_str = times[-1] if times else "0:00"
                        files_m = RE_FILES1.search(s) or RE_FILES2.search(s)
                        files = _int(files_m.group(1)) if files_m else 0
                        secs = _hms_to_seconds(rt_str)
                    
                        # derive final range (printed → from link text → index mapping)
                        link_hint, idx, total = _safe_link_hint()
                        fallback_range = ""
                        try:
                            if hasattr(self, "_extract_year_range_from_link"):
                                fallback_range = self._extract_year_range_from_link(link_hint) or ""
                        except Exception:
                            pass
                        if not fallback_range:
                            fallback_range = _index_range_label(idx, total)
                    
                        final_range = printed_yr or fallback_range
                    
                        self.report_stats.setdefault(mk, {"total_time": 0, "total_files": 0, "links": []})
                        sig = (final_range, secs, files)
                        existing = {(l["range"], l["time"], l["files"]) for l in self.report_stats[mk]["links"]}
                        if sig not in existing:
                            self.report_stats[mk]["links"].append({
                                "range": final_range,
                                "time": secs,
                                "files": files
                            })
                            if final_range in self._report_year_totals:
                                self._report_year_totals[final_range] += files
                    
    
            # 4c) Totals lines (various orders). Use context make if present.
            mk_ctx = getattr(self, "_report_current_make", None) or current_make
            if mk_ctx:
                hit_total = False
                m = RE_TOT_A.search(s)
                if m:
                    rt_str, files_str = m.groups()
                    self.report_stats.setdefault(mk_ctx, {"total_time": 0, "total_files": 0, "links": []})
                    self.report_stats[mk_ctx]["total_time"]  = max(self.report_stats[mk_ctx]["total_time"], _hms_to_seconds(rt_str))
                    self.report_stats[mk_ctx]["total_files"] = max(self.report_stats[mk_ctx]["total_files"], _int(files_str))
                    hit_total = True
                else:
                    m = RE_TOT_B.search(s)
                    if m:
                        files_str, rt_str = m.groups()
                        self.report_stats.setdefault(mk_ctx, {"total_time": 0, "total_files": 0, "links": []})
                        self.report_stats[mk_ctx]["total_time"]  = max(self.report_stats[mk_ctx]["total_time"], _hms_to_seconds(rt_str))
                        self.report_stats[mk_ctx]["total_files"] = max(self.report_stats[mk_ctx]["total_files"], _int(files_str))
                        hit_total = True
                    else:
                        m = RE_TOT_TIME_ONLY.search(s)
                        if m:
                            rt_str = m.group(1)
                            self.report_stats.setdefault(mk_ctx, {"total_time": 0, "total_files": 0, "links": []})
                            self.report_stats[mk_ctx]["total_time"] = max(self.report_stats[mk_ctx]["total_time"], _hms_to_seconds(rt_str))
                            hit_total = True
                        m2 = RE_TOT_FILES_ONLY.search(s)
                        if m2:
                            files_str = m2.group(1)
                            self.report_stats.setdefault(mk_ctx, {"total_time": 0, "total_files": 0, "links": []})
                            self.report_stats[mk_ctx]["total_files"] = max(self.report_stats[mk_ctx]["total_files"], _int(files_str))
                            hit_total = True
                        m3 = RE_TOT_SECONDS.search(s)
                        if m3:
                            secs = int(float(m3.group(1)))
                            self.report_stats.setdefault(mk_ctx, {"total_time": 0, "total_files": 0, "links": []})
                            self.report_stats[mk_ctx]["total_time"] = max(self.report_stats[mk_ctx]["total_time"], secs)
                            hit_total = True
                # if totals hit but files still 0 and we have links, backfill per-make files from links
                if hit_total and self.report_stats.get(mk_ctx, {}).get("total_files", 0) == 0:
                    links = self.report_stats.get(mk_ctx, {}).get("links", [])
                    if links:
                        self.report_stats[mk_ctx]["total_files"] = sum(l["files"] for l in links)
    
            # collect for echo, respecting hide flags
            if show_this_line:
                echo_lines.append(raw)
    
        # Append echo lines (rather than raw chunk)
        if echo_lines:
            self.terminal.append_output("\n".join(echo_lines) + ("\n" if orig_ended_nl else ""))



   
    def mark_manual_stop(self):
        """
        Reflect a manual stop in labels and progress bars (works for Cleanup + Regular).
        Does NOT touch buttons; your on_start_stop() already handles swapping Start/Stop.
        """
        # Labels
        if hasattr(self, "current_manufacturer_label"):
            self.current_manufacturer_label.setText("Current Manufacturer: Manually Stopped")
        if hasattr(self, "manufacturer_hyperlink_label"):
            self.manufacturer_hyperlink_label.setText("Manufacturer Hyperlinks: Manually Stopped")
        if hasattr(self, "overall_progress_label"):
            self.overall_progress_label.setText("Overall Progress: Manually Stopped")
    
        # Progress bars → reset to 0
        for bar_name in ("current_manufacturer_progress", "manufacturer_hyperlink_bar", "overall_progress_bar"):
            bar = getattr(self, bar_name, None)
            if bar:
                try:
                    bar.setMaximum(100)  # in case it was set to a link-count
                    bar.setValue(0)
                except Exception:
                    pass
    
        # Clear any counters so next run starts fresh
        for attr in ("_initial_broken", "_fixed_count", "_initial_folder_count",
                     "_hyperlinks_total_links", "_hyperlinks_done_links"):
            if hasattr(self, attr):
                setattr(self, attr, None)
       
    def _style_single_bar(self, bar: QProgressBar, stopped: bool) -> None:
        if not bar:
            return
        bar.setStyleSheet(self._bar_style_stopped if stopped else self._bar_style_normal)
    
    def _apply_stopped_style_to_all_bars(self, stopped: bool):
        """
        Ensure bars use the baseline CSS (no lingering red background),
        then flip only the CHUNK to red via the 'stopped' dynamic property.
        Also force a visible red fill even if the bar value is 0%.
        """
        bars = (
            getattr(self, "current_manufacturer_progress", None),
            getattr(self, "manufacturer_hyperlink_bar", None),
            getattr(self, "overall_progress_bar", None),
        )
        for name_hint, bar in zip(("cmBar", "mhBar", "ovBar"), bars):
            if not bar:
                continue
    
            # Reset and reapply baseline CSS
            bar.setStyleSheet("")
            bar.setStyleSheet(getattr(self, "_progress_css", ""))
    
            # Apply stopped property
            bar.setProperty("stopped", bool(stopped))
            bar.style().unpolish(bar)
            bar.style().polish(bar)
            bar.update()
    
            # Force red fill to appear even if value == 0
            self._force_zero_red(bar, enable=stopped, full=True)

          
    def on_si_mode_toggled(self, state):
        """Enable one list & button, disable—and clear—the other."""
        is_repair = (state == Qt.Checked)
    
        # Repair group gets enabled; ADAS gets disabled & cleared
        for cb in self.repair_checkboxes:
            cb.setEnabled(is_repair)
        if not is_repair:
            for cb in self.repair_checkboxes:
                cb.setChecked(False)
        self.select_all_repair_button.setEnabled(is_repair)
        if not is_repair:
            self.select_all_repair_button.setChecked(False)
    
        # ADAS group gets enabled; Repair gets disabled & cleared
        for cb in self.adas_checkboxes:
            cb.setEnabled(not is_repair)
        if is_repair:
            for cb in self.adas_checkboxes:
                cb.setChecked(False)
        self.select_all_adas_button.setEnabled(not is_repair)
        if is_repair:
            self.select_all_adas_button.setChecked(False)
        # ✅ Disable Excel Format toggle if Repair SI is active
        # ✅ Disable Excel Format toggle and reset to OG when Repair SI is active
        #if self.excel_mode_switch:
        #    if is_repair:
        #        self.excel_mode_switch.setChecked(False)   # ← Reset to OG
        #        self.excel_mode_switch.setEnabled(False)   # ← Gray out
        #    else:
        #        self.excel_mode_switch.setEnabled(True)

    # Function to select/unselect all manufacturers
    def select_all_manufacturers(self):
        select_all_checked = True
        for i in range(self.manufacturer_tree.topLevelItemCount()):
            item = self.manufacturer_tree.topLevelItem(i)
            if item.checkState(0) != Qt.Checked:
                select_all_checked = False
                break
    
        for i in range(self.manufacturer_tree.topLevelItemCount()):
            item = self.manufacturer_tree.topLevelItem(i)
            item.setCheckState(0, Qt.Checked if not select_all_checked else Qt.Unchecked)
    
    # Function to select/unselect all ADAS systems
    def select_all_adas(self):
        select_all_checked = all(checkbox.isChecked() for checkbox in self.adas_checkboxes)
    
        for checkbox in self.adas_checkboxes:
            checkbox.setChecked(not select_all_checked)

    def select_all_repair(self):
        select_all_checked = all(checkbox.isChecked() for checkbox in self.repair_checkboxes)
        for checkbox in self.repair_checkboxes:
            checkbox.setChecked(not select_all_checked)
    
    def select_excel_files(self):
        self.excel_paths, _ = QFileDialog.getOpenFileNames(
            self, 'Open files', 'C:/Users/', "Excel files (*.xlsx *.xls)"
        )
        if self.excel_paths:
            # Trim any stray whitespace
            self.excel_paths = [p.strip() for p in self.excel_paths]
    
            # 1) Show numbered filenames in the label
            numbered = [f"{i+1}. {os.path.basename(p)}"
                        for i, p in enumerate(self.excel_paths)]
            self.excel_path_label.setText("\n".join(numbered))
    
            # 2) Fill the scrollable list with the same numbering
            self.excel_list.clear()
            for i, p in enumerate(self.excel_paths):
                self.excel_list.addItem(f"{i+1}. {os.path.basename(p)}")
    
        else:
            # No files chosen: clear both widgets
            self.excel_path_label.setText('No files selected')
            self.excel_list.clear()
            self.excel_list.addItem('No files selected, please select files')
         
    def toggle_theme(self):
        if self.theme_toggle.isChecked():
            # light
            self.setStyleSheet("background-color: #ffffff; color: black;")
            self.manufacturer_tree.setStyleSheet(
                "background-color: #f0f0f0; color: black; "
                "border: 1px solid #cccccc; border-radius: 5px;"
            )
            self.excel_path_label.setStyleSheet(
                "font-size: 14px; padding: 5px; "
                "border: 1px solid #cccccc; border-radius: 5px; "
                "background-color: #f0f0f0;"
            )
            self.excel_list.setStyleSheet(
                "font-size: 14px; padding: 5px; "
                "border: 1px solid #cccccc; border-radius: 5px; "
                "background-color: #f0f0f0; color: black;"
            )
            self.repair_scroll_area.setStyleSheet(
                "background-color: #f0f0f0; "
                "border: 1px solid #cccccc; border-radius: 5px;"
            )
        else:
            # dark
            self.setStyleSheet("background-color: #2e2e2e; color: white;")
            self.manufacturer_tree.setStyleSheet(
                "background-color: #3e3e3e; color: white; "
                "border: 1px solid #555555; border-radius: 5px;"
            )
            self.excel_path_label.setStyleSheet(
                "font-size: 14px; padding: 5px; "
                "border: 1px solid #555555; border-radius: 5px; "
                "background-color: #3e3e3e;"
            )
            self.excel_list.setStyleSheet(
                "font-size: 14px; padding: 5px; "
                "border: 1px solid #555555; border-radius: 5px; "
                "background-color: #3e3e3e; color: white;"
            )
            self.repair_scroll_area.setStyleSheet(
                "background-color: #3e3e3e; "
                "border: 1px solid #555555; border-radius: 5px;"
            )

    # ========= REPORT LOG BACKFILL HELPERS =========
    def _find_latest_log_file(self):
        """
        Return the most recent log file path. Prefer FileHandler's file; otherwise
        scan the Hyper Logs directory for the freshest .log/.txt (excluding our report files).
        """
        import os, glob, logging
        # Try logging.FileHandler first
        try:
            paths = []
            for h in logging.getLogger().handlers:
                if hasattr(h, "baseFilename"):
                    paths.append(h.baseFilename)
            # choose newest by mtime
            paths = [p for p in paths if p and os.path.isfile(p)]
            if paths:
                paths.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                return paths[0]
        except Exception:
            pass
    
        # Fallback: scan Hyper Logs folder
        log_dir = self._get_hyper_logs_dir()
        candidates = []
        for pattern in ("*.log", "*.txt"):
            candidates.extend(glob.glob(os.path.join(log_dir, pattern)))
        # exclude report files themselves
        candidates = [p for p in candidates if "Hyper Report " not in os.path.basename(p)]
        if not candidates:
            return None
        candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return candidates[0]
    
    def _backfill_report_from_log(self, log_path=None):
        """
        Parse the latest log and populate/merge self.report_stats and self._report_year_totals.
        Works with segments like:
          - 'Configured new SharepointExtractor for <Make> correctly!'  (segment start)
          - '... Folders Remain | ... Files Indexed'                     (we keep last files count)
          - 'Extraction and population for <Make> is complete!'          (segment end)
          - 'Finished SharePoint link X/3 for <Make>'                    (link index => year range)
        Returns True if anything was parsed.
        """
        import os, re, datetime
    
        # Buckets
        if not hasattr(self, "report_stats"):
            self.report_stats = {}
        if not hasattr(self, "_report_year_totals"):
            self._report_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0}
        # Track freshest epoch per make (when idx==1 finalizes)
        if not hasattr(self, "_report_last_mk_ts"):
            self._report_last_mk_ts = {}  # { "Acura": datetime, ... }
    
        # Find a log file if none provided
        if not log_path:
            log_path = self._find_latest_log_file()
        if not log_path or not os.path.isfile(log_path):
            print("ℹ️ _backfill_report_from_log: no log file found.")
            return False
    
        # Known manufacturers (for clean casing)
        manu_map = {}
        try:
            if hasattr(self, "selected_manufacturers"):
                manu_map = {m.lower(): m for m in self.selected_manufacturers}
            elif hasattr(self, "manufacturer_links"):
                manu_map = {m.lower(): m for m in self.manufacturer_links.keys()}
        except Exception:
            pass
    
        def canon_make(name):
            if not name: return None
            return manu_map.get(name.strip().lower(), name.strip())
    
        # Time parsing
        def _parse_ts(line):
            m = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),(\\d{3})', line)
            if not m:
                m = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),(\d{3})', line)  # in case backslash escaped
            if not m:
                m = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}),(\d{1,3})', line)
            if not m:
                return None
            base, ms = m.group(1), int(m.group(2))
            try:
                dt = datetime.datetime.strptime(base, "%Y-%m-%d %H:%M:%S")
                return dt + datetime.timedelta(milliseconds=ms)
            except Exception:
                return None
    
        # Year-range by index
        def _yr_for(idx, total):
            if total == 3 and 1 <= idx <= 3:
                return ["2012–2016", "2017–2021", "2022–2026"][idx - 1]
            if total == 2 and 1 <= idx <= 2:
                return ["2012–2020", "2021–2026"][idx - 1]
            return ""
    
        def _int(s):
            return int(str(s).replace(",", "").strip())
    
        # Regexes for this log format
        RE_CFG   = re.compile(r"Configured new SharepointExtractor for (.+?) correctly!", re.IGNORECASE)
        RE_DONE  = re.compile(r"Extraction and population for (.+?) is complete!", re.IGNORECASE)
        RE_FIN   = re.compile(r"Finished SharePoint link\s+(\d+)\s*/\s*(\d+)\s*for\s*(.+)$", re.IGNORECASE)
        RE_FIDX  = re.compile(r"(\d+)\s+Folders Remain\s*\|\s*(\d{1,3}(?:,\d{3})*)\s+Files Indexed", re.IGNORECASE)
        RE_TOT   = re.compile(r"(\d{1,2}:\d{2}(?::\d{2})?)\s+Total Time\s*\|\s*Total Files:\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE)
    
        # State while scanning
        any_hit = False
        seg = {}         # { make: {"start": ts, "end": ts|None, "last_files": int|0} }
        finalized = {}   # per-run de-dupe by idx (reset on every RE_CFG for that make)
    
        with open(log_path, "r", encoding="utf-8", errors="ignore") as fh:
            for raw in fh:
                s = raw.strip()
                if not s:
                    continue
                ts = _parse_ts(raw)
    
                # Segment start
                m = RE_CFG.search(s)
                if m and ts:
                    mk = canon_make(m.group(1))
                    any_hit = True
                    seg[mk] = {"start": ts, "end": None, "last_files": 0}
                    finalized[mk] = set()  # reset per-run idx dedupe for this make
                    self.report_stats.setdefault(mk, {"total_time": 0, "total_files": 0, "links": []})
                    continue
    
                # Track last files count
                m = RE_FIDX.search(s)
                if m:
                    _, files = m.groups()
                    files = _int(files)
                    for mk, st in seg.items():
                        if st.get("start") and not st.get("end"):
                            st["last_files"] = files
                    continue
    
                # Segment end (extraction complete)
                m = RE_DONE.search(s)
                if m and ts:
                    mk = canon_make(m.group(1))
                    if mk in seg and seg[mk].get("start"):
                        seg[mk]["end"] = ts
                    continue
    
                # Link finalize (index/total)
                m = RE_FIN.search(s)
                if m:
                    idx, total, mk_raw = m.groups()
                    idx, total = int(idx), int(total)
                    mk = canon_make(mk_raw)
                    if not mk:
                        continue
                    any_hit = True
                    if mk not in seg or not seg[mk].get("start"):
                        continue
                    if idx in finalized.setdefault(mk, set()):
                        continue
    
                    start_ts = seg[mk].get("start")
                    end_ts   = seg[mk].get("end") or ts or start_ts
                    duration = (end_ts - start_ts).total_seconds() if (start_ts and end_ts) else 0
                    files    = int(seg[mk].get("last_files", 0))
                    yr       = _yr_for(idx, total)
    
                    # Epoch decision: only when idx==1 (first link) do we treat this as a new run and overwrite
                    seg_ts  = end_ts or start_ts
                    last_ts = self._report_last_mk_ts.get(mk)
                    if idx == 1 and seg_ts and (last_ts is None or seg_ts > last_ts):
                        # Newer epoch for this make → wipe and start fresh
                        self._report_last_mk_ts[mk] = seg_ts
                        self.report_stats[mk] = {"total_time": 0, "total_files": 0, "links": []}
    
                    # Latest-wins per year-range: replace existing range entry, else append
                    self.report_stats.setdefault(mk, {"total_time": 0, "total_files": 0, "links": []})
                    links = self.report_stats[mk]["links"]
                    replaced = False
                    for i, l in enumerate(links):
                        if l.get("range") == yr:
                            links[i] = {"range": yr, "time": int(round(duration)), "files": files}
                            replaced = True
                            break
                    if not replaced:
                        links.append({"range": yr, "time": int(round(duration)), "files": files})
    
                    finalized[mk].add(idx)
                    # keep segment active; idx 2/3 will use the same seg
                    continue
    
                # Optional combined total line (kept but harmless when present)
                m = RE_TOT.search(s)
                if m:
                    rt_str, files_str = m.groups()
                    secs = sum(int(x) * f for x, f in zip(rt_str.split(":")[-3:], (3600, 60, 1)))
                    files = _int(files_str)
                    if seg:
                        mk = list(seg.keys())[-1]
                        self.report_stats.setdefault(mk, {"total_time": 0, "total_files": 0, "links": []})
                        self.report_stats[mk]["total_time"]  = max(self.report_stats[mk]["total_time"], secs)
                        self.report_stats[mk]["total_files"] = max(self.report_stats[mk]["total_files"], files)
                    continue
    
        # Backfill per-make totals from links (latest-wins already applied per range)
        for mk, data in self.report_stats.items():
            link_files = sum(l["files"] for l in data.get("links", []))
            link_secs  = sum(l["time"]  for l in data.get("links", []))
            if data.get("total_files", 0) == 0:
                data["total_files"] = link_files
            else:
                data["total_files"] = max(data["total_files"], link_files)
            if data.get("total_time", 0) == 0:
                data["total_time"] = link_secs
            else:
                data["total_time"] = max(data["total_time"], link_secs)
    
        # Recompute year totals fresh from report_stats to avoid double-counting
        self._report_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0}
        for mk, data in self.report_stats.items():
            for l in data.get("links", []):
                if l["range"] in self._report_year_totals:
                    self._report_year_totals[l["range"]] += int(l.get("files", 0))
    
        return any_hit
    



    def start_automation(self):
        # 1) gather selected manufacturers
        selected_manufacturers = []
        for i in range(self.manufacturer_tree.topLevelItemCount()):
            item = self.manufacturer_tree.topLevelItem(i)
            if item.checkState(0) == Qt.Checked:
                selected_manufacturers.append(item.text(0))
    
        # 2) gather selected systems based on the slide‐toggle
        if self.mode_switch.isChecked():   # Repair mode
            selected_systems = [cb.text() for cb in self.repair_checkboxes if cb.isChecked()]
        else:                              # ADAS mode
            selected_systems = [cb.text() for cb in self.adas_checkboxes if cb.isChecked()]
    
        # 3) sanity check
        if not (self.excel_paths and selected_manufacturers and selected_systems):
            QMessageBox.warning(self, 'Warning',
                "Please select Excel files, manufacturers, and at least one system.", QMessageBox.Ok)
            return
    
        # 4) confirm and kick off
        excel_list = "\n".join(f"{i+1}. {os.path.basename(path)}" for i, path in enumerate(self.excel_paths))
        manu_list  = "\n".join(f"{i+1}. {m}" for i, m in enumerate(selected_manufacturers))
    
        cleanup_note = ""
        if self.cleanup_checkbox.isChecked():
            cleanup_note = (
                "\n\n⚠️ Broken Hyperlink Mode Activated:\n"
                "With this selected, it will ignore all the ADAS/Repair arguments\n"
                "and find the broken links. Based off of those results, it will\n"
                "find the matching links and repair them."
            )
    
        excel_format   = "Repair SI" if self.mode_switch.isChecked() else "ADAS SI"
        #version_format = "NEW" if self.excel_mode_switch.isChecked() else "OG"
    
        # Format the selected year ranges like "2012–2016, 2017–2021"
        ranges = self.get_selected_year_ranges() if hasattr(self, "get_selected_year_ranges") else []
        years_list = ", ".join(f"{a}–{b}" for (a, b) in ranges) if ranges else "None"
        
        confirm_message = (
            "Excel files selected:\n"
            f"{excel_list}\n\n"
            "Manufacturers selected:\n"
            f"{manu_list}\n\n"
            "Years selected:\n"
            f"{years_list}\n\n"
            "Systems selected:\n"
            + ", ".join(selected_systems) + "\n\n"
            "Excel Format:\n"
            f"{excel_format}\n\n"
            #"Version Format:\n"
            #f"{version_format}"
            + cleanup_note + "\n\nContinue?"
        )
        
    
        if QMessageBox.question(self, 'Confirmation', confirm_message,
               QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
            return
    
        # user clicked YES → mark running
        self.is_running     = True
        self.stop_requested = False
        self._report_written = False  # <-- reset "written" flag for this batch
    
        # rip out the old “Start” button and insert a red “Stop Automation”
        layout = self.button_layout
        layout.removeWidget(self.start_button)
        self.start_button.deleteLater()
        self.start_button = CustomButton("Stop Automation", "#e63946", self)
        self.start_button.clicked.connect(self.on_start_stop)
        layout.addWidget(self.start_button)
    
        # enable Pause
        self.pause_button.setEnabled(True)
        self.pause_button.setText('Pause Automation')
        self.pause_requested = False
    
        # 5) stash for process_next_manufacturer
        self.selected_manufacturers = selected_manufacturers
        self.selected_systems       = selected_systems
        self.mode_flag              = "repair" if self.mode_switch.isChecked() else "adas"
    
        # REPORT: fresh state for each batch (AFTER mode_flag is set)
        self.report_stats = {}                        # per-make stats bucket
        self._report_year_totals = {                  # grand-total per range
            "2012–2016": 0,
            "2017–2021": 0,
            "2022–2026": 0,
        }
        self.run_start = time.time()
        self._report_header_label = f"{'Repair SI' if self.mode_flag == 'repair' else 'ADAS SI'} PDF Document"
    
        self.current_index          = 0
        self.total_manufacturers    = len(self.selected_manufacturers)
    
        # ---- PRE-PRIME LABELS TO AVOID "None" / "0 / 0" FLASH ----
        # ---- PRE-PRIME LABELS TO AVOID "None" / "0 / 0" FLASH ----
        first_manufacturer = self.selected_manufacturers[0]
        self.current_manufacturer_label.setText(f"Current Manufacturer: {first_manufacturer}")
        
        # Build the preview list once (already applies year-range + cleanup filters)
        # so the progress bar shows the real count.
        sp_links = self._links_for_manufacturer_preview(first_manufacturer)
        
        total_links = len(sp_links) or 1  # keep bar valid even if empty
        self._hyperlinks_total_links = total_links
        
        # keep bar in sync so nothing can overwrite back to 0/0
        self.manufacturer_hyperlink_bar.setMaximum(max(1, total_links))
        self.manufacturer_hyperlink_bar.setValue(0)
        self.manufacturer_hyperlink_label.setText(
            f"Manufacturer Hyperlinks Indexed: 0 / {total_links}"
        )
        
    
        # get links for this manufacturer
        link_dict = self.repair_links if self.mode_flag == "repair" else self.manufacturer_links
        sharepoint_links = link_dict.get(first_manufacturer) or []
        if isinstance(sharepoint_links, str):
            sharepoint_links = [sharepoint_links]
        
        # NEW: honor Year Ranges checkboxes for the actual run
        sharepoint_links = self._filter_links_by_selected_years(sharepoint_links)
        if not sharepoint_links:
            self.terminal.append_output(
                f"ℹ️ No links match selected years for {first_manufacturer}; skipping."
            )
            # make sure you advance to the next manufacturer cleanly
            self.on_manufacturer_finished(first_manufacturer, True)
            return
        
        # proceed as before
        self._multi_links = sharepoint_links
        self._multi_link_index = 0
        self._hyperlinks_total_links = len(self._multi_links)
        # (keep your existing label/bar updates here)
        
        
    
        if self.cleanup_checkbox.isChecked():
            years_needed = self.get_broken_hyperlink_years_for_manufacturer(first_manufacturer)
            filtered = []
            for link in sp_links:
                m = re.search(r'\((\d{4})\s*-\s*(\d{4})\)', link)
                if m:
                    start_y, end_y = int(m.group(1)), int(m.group(2))
                    if any(start_y <= y <= end_y for y in years_needed):
                        filtered.append(link)
            if filtered:
                sp_links = filtered
    
        total_links = max(1, len(sp_links)) if not sp_links else len(sp_links)
        self.manufacturer_hyperlink_bar.setMaximum(max(1, total_links))
        self.manufacturer_hyperlink_bar.setValue(0)
        self.manufacturer_hyperlink_label.setText(f"Manufacturer Hyperlinks: 0 / {total_links}")
        self.overall_progress_label.setText(f"Overall Progress: 0 / {self.total_manufacturers}")
    
        # progress bars reset
        self.overall_progress_bar.setValue(0)
        self.current_manufacturer_progress.setValue(0)
        self.overall_progress_label.setText(f"Overall Progress: 0 / {self.total_manufacturers}")
    
        try:
            self.manufacturer_hyperlink_label.setText("Manufacturer Hyperlinks: 0 / 0")
            if hasattr(self, "manufacturer_hyperlink_bar"):
                self.manufacturer_hyperlink_bar.setMaximum(100)
                self.manufacturer_hyperlink_bar.setValue(0)
        except Exception:
            pass
    
        self._hyperlinks_total_links = 0
        self._hyperlinks_done_links  = 0
        self._initial_broken         = None
        self._fixed_count            = 0
        self._initial_folder_count   = None
    
        if hasattr(self, "_apply_stopped_style_to_all_bars"):
            self._apply_stopped_style_to_all_bars(False)
    
        # terminal
        if getattr(self, 'terminal', None) is None or not self.terminal.isVisible():
            self.terminal = TerminalDialog(self)
            # ── MONKEY‐PATCH for live logging ──
            _orig_append = self.terminal.append_output
            def _live_append(text: str):
                # Parse first (no UI writes inside)
                try:
                    # Avoid double parsing if caller already parsed this line
                    if not getattr(self, "_skip_parse_in_monkeypatch", False):
                        self._parse_and_update_report(text)
                except Exception as e:
                    logging.exception("Report parser error: %s", e)
                # Then show in UI and log file
                _orig_append(text)
                logging.info(text)
            self.terminal.append_output = _live_append
            
    
        self.terminal.show()
        self.terminal.raise_()
    
        # start batch
        self.queue_active = True
        self.current_index = 0
        self._clear_queue_state()
        self.process_next_manufacturer()

    def _get_hyper_logs_dir(self) -> str:
        import os, logging
        try:
            for h in logging.getLogger().handlers:
                if hasattr(h, "baseFilename"):
                    d = os.path.dirname(h.baseFilename)
                    if d and os.path.isdir(d):
                        return d
        except Exception:
            pass
        d = os.path.join(os.getcwd(), "Hyper Logs")
        os.makedirs(d, exist_ok=True)
        return d
    
    def _write_hyper_report(self) -> str:
        import os, time, datetime, re  # re reserved if needed elsewhere
    
        if not hasattr(self, "report_stats"):
            self.report_stats = {}
        if not hasattr(self, "_report_year_totals"):
            self._report_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0}
    
        header_label = getattr(self, "_report_header_label", "ADAS SI PDF Document")
    
        def hms(seconds: int) -> str:
            h = seconds // 3600; m = (seconds % 3600) // 60; s = seconds % 60
            return f"{h}:{m:02d}:{s:02d}"
    
        def human_dhm(seconds: int) -> str:
            days = seconds // 86400; rem = seconds % 86400
            hrs = rem // 3600; mins = (rem % 3600) // 60
            parts = []
            if days: parts.append(f"{days} day{'s' if days != 1 else ''}")
            if hrs:  parts.append(f"{hrs} hour{'s' if hrs != 1 else ''}")
            parts.append(f"{mins} minute{'s' if mins != 1 else ''}")
            return ", ".join(parts)
    
        # ---- Year selection + labels (we'll reuse in recompute AND header/link print)
        try:
            # returns [(2012,2016), (2017,2021), ...] or [] if none checked
            selected_ranges = self.get_selected_year_ranges()
        except Exception:
            selected_ranges = []
    
        tuple_to_label = {
            (2012, 2016): "2012–2016",
            (2017, 2021): "2017–2021",
            (2022, 2026): "2022–2026",
        }
        # If none selected, treat as "all" (legacy behavior)
        if not selected_ranges:
            display_labels = ["2012–2016", "2017–2021", "2022–2026"]
        else:
            display_labels = [tuple_to_label[t] for t in selected_ranges if t in tuple_to_label]
    
        # Fixed legacy map used only when NO year filter is applied
        legacy_map_by_index = {0: "2012–2016", 1: "2017–2021", 2: "2022–2026"}
        valid_labels = {"2012–2016", "2017–2021", "2022–2026"}
    
        # helper: compute the intended label for a link by zero-based index
        def _label_for_link_index(zero_idx: int) -> str:
            # exactly one selected → force it
            if len(display_labels) == 1:
                return display_labels[0]
            # some selected (>=2) → map by index through selected list; if out of range, use last
            if len(display_labels) >= 2:
                if 0 <= zero_idx < len(display_labels):
                    return display_labels[zero_idx]
                return display_labels[-1]
            # none selected → legacy mapping
            return legacy_map_by_index.get(zero_idx, "")
    
        # 🔁 Self-heal: recompute year totals and normalize per-link labels
        recomputed_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0}
    
        for make, data in self.report_stats.items():
            links = data.get("links", []) or []
            # Fill per-make totals if missing
            if data.get("total_files", 0) == 0 and links:
                data["total_files"] = sum(int(l.get("files", 0)) for l in links)
            if data.get("total_time", 0) == 0 and links:
                data["total_time"] = sum(int(l.get("time", 0)) for l in links)  # heuristic
    
            # Normalize per-link ranges and re-accumulate year totals
            for idx, l in enumerate(links):  # idx is 0-based
                yr = (l.get("range") or "").strip()
    
                # If range present and valid, keep it; otherwise compute from selection/index
                if not yr or yr not in valid_labels:
                    yr = _label_for_link_index(idx)
                    if yr:
                        l["range"] = yr  # persist normalized label for print step
    
                if yr in recomputed_year_totals:
                    try:
                        recomputed_year_totals[yr] += int(l.get("files", 0))
                    except Exception:
                        pass
    
        # Replace live totals with the recomputed values (guarantees header consistency)
        self._report_year_totals = recomputed_year_totals
    
        # ---------- Write the file ----------
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = os.path.join(self._get_hyper_logs_dir(), f"ADAS_SI_Report_{ts}.txt")
        with open(filename, "w", encoding="utf-8") as f:
            # Header (dynamic by selected year ranges)
            grand_runtime = sum(int(v.get("total_time", 0)) for v in self.report_stats.values())
            f.write("Grand Totals (All Makes Combined)\n\n")
            f.write(f"Complete Runtime: {human_dhm(grand_runtime)} ({hms(grand_runtime)})\n")
    
            # Write only the selected bucket lines and compute total from those
            selected_total = 0
            for lbl in display_labels:
                count = int(self._report_year_totals.get(lbl, 0))
                f.write(f"{header_label} {lbl} Files: {count}\n")
                selected_total += count
    
            f.write(f"Total Files (All Years): {selected_total}\n\n")
            f.write("-" * 90 + "\n\n")
    
            # Per-make blocks
            def ord_label(i: int) -> str:
                return {1:"1st", 2:"2nd", 3:"3rd"}.get(i, f"{i}th")
    
            for make in sorted(self.report_stats.keys(), key=str.lower):
                data = self.report_stats[make]
                f.write(f"{make}\n\n")
                f.write(f"{hms(int(data.get('total_time', 0)))} Total Time | Total Files: {int(data.get('total_files', 0)):,}\n")
    
                # ---- Per-link rows (dynamic label per selection/index) ----
                for i, link in enumerate(data.get("links", []), start=1):
                    # i is 1-based for display; use (i-1) for zero-based index mapping
                    link_label = (link.get("range") or "").strip()
                    if not link_label or link_label not in valid_labels:
                        link_label = _label_for_link_index(i - 1)
                    t  = hms(int(link.get("time", 0)))
                    files = int(link.get("files", 0))
                    f.write(f"-----{ord_label(i)} {make} Link ({link_label}): {t} | Files: {files:,}\n")
                f.write("\n")
            f.flush(); os.fsync(f.fileno())
    
        msg = f"📄 Report written to: {os.path.abspath(filename)} (exists={os.path.exists(filename)})"
        print(msg)
        try:
            if getattr(self, "terminal", None):
                self.terminal.append_output(msg)
        except Exception:
            pass
        return filename
  
    def _try_write_report_once(self, reason: str = "") -> None:
        """Write the report only once per batch, with backfill from the latest log if needed."""
        if getattr(self, "_report_written", False):
            return
        try:
            # 🔄 Always attempt a backfill from the latest log so counts/times are present
            self._backfill_report_from_log()  # safe no-op if nothing found
    
            fn = self._write_hyper_report()
            self._report_written = True
            note = f"✅ Report saved ({reason}): {fn}"
            print(note)
            if getattr(self, "terminal", None):
                self.terminal.append_output(note)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            err = f"❌ Report write failed ({reason}): {e}\n{tb}"
            print(err)
            if getattr(self, "terminal", None):
                self.terminal.append_output(err)


    def _parse_and_update_report(self, line: str) -> None:
        """
        Parse a single log line to update progress bars and the dynamic report.
        This function NEVER writes to the terminal; it's parse-only.
        """
        import re
    
        # 0) Progress hooks ----------------------------------------------------
        m_cm = re.match(r"\s*CM_PROGRESS\s+(\d+)\s*/\s*(\d+)\s*\((\d+)%\)", line, re.IGNORECASE)
        if m_cm:
            done = int(m_cm.group(1))
            total = max(1, int(m_cm.group(2)))
            pct_from_text = int(m_cm.group(3))
            pct = max(0, min(100, pct_from_text if 0 <= pct_from_text <= 100 else int(done/total*100)))
            self.current_manufacturer_progress.setValue(pct)
            return
    
        if getattr(self, '_cleanup_mode', False):
            m_total = re.search(r'Total broken hyperlinks:\s*(\d+)', line)
            if m_total:
                self._initial_broken = int(m_total.group(1))
                self._fixed_count = 0
                return
            if line.startswith(("Fixed hyperlink for", "✅", "❌")) and getattr(self, "_initial_broken", None):
                self._fixed_count += 1
                pct = int(self._fixed_count / max(1, self._initial_broken) * 100)
                self.current_manufacturer_progress.setValue(pct)
                return
    
        m_fr = re.search(r'(\d+)\s+Folders Remain', line)
        if m_fr:
            remaining = int(m_fr.group(1))
            if not hasattr(self, '_initial_folder_count') or self._initial_folder_count is None:
                self._initial_folder_count = remaining
            else:
                self._initial_folder_count = max(self._initial_folder_count, remaining)
            initial = max(1, self._initial_folder_count)
            pct = max(0, min(100, int((initial - remaining) / initial * 100)))
            self.current_manufacturer_progress.setValue(pct)
    
        # 1) Reporting buckets -------------------------------------------------
        if not hasattr(self, "report_stats"):
            self.report_stats = {}
        if not hasattr(self, "_report_year_totals"):
            self._report_year_totals = {"2012–2016": 0, "2017–2021": 0, "2022–2026": 0}
    
        current_make = getattr(self, "_report_current_make", None)
        if not current_make and hasattr(self, "selected_manufacturers") and 0 <= self.current_index < len(self.selected_manufacturers):
            current_make = self.selected_manufacturers[self.current_index]
        if not current_make:
            return
    
        # 2) Helpers -----------------------------------------------------------
        def _norm_range(rng: str) -> str:
            rng = rng.strip().replace("—", "-").replace("–", "-")
            m = re.search(r"(\d{4})\s*-\s*(\d{4})", rng)
            if not m:
                return rng.strip()
            a, b = int(m.group(1)), int(m.group(2))
            return f"{a}–{b}"
    
        def _hms_to_seconds(s: str) -> int:
            parts = [int(x) for x in s.split(":")]
            if len(parts) == 3:
                h, m, sec = parts
            elif len(parts) == 2:
                h, m, sec = 0, parts[0], parts[1]
            else:
                h, m, sec = 0, 0, parts[0]
            return h * 3600 + m * 60 + sec
    
        def _int(s: str) -> int:
            return int(str(s).replace(",", "").strip())
    
        # We’ll try a few tolerant formats. If a per-link line omits the year
        # we’ll fill it using the running link index (1→2012–2016, 2→2017–2021, 3→2022–2026).
        RE_LINK_CANON = re.compile(
            r"-{2,}\s*(\d+)(?:st|nd|rd|th)\s+[A-Za-z][A-Za-z &\-/]+?\s+Link\s*\(([^)]*)\)\s*[:\-–]\s*"
            r"(\d{1,2}:\d{2}(?::\d{2})?)\s*\|\s*(?:Files(?:\s*(?:Found|Indexed))?|Files)\s*[:=]?\s*(\d{1,3}(?:,\d{3})*)",
            re.IGNORECASE
        )
        RE_YEAR   = re.compile(r"\((\s*\d{4}\s*[-–—]\s*\d{4}\s*)\)")
        RE_TIME   = re.compile(r"\b(\d{1,2}:\d{2}(?::\d{2})?)\b")
        RE_FILES1 = re.compile(r"(?:Files(?:\s*(?:Found|Indexed))?|Found)\s*[:=]?\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE)
        RE_FILES2 = re.compile(r"\b(\d{1,3}(?:,\d{3})*)\s+Files\b", re.IGNORECASE)
    
        # 3) Per-link (canonical)
        m = RE_LINK_CANON.search(line)
        if m:
            ord_idx_str, yr_raw, rt_str, files_str = m.groups()
            ord_idx = max(1, int(ord_idx_str))
            files   = _int(files_str)
            secs    = _hms_to_seconds(rt_str)
    
            # fallback year if missing in the line
            idx_map = {1: "2012–2016", 2: "2017–2021", 3: "2022–2026"}
            yr = _norm_range(yr_raw) if yr_raw.strip() else idx_map.get(ord_idx, "")
    
            self.report_stats.setdefault(current_make, {"total_time": 0, "total_files": 0, "links": []})
            sig = (yr, secs, files)
            existing = {(l.get("range",""), l.get("time",0), l.get("files",0)) for l in self.report_stats[current_make]["links"]}
            if sig not in existing:
                self.report_stats[current_make]["links"].append({"range": yr, "time": secs, "files": files})
                if yr in self._report_year_totals:
                    self._report_year_totals[yr] += files
            return
    
        # 4) Per-link (fallback: year-range + time + files scattered)
        m_year = RE_YEAR.search(line)
        if m_year and (RE_FILES1.search(line) or RE_FILES2.search(line)) and RE_TIME.search(line):
            yr = _norm_range(m_year.group(1))
            times = RE_TIME.findall(line)
            rt_str = times[-1] if times else "0:00"
            files_m = RE_FILES1.search(line) or RE_FILES2.search(line)
            files = _int(files_m.group(1)) if files_m else 0
            secs = _hms_to_seconds(rt_str)
    
            # If the year still somehow empty, map by index from the running sub-link index.
            if not yr:
                try:
                    # _multi_link_index is 0-based for in-flight link; report lines are printed after finish,
                    # so add 1 to get human-ordinal.
                    ord_idx = int(getattr(self, "_multi_link_index", 0)) + 1
                    yr = {1:"2012–2016", 2:"2017–2021", 3:"2022–2026"}.get(ord_idx, "")
                except Exception:
                    pass
    
            self.report_stats.setdefault(current_make, {"total_time": 0, "total_files": 0, "links": []})
            sig = (yr, secs, files)
            existing = {(l.get("range",""), l.get("time",0), l.get("files",0)) for l in self.report_stats[current_make]["links"]}
            if sig not in existing:
                self.report_stats[current_make]["links"].append({"range": yr, "time": secs, "files": files})
                if yr in self._report_year_totals:
                    self._report_year_totals[yr] += files
            return
    
        # 5) Totals (per-make) from various orders
        for pat in (
            re.compile(r"(\d{1,2}:\d{2}(?::\d{2})?)\s+Total Time\s*\|\s*Total Files\s*:\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE),
            re.compile(r"Total Files\s*:\s*(\d{1,3}(?:,\d{3})*)\s*\|\s*Total Time\s*:\s*(\d{1,2}:\d{2}(?::\d{2})?)", re.IGNORECASE),
            re.compile(r"Total Time\s*:\s*(\d{1,2}:\d{2}(?::\d{2})?)", re.IGNORECASE),
            re.compile(r"(?:Total Files|Files Indexed|Files Found)\s*:\s*(\d{1,3}(?:,\d{3})*)", re.IGNORECASE),
            re.compile(r"Indexing routine took\s*(\d+(?:\.\d+)?)\s*seconds", re.IGNORECASE),
        ):
            m = pat.search(line)
            if not m:
                continue
            self.report_stats.setdefault(current_make, {"total_time": 0, "total_files": 0, "links": []})
            if pat.pattern.startswith("("):  # time | files
                t, f = m.groups(); self.report_stats[current_make]["total_time"] = max(self.report_stats[current_make]["total_time"], _hms_to_seconds(t)); self.report_stats[current_make]["total_files"] = max(self.report_stats[current_make]["total_files"], _int(f)); break
            if "Total Files" in pat.pattern and "Total Time" in pat.pattern:  # files | time
                f, t = m.groups(); self.report_stats[current_make]["total_time"] = max(self.report_stats[current_make]["total_time"], _hms_to_seconds(t)); self.report_stats[current_make]["total_files"] = max(self.report_stats[current_make]["total_files"], _int(f)); break
            if "Total Time" in pat.pattern:  # time only
                t = m.group(1); self.report_stats[current_make]["total_time"] = max(self.report_stats[current_make]["total_time"], _hms_to_seconds(t)); break
            if "Total Files" in pat.pattern or "Files Indexed" in pat.pattern or "Files Found" in pat.pattern:
                f = m.group(1); self.report_stats[current_make]["total_files"] = max(self.report_stats[current_make]["total_files"], _int(f)); break
            if "Indexing routine took" in pat.pattern:
                secs = int(float(m.group(1))); self.report_stats[current_make]["total_time"] = max(self.report_stats[current_make]["total_time"], secs); break
    
        # Backfill per-make files if totals seen but zero
        data = self.report_stats.get(current_make, {})
        if data.get("total_files", 0) == 0 and data.get("links"):
            data["total_files"] = sum(int(l.get("files", 0)) for l in data["links"])


    def process_next_manufacturer(self):
        import re
    
        # ── HARD STOPS ─────────────────────────────────────────────────────
        if getattr(self, "stop_requested", False):
            # Write whatever we have so far on manual stop
            self._try_write_report_once("manual stop")
            return
    
        # Only proceed if we’re in an active batch (prevents stray calls)
        if not getattr(self, "queue_active", False):
            return
        # ──────────────────────────────────────────────────────────────────
    
        # ── INDEX/BOUNDS GUARD ────────────────────────────────────────────
        total = len(self.selected_manufacturers)
        if self.current_index >= total:
            # 🆕 If cleanup mode, run final unresolved broken link removal
            if self.cleanup_checkbox.isChecked():
                try:
                    if hasattr(self, "extractor") and hasattr(self.extractor, "broken_entries"):
                        #if self.extractor.repair_mode and self.extractor.excel_mode == "og":
                        #    hyperlink_col = 8
                        #elif not self.extractor.repair_mode and self.extractor.excel_mode == "og":
                        #    hyperlink_col = 12
                        #elif not self.extractor.repair_mode and self.extractor.excel_mode == "new":
                        #    hyperlink_col = 11
                        #else:
                        #    hyperlink_col = None
    
                        if hyperlink_col:
                            print("🧹 Finalizing cleanup — removing unresolved broken links...")
                            import openpyxl
                            wb = openpyxl.load_workbook(self.excel_paths[0])
                            ws = wb['Model Version']
                            removed_count = 0
                            for row, (yr, mk, mdl, sys) in self.extractor.broken_entries:
                                cell = ws.cell(row=row, column=hyperlink_col)
                                link_to_test = (
                                    cell.hyperlink.target if cell.hyperlink
                                    else (str(cell.value).strip() if cell.value else "")
                                )
                                if not link_to_test or link_to_test.lower() == "hyperlink not available":
                                    continue
                                if self.extractor.is_broken_sharepoint_link(link_to_test, file_name=sys):
                                    print(f"🗑 Removing unresolved broken link at row {row}: {link_to_test}")
                                    cell.value = None
                                    cell.hyperlink = None
                                    removed_count += 1
                            wb.save(self.excel_paths[0])
                            wb.close()
                            print(f"✅ Cleanup complete — removed {removed_count} unresolved links")
                except Exception as e:
                    print(f"⚠️ Final cleanup pass failed: {e}")
    
            # 🏁 Done with the batch — cleanly wrap up and reset state
            completed = "\n".join(sorted(self.selected_manufacturers, key=str.lower))
            QMessageBox.information(
                self,
                'Completed',
                f"The Following Manufacturers have been completed:\n{completed}",
                QMessageBox.Ok
            )
    
            # ✅ ALWAYS write report here (only once)
            self._try_write_report_once("batch complete")
    

        # ──────────────────────────────────────────────────────────────────
    
        # Reset per‐manufacturer progress tracking
        self._initial_folder_count = None
    
        manufacturer = self.selected_manufacturers[self.current_index]
        # REPORT: remember which make's lines we are parsing right now
        self._report_current_make = manufacturer
        self.report_stats.setdefault(manufacturer, {"total_time": 0, "total_files": 0, "links": []})
    
        self.current_manufacturer_label.setText(f"Current Manufacturer: {manufacturer}")
        self.current_manufacturer_progress.setValue(0)
    
        # ── SAFELY FETCH PARALLEL LIST ENTRY ──────────────────────────────
        if self.current_index >= len(self.excel_paths):
            print("⚠️ excel_paths shorter than selected_manufacturers; aborting this step.")
            self.current_index += 1
            return self.process_next_manufacturer()
        excel_path = self.excel_paths[self.current_index]
        # ──────────────────────────────────────────────────────────────────
    
        link_dict = self.repair_links if self.mode_flag == "repair" else self.manufacturer_links
        sharepoint_links = link_dict.get(manufacturer)
        if not sharepoint_links:
            QMessageBox.warning(
                self,
                'Error',
                f"No SharePoint link found for {manufacturer} in {self.mode_flag} mode.",
                QMessageBox.Ok
            )
            self.current_index += 1
            return self.process_next_manufacturer()
    
        if isinstance(sharepoint_links, str):
            sharepoint_links = [sharepoint_links]
    
        # === NEW: honor Year Ranges checkboxes (regex first; index fallback 0/1/2) ===
        sharepoint_links = self._filter_links_by_selected_years(sharepoint_links)
        if not sharepoint_links:
            self.terminal.append_output(f"ℹ️ No links match selected years for {manufacturer}; skipping.")
            self.on_manufacturer_finished(manufacturer, True)
            return
    
        # Cleanup mode: intersect with broken-link years (keep previous set if no regex match)
        if self.cleanup_checkbox.isChecked():
            years_needed = self.get_broken_hyperlink_years_for_manufacturer(manufacturer)
            if years_needed:
                filtered_links = []
                for link in sharepoint_links:
                    m = re.search(r'\((\d{4})\s*-\s*(\d{4})\)', link)
                    if m:
                        start_year, end_year = int(m.group(1)), int(m.group(2))
                        if any(start_year <= y <= end_year for y in years_needed):
                            filtered_links.append(link)
                if filtered_links:
                    sharepoint_links = filtered_links
    
        # ── set up multi-link run state ────────────────────────────────────
        self._multi_links        = sharepoint_links
        self._multi_link_index   = 0
        self._multi_excel_path   = excel_path
        self._multi_manufacturer = manufacturer
    
        self._hyperlinks_total_links = len(self._multi_links)
        self._hyperlinks_done_links  = 0
        self.update_manufacturer_progress_bar()
    
        self._cleanup_mode = self.cleanup_checkbox.isChecked()
        if self._cleanup_mode:
            self._initial_broken = None
            self._fixed_count    = 0
    
        # Kick it off
        self.run_all_links_batch() if self._cleanup_mode else self.run_next_sub_link()

    def _extract_year_range_label(link: str, index: int | None = None) -> str:
        """
        Pull a '(YYYY - YYYY)' or '(YYYY–YYYY)' suffix out of the link text.
        Falls back to index mapping (0/1/2/3) → 2012–2016 / 2017–2021 / 2022–2026 / 2027–2031.
        Returns an empty string if unknown.
        """
        import re
    
        # try a suffix like "... (2012 - 2016)" or "(2017-2021)" or with en dash
        m = re.search(r'\(\s*(20\d{2})\s*[–-]\s*(20\d{2})\s*\)\s*$', link)
        if not m:
            # try anywhere in the string inside parentheses, just in case
            m = re.search(r'\((?:[^)]*?)(20\d{2})\s*[–-]\s*(20\d{2})(?:[^)]*?)\)', link)
    
        if m:
            return f"{m.group(1)}–{m.group(2)}"
    
        # index fallback (0-based) — now includes the 4th range
        mapping = {0: "2012–2016", 1: "2017–2021", 2: "2022–2026", 3: "2027–2031"}  # NEW
        if index is not None and index in mapping:
            return mapping[index]
    
        return ""
    
    
    def _links_for_manufacturer_preview(self, manufacturer: str):
        """Return the exact list of SharePoint links this run will use for `manufacturer`.
        Mirrors the logic in process_next_manufacturer(), including cleanup-mode filtering."""
        link_dict = self.repair_links if self.mode_flag == "repair" else self.manufacturer_links
        links = link_dict.get(manufacturer) or []
        if isinstance(links, str):
            links = [links]
            
        # ⬇️ NEW: apply UI year-range filter for both normal & cleanup previews
        links = self._filter_links_by_selected_years(links)
    
        if self.cleanup_checkbox.isChecked():
            years_needed = self.get_broken_hyperlink_years_for_manufacturer(manufacturer)
            if years_needed:
                filtered = []
                for link in links:
                    m = re.search(r'\((\d{4})\s*-\s*(\d{4})\)', link)
                    if m:
                        start_y, end_y = int(m.group(1)), int(m.group(2))
                        if any(start_y <= y <= end_y for y in years_needed):
                            filtered.append(link)
                    else:
                        # If no range on the link, keep it (conservative) so count never under-reports
                        filtered.append(link)
                if filtered:
                    links = filtered
        return links
    

    def _clear_queue_state(self):
        """Reset per-run tracking to avoid carryover between batches."""
        try:
            self.manufacturers_completed = []
        except Exception:
            pass
        try:
            self.manufacturers_failed = []
        except Exception:
            pass
        # Stop any pending next-timer you might have elsewhere
        if getattr(self, "_next_timer", None):
            try:
                self._next_timer.stop()
            except Exception:
                pass
            self._next_timer = None
    
    def _log_all_done(self):
        print("🏁 All Manufacturers finished.")
        # Also reset/disable any UI as you already do when everything completes
        
    def _schedule_next_manufacturer(self):
        # stop previous timer if it exists
        if self._next_timer:
            try:
                self._next_timer.stop()
            except Exception:
                pass
            self._next_timer = None
    
        self._next_timer = QTimer(self)
        self._next_timer.setSingleShot(True)
        self._next_timer.timeout.connect(self.process_next_manufacturer)
        print("⏱ Checking in 10s if i Need to run another Manufacturer…")
        self._next_timer.start(10_000)
        
    def run_next_sub_link(self):
        self._batch_links_mode = False
        if self._multi_link_index >= len(self._multi_links):
            # All links processed for this manufacturer
            self.on_manufacturer_finished(self._multi_manufacturer, True)
            return
    
        # Reset Current Manufacturer progress bar for this sub-link
        self.current_manufacturer_progress.setValue(0)
        self._initial_folder_count = None  # 🆕 Reset baseline for "Folders Remain"
    
        script_path = os.path.join(os.path.dirname(__file__), "SharepointExtractor.py")
        #excel_mode = "new" if self.excel_mode_switch.isChecked() else "og"
    
        current_link = self._multi_links[self._multi_link_index]
        args = [
            sys.executable,
            script_path,
            current_link,
            self._multi_excel_path,
            ",".join(self.selected_systems),
            self.mode_flag,
            "cleanup" if self.cleanup_checkbox.isChecked() else "full",
            #excel_mode
        ]
    
        self._cleanup_mode = self.cleanup_checkbox.isChecked()
        if self._cleanup_mode:
            self._initial_broken = None
            self._fixed_count = 0
    
        self.update_manufacturer_progress_bar()
    
        thread = WorkerThread(args, self._multi_manufacturer, parent=self)
        self.thread = thread
        thread.output_signal.connect(self.handle_extractor_output)
        thread.finished_signal.connect(self.on_sub_link_finished)
        thread.start()
        self.threads.append(thread)
        
    def run_all_links_batch(self):
        """
        Cleanup Mode: launch ONE extractor process passing ALL SharePoint links joined by '||'.
        This avoids per-link rescans and lets the extractor try all links in a single pass.
        """
        # Reset Current Manufacturer progress bar at batch start
        self.current_manufacturer_progress.setValue(0)
        self._initial_folder_count = None

        script_path = os.path.join(os.path.dirname(__file__), "SharepointExtractor.py")
        #excel_mode = "new" if self.excel_mode_switch.isChecked() else "og"

        # Join all links for this manufacturer into a single argument
        all_links_arg = "||".join(self._multi_links)

        args = [
            sys.executable,
            script_path,
            all_links_arg,                      # ← multiple links combined
            self._multi_excel_path,
            ",".join(self.selected_systems) if hasattr(self, "selected_systems") else "",
            self.mode_flag,
            "cleanup",
            #excel_mode
        ]

        # Batch mode marker so we can adjust hyperlink progress semantics
        self._batch_links_mode = True

        # Reset cleanup counters
        self._cleanup_mode = True
        self._initial_broken = None
        self._fixed_count = 0

        self.update_manufacturer_progress_bar()

        thread = WorkerThread(args, self._multi_manufacturer, parent=self)
        self.thread = thread
        thread.output_signal.connect(self.handle_extractor_output)
        # In batch mode we finish the whole manufacturer on single thread end
        thread.finished_signal.connect(self.on_manufacturer_finished)
        thread.start()
        self.threads.append(thread)


    def finalize_cleanup_for_file(self, excel_path, broken_entries, hyperlink_col):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Model Version']
    
        removed_count = 0
        for row, (yr, mk, mdl, sys) in broken_entries:
            cell = ws.cell(row=row, column=hyperlink_col)
    
            # Prefer the actual hyperlink target if it exists
            link_to_test = cell.hyperlink.target if cell.hyperlink else str(cell.value).strip() if cell.value else ""
    
            # Skip placeholders or empty links
            if not link_to_test or link_to_test.lower() == "hyperlink not available":
                continue
    
            # Check link validity
            if self.extractor.is_broken_sharepoint_link(link_to_test, file_name=sys):
                print(f"🗑 Removing unresolved broken link at row {row}: {link_to_test}")
                cell.value = None
                cell.hyperlink = None
                removed_count += 1
    
        wb.save(excel_path)
        wb.close()
        print(f"✅ Cleanup complete — removed {removed_count} unresolved links")
           
    def count_expected_hyperlinks_for_link(self, manufacturer, sharepoint_link):
        from openpyxl import load_workbook
        import re
    
        # Extract year range from link (assuming in format "(2012 - 2016)")
        m = re.search(r'\((\d{4})\s*-\s*(\d{4})\)', sharepoint_link)
        year_range = None
        if m:
            year_range = (int(m.group(1)), int(m.group(2)))
    
        wb = load_workbook(self.excel_paths[0])
        ws = wb.active
        count = 0
    
        for row in ws.iter_rows(min_row=2):
            year_val = row[0].value
            make_val = str(row[1].value).strip().lower() if row[1].value else ""
            link_val = row[11].hyperlink if len(row) > 11 else None  # Column L
    
            if make_val == manufacturer.lower():
                if year_range and isinstance(year_val, int) and not (year_range[0] <= year_val <= year_range[1]):
                    continue
                if not link_val:
                    count += 1
    
        return count
    
    def _extract_year_range_from_link(self, link: str) -> str:
        """
        Extract 'YYYY–YYYY' from the link text (anywhere).
        Returns '' if not found.
        """

        m = re.search(r'\(\s*(20\d{2})\s*[-–]\s*(20\d{2})\s*\)', link)
        if m:
            return f"{m.group(1)}–{m.group(2)}"
        return ""

    # ────────────────────────────────────────────────
    # 🧭 Header Column Utility
    # ────────────────────────────────────────────────
    def _header_col_index(self, ws, *names):
        """
        Find a header by name(s) in row 1 and return its 1-based column index.
        Example:
            idx = self._header_col_index(ws, "Hyperlink", "Link", "URL")
        Returns None if no match found.
        """
        header = next(ws.iter_rows(min_row=1, max_row=1))
        hmap = { (str(c.value).strip().upper() if c.value else ""): i + 1
                 for i, c in enumerate(header) }
        for n in names:
            if n.upper() in hmap:
                return hmap[n.upper()]
        return None
                
    def on_sub_link_finished(self, manufacturer, success):
        if not self.is_running:
            return
    
        if success:
            msg = f"✅ Finished SharePoint link {self._multi_link_index+1}/{len(self._multi_links)} for {manufacturer}"
            self.terminal.append_output(msg)
            logging.info(msg)
    
            # Increment completed links count & update bar
            self._hyperlinks_done_links += 1
            self.update_manufacturer_progress_bar()
    
            self._multi_link_index += 1
    
            if self._multi_link_index >= len(self._multi_links):
                # All links for this manufacturer complete
                self.manufacturer_hyperlink_label.setText("Manufacturer Hyperlinks: Complete")
                self.manufacturer_hyperlink_bar.setValue(self.manufacturer_hyperlink_bar.maximum())
                self.on_manufacturer_finished(manufacturer, True)
            else:
                # Move to next sub-link
                self.run_all_links_batch() if self.cleanup_checkbox.isChecked() else self.run_next_sub_link()
    
        else:
            msg = f"❌ SharePoint link {self._multi_link_index+1}/{len(self._multi_links)} for {manufacturer} failed"
            self.terminal.append_output(msg)
            logging.warning(msg)
        
            # NEW: treat as incomplete but keep going
            self._hyperlinks_done_links += 1
            self.update_manufacturer_progress_bar()
            self._multi_link_index += 1
        
            if self._multi_link_index >= len(self._multi_links):
                # Finished this manufacturer (some links may be incomplete)
                self.manufacturer_hyperlink_label.setText("Manufacturer Hyperlinks: Complete")
                self.manufacturer_hyperlink_bar.setValue(self.manufacturer_hyperlink_bar.maximum())
                self.on_manufacturer_finished(manufacturer, True)
            else:
                # Try the next sub-link
                self.run_all_links_batch() if self.cleanup_checkbox.isChecked() else self.run_next_sub_link()

    
    def update_manufacturer_progress_bar(self):
        total_links = getattr(self, "_hyperlinks_total_links", 1)
        done_links = getattr(self, "_hyperlinks_done_links", 0)
    
        # Keep bar in sync
        self.manufacturer_hyperlink_bar.setMaximum(total_links)
        self.manufacturer_hyperlink_bar.setValue(done_links)
    
        # Always show x/y format
        self.manufacturer_hyperlink_label.setText(
            f"Manufacturer Hyperlinks: {done_links} / {total_links}"
        )
      
    # --- Quick syntax/tenant check used by the GUI pre-scan ---
    def is_broken_sharepoint_link(self, url: str) -> bool:
        """
        Cheap/fast check (no Selenium): flags obviously bad links so we can
        pick years to re-run. The extractor will do the real validation later.
        """
        try:
            if not url or not isinstance(url, str):
                return True
            url = url.strip()
            if not url.lower().startswith("http"):
                return True
    
            from urllib.parse import urlparse, parse_qs
            pu = urlparse(url)
    
            # Must have scheme + host
            if not pu.scheme or not pu.netloc:
                return True
    
            # Restrict to your tenant hosts (adjust if you have others)
            allowed_hosts = {
                "calibercollision.sharepoint.com",
                "calibercollision-my.sharepoint.com",
            }
            if pu.netloc.lower() not in allowed_hosts:
                return True
    
            # If it's an AllItems.aspx-style link, make sure the id= param is present
            if pu.path.lower().endswith("allitems.aspx"):
                qs = parse_qs(pu.query or "")
                if "id" not in qs:
                    return True
    
            # Looks syntactically fine; extractor will verify live later
            return False
    
        except Exception:
            # Any parsing error -> treat as broken
            return True
    
        
    def get_broken_hyperlink_years_for_manufacturer(self, manufacturer):
        years = set()
        from openpyxl import load_workbook
    
        wb = load_workbook(self.excel_paths[0])
        ws = wb.active
    
        for row in ws.iter_rows(min_row=2):
            year_val = row[0].value  # Assuming column A = Year
            make_val = str(row[1].value).strip().lower() if row[1].value else ""
            link_val = row[11].hyperlink if len(row) > 11 else None  # Assuming column L for hyperlink
    
            if make_val == manufacturer.lower():
                # No hyperlink OR broken link
                if not link_val or self.is_broken_sharepoint_link(link_val.target):
                    try:
                        years.add(int(year_val))
                    except (TypeError, ValueError):
                        pass
    
        return sorted(years)
    
    def on_manufacturer_finished(self, manufacturer, success):
        # ── Reentrancy guard: prevents double entry from duplicate signals ──
        if getattr(self, "_finish_guard", False):
            return
        self._finish_guard = True
        try:
            # If we ran in batch-links mode, mark all links done for the hyperlink counter
            if getattr(self, '_batch_links_mode', False):
                self._hyperlinks_done_links = self._hyperlinks_total_links
                self.update_manufacturer_progress_bar()
                self.manufacturer_hyperlink_label.setText('Manufacturer Hyperlinks: Complete')
    
            # ── HARD BAIL-OUT: if we're not running, stop immediately ──
            if not self.is_running:
                # Reset UI to Stopped state
                self.current_manufacturer_progress.setValue(0)
                self.overall_progress_bar.setValue(0)
                self.current_manufacturer_label.setText("Current Manufacturer: Manually Stopped")
                self.overall_progress_label.setText("Overall Progress: Manually Stopped")
                self.manufacturer_hyperlink_label.setText('Manufacturer Hyperlinks: Manually Stopped')

                    # 🆕 force bars into red "stopped" style
                self._apply_stopped_style_to_all_bars(True)
    
                # Swap back to Start button
                layout = self.button_layout
                layout.removeWidget(self.start_button)
                self.start_button.deleteLater()
                self.start_button = CustomButton("Start Automation", "#008000", self)
                self.start_button.clicked.connect(self.on_start_stop)
                layout.addWidget(self.start_button)
    
                # Disable Pause
                self.pause_button.setEnabled(False)
                return
    
            # ── YOUR ORIGINAL LOGIC (kept) ──────────────────────────────────
    
            # 1) count this run
            prev = self.attempts.get(manufacturer, 0)
            self.attempts[manufacturer] = prev + 1
            attempt_no = self.attempts[manufacturer]
    
            # 2) route based on success / attempt count
            if success:
                # Avoid duplicate entries if a stray double-finish happens
                if manufacturer not in self.completed_manufacturers:
                    self.completed_manufacturers.append(manufacturer)
                msg = f"✅ {manufacturer} succeeded on attempt {attempt_no}."
                self.terminal.append_output(msg)
                logging.info(msg)
    
                # update overall on success
                finalized = len(self.completed_manufacturers) + len(self.given_up_manufacturers)
                percent   = int(finalized / self.total_manufacturers * 100) if self.total_manufacturers else 100
                self.overall_progress_bar.setValue(percent)
                self.overall_progress_label.setText(
                    f"Overall Progress: {finalized} / {self.total_manufacturers}"
                )
            else:
                if attempt_no < self.max_attempts:
                    # Record failure for retry pass (keep lists in sync, avoid dupes)
                    err_excel = self.excel_paths[self.current_index] if self.current_index < len(self.excel_paths) else None
                    if manufacturer not in self.failed_manufacturers:
                        self.failed_manufacturers.append(manufacturer)
                    if err_excel and (err_excel not in self.failed_excels):
                        self.failed_excels.append(err_excel)
                    msg = f"❗ {manufacturer} failed on attempt {attempt_no}; will retry later."
                    self.terminal.append_output(msg)
                    logging.warning(msg)
                else:
                    if manufacturer not in self.given_up_manufacturers:
                        self.given_up_manufacturers.append(manufacturer)
                    msg = (
                        f"❌ {manufacturer} failed on attempt {attempt_no}; "
                        f"giving up after {self.max_attempts} tries."
                    )
                    self.terminal.append_output(msg)
                    logging.error(msg)
    
                    # update overall on final give-up
                    finalized = len(self.completed_manufacturers) + len(self.given_up_manufacturers)
                    percent   = int(finalized / self.total_manufacturers * 100) if self.total_manufacturers else 100
                    self.overall_progress_bar.setValue(percent)
                    self.overall_progress_label.setText(
                        f"Overall Progress: {finalized} / {self.total_manufacturers}"
                    )
    
            # 3) Debounced, UI-safe delay (replaces sleep(10))
            msg = "⏱ Checking in 10s if i Need to run another Manufacturer…"
            self.terminal.append_output(msg)
            logging.info(msg)
    
            def _continue_after_delay():
                # Advance index ONCE here (not before the delay)
                self.current_index += 1
    
                # 4) if still in this pass, keep going
                if self.current_index < len(self.selected_manufacturers):
                    self.process_next_manufacturer()
                    return
    
                # 5) end of pass: retry logic…
                if self.failed_manufacturers:
                    retry_list = ", ".join(self.failed_manufacturers)
                    self.terminal.append_output(f"🔄 Retrying: {retry_list}")
    
                    # Small extra debounce so logs render nicely
                    QTimer.singleShot(200, lambda: None)
    
                    # swap in failed sets, reset index, and go again
                    self.selected_manufacturers = list(self.failed_manufacturers)
                    self.excel_paths            = list(self.failed_excels)
                    self.current_index          = 0
                    self.failed_manufacturers   = []
                    self.failed_excels          = []
                    self.process_next_manufacturer()
                    return
    
                # 6) final summary
                completed_sorted = sorted(self.completed_manufacturers, key=str.lower)
                given_up_sorted  = sorted(self.given_up_manufacturers,  key=str.lower)
    
                self.terminal.append_output("")
                self.terminal.append_output("🏁 All Manufacturers finished.")
                self.terminal.append_output(f"✅ Completed: {', '.join(completed_sorted)}" if completed_sorted else "✅ Completed: ")
                self.terminal.append_output(f"❌ Gave up:   {', '.join(given_up_sorted)}" if given_up_sorted else "❌ Gave up:   ")
    
                            # >>> NEW <<< Write the final dynamic report once the entire batch is finished
                try:
                    self._try_write_report_once("batch complete")
                except Exception as e:
                    print(f"⚠️ Could not write final report: {e}")

                # lock bars at 100%
                self.current_manufacturer_progress.setValue(100)
                self.overall_progress_bar.setValue(100)
                self.current_manufacturer_label.setText("Current Manufacturer: Complete")
                self.overall_progress_label.setText("Overall Progress: Complete")
                self.terminal.append_output("=" * 68)
    
                # reset tracking for next run
                self.completed_manufacturers    = []
                self.failed_manufacturers       = []
                self.failed_excels              = []
                self.given_up_manufacturers     = []
                self.attempts                   = {}
    
                # swap back to Start button
                layout = self.button_layout
                layout.removeWidget(self.start_button)
                self.start_button.deleteLater()
                self.start_button = CustomButton("Start Automation", "#008000", self)
                self.start_button.clicked.connect(self.on_start_stop)
                layout.addWidget(self.start_button)
                self.pause_button.setEnabled(False)
                self.is_running = False
    
            # Fire the continuation in 10 seconds without blocking the UI
            QTimer.singleShot(10_000, _continue_after_delay)
    
        finally:
            # Release the guard so future *distinct* completions can run
            self._finish_guard = False


    def select_all(self):
        select_all_checked = True
        for i in range(self.manufacturer_tree.topLevelItemCount()):
            item = self.manufacturer_tree.topLevelItem(i)
            if item.checkState(0) != Qt.Checked:
                select_all_checked = False
                break
        
        for i in range(self.manufacturer_tree.topLevelItemCount()):
            item = self.manufacturer_tree.topLevelItem(i)
            item.setCheckState(0, Qt.Checked if not select_all_checked else Qt.Unchecked)            

    def closeEvent(self, event):
        # when the GUI closes, dump the terminal contents (if any) to Documents/Logs
        super().closeEvent(event)
     
    def _set_bar_stopped(self, bar, stopped: bool, name_hint: str):
        """
        Make a single QProgressBar turn red when stopped, while preserving its exact original look.
        Uses a per-widget stylesheet with higher specificity and !important so global styles can't override it.
        """
        if not bar:
            return
    
        # Cache original style once
        if not hasattr(bar, "_orig_stylesheet"):
            bar._orig_stylesheet = bar.styleSheet() or ""
    
        # Ensure we can target this bar specifically
        if not bar.objectName():
            bar.setObjectName(name_hint)
    
        if stopped:
            base = bar._orig_stylesheet
            obj  = bar.objectName()
            # Per‑widget rules (highest precedence) + !important
            red_css = (
                f'QProgressBar#{obj} {{ '
                f'  background-color: #470000 !important; '   # deep red even at 0%
                f'  color: white !important; '
                f'  border: 1px solid #aa4444; '
                f'  border-radius: 4px; '
                f'}} '
                f'QProgressBar#{obj}::chunk {{ '
                f'  background-color: #e53935 !important; '   # red fill when >0%
                f'}}'
            )
            bar.setStyleSheet(base + "\n" + red_css)
        else:
            # Restore exactly what the bar had before
            bar.setStyleSheet(bar._orig_stylesheet)
    
        # Force immediate repaint
        bar.style().unpolish(bar)
        bar.style().polish(bar)
        bar.update()
    
    def _style_bar(self, bar, *, stopped: bool, name_hint: str):
        """Per-widget stylesheet: groove stays grey; only the chunk color changes."""
        if not bar:
            return
        if not bar.objectName():
            bar.setObjectName(name_hint)
        obj = bar.objectName()
    
        normal_css = f"""
        QProgressBar#{obj} {{
            font-size: 12px; padding: 4px; text-align: center;
            color: black; border: 1px solid #555; border-radius: 4px;
            background: #e0e0e0;                /* grey groove */
        }}
        QProgressBar#{obj}::chunk {{
            background-color: #19A602;           /* GREEN fill in normal runs */
            margin: 0px; border-radius: 3px;
        }}
        """
    
        stopped_css = f"""
        QProgressBar#{obj} {{
            font-size: 12px; padding: 4px; text-align: center;
            color: black; border: 1px solid #555; border-radius: 4px;
            background: #e0e0e0 !important;      /* keep grey groove when stopped */
        }}
        QProgressBar#{obj}::chunk {{
            background-color: #B30000 !important;/* RED fill when stopped */
            margin: 0px; border-radius: 3px;
        }}
        """
    
        bar.setStyleSheet(stopped_css if stopped else normal_css)
        bar.style().unpolish(bar); bar.style().polish(bar); bar.update()
    
    def _force_zero_red(self, bar, enable: bool, full: bool = True):
        """
        When enable=True:
          - keep the text at "0%" (format override)
          - but set the value to full width (100%) so the red CHUNK fills the bar
        When enable=False:
          - restore the original format and value
        """
        if not bar:
            return
    
        if enable:
            if not hasattr(bar, "_orig_format"):
                bar._orig_format = bar.format()
            if not hasattr(bar, "_orig_value"):
                bar._orig_value = bar.value()
    
            bar.setFormat("0%")  # force the text to show 0%
            target = bar.maximum() if full else max(1, bar.maximum() // 100)
            bar.setValue(target)  # fill the bar (chunk draws full-width in red)
            bar._forced_zero_red = True
        else:
            if getattr(bar, "_forced_zero_red", False):
                bar.setFormat(getattr(bar, "_orig_format", "%p%"))
                bar.setValue(getattr(bar, "_orig_value", 0))
                bar._forced_zero_red = False
           
    def _apply_stopped_style_to_all_bars(self, stopped: bool):
        bars = (
            ("cmBar", self.current_manufacturer_progress),
            ("mhBar", self.manufacturer_hyperlink_bar),
            ("ovBar", self.overall_progress_bar),
        )
        for name_hint, bar in bars:
            if not bar:
                continue
            # Always restyle per-bar
            self._style_bar(bar, stopped=stopped, name_hint=name_hint)
            # Always force red 0% fill if stopped
            if stopped:
                self._force_zero_red(bar, enable=True)
            else:
                self._force_zero_red(bar, enable=False)
    
       
    def on_start_stop(self):
        # — START path —
        if not self.is_running:
            # Try to start; this might show a confirmation and bail out.
            self.start_automation()
    
            # ⛔ If user clicked "No" (or start failed), do NOTHING to the bars/styles.
            # Leave the "Manually Stopped" red state intact.
            if not self.is_running:
                return
    
            # ✅ We are actually starting now → safe to restore normal styles
            self.pause_requested = False
            self.pause_button.setText('Pause Automation')
            self.pause_button.setEnabled(True)
    
            # Back to normal look for a fresh run
            self._apply_stopped_style_to_all_bars(False)
            return
    
        # — STOP path —   (unchanged)
        reply = QMessageBox.question(
            self,
            "Confirm Stop",
            "Are you sure you want to end this automation?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
    
        # If we were paused, first resume the subprocess so it can be cleanly terminated
        if self.pause_requested and self.thread is not None and hasattr(self.thread, "process"):
            try:
                proc = psutil.Process(self.thread.process.pid)
                proc.resume()
                for child in proc.children(recursive=True):
                    child.resume()
            except Exception as e:
                self.terminal.append_output(f"⚠️ Couldn’t resume before stopping: {e}")
    
        # clear pause flag and tell loops not to launch any more work
        self.pause_requested = False
        self.stop_requested  = True
    
        # 1) Ask the Python extractor to shut down nicely
        if self.thread is not None and hasattr(self.thread, "process"):
            try:
                if os.name == "nt":
                    self.thread.process.send_signal(signal.CTRL_BREAK_EVENT)
                else:
                    os.killpg(os.getpgid(self.thread.process.pid), signal.SIGTERM)
            except Exception:
                pass
    
            # 2) Fallback: kill Chrome/Chromedriver children and parent
            def kill_children(pid: int):
                try:
                    parent = psutil.Process(pid)
                except psutil.NoSuchProcess:
                    return
                for child in parent.children(recursive=True):
                    try:
                        pname = child.name().lower()
                    except psutil.NoSuchProcess:
                        continue
                    if "chrome" in pname or "chromedriver" in pname:
                        try: child.kill()
                        except psutil.NoSuchProcess: pass
                try: parent.kill()
                except psutil.NoSuchProcess: pass
    
            kill_children(self.thread.process.pid)
    
        # ── reset & disable Pause/Resume button when stopping ──
        self.pause_button.setText('Pause Automation')
        self.pause_button.setEnabled(False)
    
        # 3) Give it a moment, then report & swap button back
        sleep(1)
        self.terminal.append_output("❌ Hyperlink Automation has stopped.")
    
        # Show 'Manually Stopped' + reset and paint bars red
        if hasattr(self, "current_manufacturer_label"):
            self.current_manufacturer_label.setText("Current Manufacturer: Manually Stopped")
        if hasattr(self, "manufacturer_hyperlink_label"):
            self.manufacturer_hyperlink_label.setText("Manufacturer Hyperlinks: Manually Stopped")
        if hasattr(self, "overall_progress_label"):
            self.overall_progress_label.setText("Overall Progress: Manually Stopped")
    
        if hasattr(self, "current_manufacturer_progress"):
            self.current_manufacturer_progress.setValue(0)
        if hasattr(self, "manufacturer_hyperlink_bar"):
            self.manufacturer_hyperlink_bar.setValue(0)
        if hasattr(self, "overall_progress_bar"):
            self.overall_progress_bar.setValue(0)
    
        # turn bars red (and force a repaint at 0%)
        self._apply_stopped_style_to_all_bars(True)
    
        # ── swap back to a fresh “Start Automation” button ──
        layout = self.button_layout
        layout.removeWidget(self.start_button)
        self.start_button.deleteLater()
        self.start_button = CustomButton("Start Automation", "#008000", self)
        self.start_button.clicked.connect(self.on_start_stop)
        layout.addWidget(self.start_button)
    
        self.pause_button.setEnabled(False)
    
        # ── NEW: clear running state so clicks now start again ──
        self.is_running     = False
        self.stop_requested = False
    
        # Ensure it goes above the progress bars
        insert_index = layout.indexOf(self.current_manufacturer_label)
        layout.insertWidget(insert_index, self.start_button)
     
    def on_pause_resume(self):
        # only when running and we have a live subprocess
        if not self.is_running or self.thread is None or not hasattr(self.thread, "process"):
            return
    
        proc = psutil.Process(self.thread.process.pid)
    
        if not self.pause_requested:
            # ── PAUSE ──
            self.pause_requested = True
            self.pause_button.setText('Resume Automation')
            self.terminal.append_output("⏸️ Pausing automation…")
    
            # suspend main process & children
            try:
                proc.suspend()
                for child in proc.children(recursive=True):
                    child.suspend()
            except Exception as e:
                self.terminal.append_output(f"⚠️ Couldn’t pause process: {e}")
    
        else:
            # ── RESUME ──
            self.pause_requested = False
            self.pause_button.setText('Pause Automation')
            self.terminal.append_output("▶️ Resuming automation…")
    
            # resume main process & children
            try:
                proc.resume()
                for child in proc.children(recursive=True):
                    child.resume()
            except Exception as e:
                self.terminal.append_output(f"⚠️ Couldn’t resume process: {e}")
    
            # <-- no call to process_next_manufacturer() here!
            # the suspended extractor will continue its own loop

if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)

        # Require login before launching the main window
        login = LoginDialog(max_attempts=5)
        if login.exec_() != QDialog.Accepted:
            sys.exit(1)  # Exit if login failed or canceled

        window = SeleniumAutomationApp()
        window.show()
        sys.exit(app.exec_())
    except Exception:
        logging.exception("Unhandled exception — crashing out")
        raise
