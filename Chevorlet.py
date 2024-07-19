import sys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from selenium.webdriver.chrome.options import Options
import os
import time
import io
from PIL import Image
import pytesseract
import re
import psutil
import tkinter as tk
from tkinter import messagebox

# Set the path for Tesseract if not in PATH
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Update this path as needed
 
def process_subdocuments(driver, wait, ws, subdocuments, year, model, adas_last_row, parent_xpath):
    for sub_doc_name, sub_doc_info in subdocuments.items():
        if isinstance(sub_doc_info, dict) and 'folder_xpath' in sub_doc_info:
            print(f"Accessing subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder_xpath'])
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder_xpath']))
            )
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments'], year, model, adas_last_row, sub_doc_info['folder_xpath'])
            
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, parent_xpath))
            )
        elif isinstance(sub_doc_info, dict) and 'folder2_xpath' in sub_doc_info:
            print(f"Accessing nested subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder2_xpath'])
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder2_xpath']))
            )
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments2'], year, model, adas_last_row, sub_doc_info['folder2_xpath'])
            driver.back()
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, parent_xpath))
            )
        else:
            print(f"Retrieving sub-document: {sub_doc_name}")
            document_url = navigate_and_extract(driver, wait, sub_doc_info['xpath'])
            update_excel(ws, year, model, sub_doc_name, document_url, adas_last_row, sub_doc_info.get('cell_address'))
            
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, parent_xpath))
            )

def process_documents(driver, wait, ws, model_data, year, model, adas_last_row):
    for doc_name, doc_info in model_data['documents'].items():
        if isinstance(doc_info, dict) and 'folder_xpath' in doc_info:
            print(f"Accessing folder: {doc_name}")
            double_click_element(driver, wait, doc_info['folder_xpath'])
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, doc_info['folder_xpath']))
            )
            process_subdocuments(driver, wait, ws, doc_info['subdocuments'], year, model, adas_last_row, doc_info['folder_xpath'])
            driver.back()
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, model_data['model_page_xpath']))
            )
        else:
            print(f"Retrieving document: {doc_name}")
            document_url = navigate_and_extract(driver, wait, doc_info)
            update_excel(ws, year, model, doc_name, document_url, adas_last_row)

def update_excel(ws, year, model, doc_name, document_url, adas_last_row, cell_address=None):
    if cell_address:
        cell = ws[cell_address]
    else:
        cell = find_row_in_excel(ws, year, "Alfa Romeo", model, doc_name)
        if cell:
            row = cell.row
        else:
            row = ws.max_row + 1
        if doc_name in adas_last_row:
            row = adas_last_row[doc_name] + 1
        else:
            adas_last_row[doc_name] = row
        cell = ws.cell(row=row, column=12)

    cell.hyperlink = document_url
    cell.value = document_url
    cell.font = Font(color="0000FF", underline='single')
    adas_last_row[doc_name] = cell.row
    ws.parent.save(excel_file_path)
    print(f"Hyperlink for {doc_name} added at {cell.coordinate}")

def screenshot_and_get_text(driver):
    screenshot = driver.get_screenshot_as_png()
    image = Image.open(io.BytesIO(screenshot))
    text = pytesseract.image_to_string(image)
    return text

def find_row_in_excel(ws, year, make, model, adas_system):
    for row in ws.iter_rows(min_row=2, max_col=8):
        year_cell, make_cell, model_cell, adas_cell = row[0], row[1], row[2], row[7]
        if (str(year_cell.value).strip() == str(year).strip() and
            str(make_cell.value).strip().lower() == make.lower().strip() and
            str(model_cell.value).strip().lower() == model.lower().strip() and
            adas_system.lower().strip() in str(adas_cell.value).lower().strip()):
            return ws.cell(row=year_cell.row, column=12)
    return None

def double_click_element(driver, wait, xpath):
    element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    ActionChains(driver).double_click(element).perform()
    
def navigate_and_extract(driver, wait, xpath):
    double_click_element(driver, wait, xpath)
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.TAG_NAME, "body"))  # Adjust to a reliable element
    )
    document_url = driver.current_url
    driver.back()
    return document_url

def navigate_to_model(driver, wait, model_xpath):
     model_link = wait.until(EC.element_to_be_clickable((By.XPATH, model_xpath)))
     model_link.click()
     time.sleep(2)  # Wait for the model's page to load

def navigate_to_year(driver, wait, year_xpath):
     year_link = wait.until(EC.element_to_be_clickable((By.XPATH, year_xpath)))
     year_link.click()

def check_if_chrome_running():
    """Check if any Chrome instances are running."""
    for process in psutil.process_iter(['name']):
        if process.info['name'] == 'chrome.exe':
            return True
    return False

def get_chrome_options(use_existing_profile):
    chrome_options = Options()
    if use_existing_profile:
        # Get the user's home directory dynamically
        home_dir = os.path.expanduser("~")
        
        # Construct the path to the Chrome user data directory
        user_data_dir = os.path.join(home_dir, "AppData", "Local", "Google", "Chrome", "User Data")
        
        # Add the user data directory to Chrome options
        chrome_options.add_argument(f"user-data-dir={user_data_dir}")
        
        # Optionally, specify the profile directory (e.g., "Default" for the default profile)
        profile_dir = "Default"  # Change to the specific profile if needed
        chrome_options.add_argument(f"profile-directory={profile_dir}")
    return chrome_options

def ask_user_choice():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    use_existing_profile = messagebox.askyesno(
        "Chrome Instance",
        "Would you like to use the currently installed version of Chrome?\n"
        "Click 'Yes' for the currently installed version or 'No' for a new instance."
    ) 
    root.destroy()  # Destroy the main window
    return use_existing_profile

def run_alfa_romeo_script(excel_path):
    if check_if_chrome_running():
        raise Exception("The program has detected an instance of Google Chrome running on your system. Please ensure that all Chrome instances are closed before proceeding.")
    
    use_existing_profile = ask_user_choice()
    chrome_options = get_chrome_options(use_existing_profile)

    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 10)
    action_chains = ActionChains(driver)
    
        # Your structured data
    years_models_documents = {
    '2012': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVALANCHE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPRICE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'S10': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            }
        }
    },
    '2013': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVALANCHE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPRICE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2014': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPRICE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2015': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPRICE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2016': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPRICE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2017': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {}
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CITY EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'S10': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2018': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'CITY EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'S10': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2019': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CRUZE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'S10': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'VOLT': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2020': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'IMPALA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'S10': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SONIC': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2021': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'ONIX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRACKER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2022': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'AVEO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]'
                }
            },
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'ONIX PLUS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 3500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SPARK': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRACKER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            }
        }
    },
    '2023': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT EUV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BOLT EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'CHEYENNE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'GROOVE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'ONIX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]'
                }
            },
            'S10 MAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 3500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRACKER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            }
        }
    },
    '2024': {
        'year_page_xpath': '//*[@data-automationid="ListCell"][1]',
        'models': {
            'BLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'BLAZER EV': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAMARO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAPTIVA': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CAVALIER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'COLORADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'CORVETTE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EQUINOX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'EXPRESS': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'GROOVE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'MALIBU': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'N400 Move-Max': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'ONIX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'S10 MAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 1500': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 2500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 3500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 4500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 5500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SILVERADO 6500HD': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'SUBURBAN': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TAHOE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TORNADO': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRACKER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAILBLAZER': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAVERSE': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            },
            'TRAX': {
                'model_page_xpath': '//*[@data-automationid="ListCell"][1]',
                'documents': {
                    'ACC': '//*[@data-automationid="ListCell"][1]',
                    'AEB': '//*[@data-automationid="ListCell"][2]',
                    'AHL': '//*[@data-automationid="ListCell"][7]',
                    'APA': '//*[@data-automationid="ListCell"][6]',
                    'BSW/RCTW': '//*[@data-automationid="ListCell"][3]',
                    'BUC': '//*[@data-automationid="ListCell"][4]',
                    'LKA': '//*[@data-automationid="ListCell"][5]',
                    'NV': '//*[@data-automationid="ListCell"][8]',
                    'SVC': '//*[@data-automationid="ListCell"][10]'
                }
            }
        }
    }
}
        

    try:
        # Navigate to the main SharePoint page for Alfa Romeo
        print("Navigating to Alfa Romeo's main SharePoint page...")
        driver.get('https://calibercollision-my.sharepoint.com/:f:/g/personal/mark_klingenhofer_protechdfw_com/EjIo8sg9qXNEt6CDCKpeRGkBj8fLo4MSiJe7w0h7hZ30rQ?e=hFT1RF')
 
        # Give the user up to 120 seconds to log in
        max_wait_time = 120
        start_time = time.time()

        try:
            # Wait until the element with the specified XPath is found, or until 60 seconds have passed
            element = WebDriverWait(driver, max_wait_time).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-automationid="ListCell"][3]'))
            )
        except:
            # If the element is not found within 120 seconds, print a message
            print("The element was not found within 120 seconds.")

        # Calculate the elapsed time
        elapsed_time = time.time() - start_time

        # If less than 60 seconds have passed and the element is found, continue with the rest of the code
        if elapsed_time < max_wait_time:
            # Continue with the code after successful login
            print("Element found, continuing with the code...")
        else:
            # Handle the situation where the element was not found within the allotted time
            print("Proceeding without finding the element...")
        
        # Clicks Alfa Romeo
        print("Locating Alfa Romeo link and clicking...")
        double_click_element(driver, wait, '//*[@data-automationid="ListCell"][3]')
        time.sleep(1)                       
        adas_last_row = {}
        wb = load_workbook(excel_path)
        ws = wb['Model Version']  # Correctly referencing the worksheet
      

        print(f"Workbook loaded successfully: {excel_path}")
    
        for year, data in years_models_documents.items():
            print(f"Processing year: {year}")
            year_page_xpath = data['year_page_xpath']
            double_click_element(driver, wait, year_page_xpath)
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, year_page_xpath))
            )

            for model, model_data in data['models'].items():
                print(f"Accessing model: {model}")
                model_page_xpath = model_data['model_page_xpath']
                double_click_element(driver, wait, model_page_xpath)
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, model_page_xpath))
                )
                adas_last_row = {}  # Reset ADAS last row tracker for each model
                process_documents(driver, wait, ws, model_data, year, model, adas_last_row)
                driver.back()
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, year_page_xpath))
                )
            driver.back()
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@data-automationid="ListCell"][3]'))
            )
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()
        print("Script completed, you may exit now.")
        
if __name__ == "__main__":
    excel_file_path = sys.argv[1]  # The Excel file path is expected as the first argument
    run_alfa_romeo_script(excel_file_path)
