from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from tkinter import messagebox
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.webdriver.chrome.options import Options
import pytesseract
import sys
import time
import psutil
import os
import tkinter as tk

# Set the path for Tesseract if not in PATH
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Update this path as needed

def process_subdocuments(driver, wait, ws, subdocuments, year, model, adas_last_row, parent_xpath):
    for sub_doc_name, sub_doc_info in subdocuments.items():
        if isinstance(sub_doc_info, dict) and 'folder_xpath' in sub_doc_info:
            print(f"Accessing subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments'], year, model, adas_last_row, sub_doc_info['folder_xpath'])
            navigate_back_to_element(driver, wait, parent_xpath)
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)
            time.sleep(2) # Go back one more time to ensure stability
        elif isinstance(sub_doc_info, dict) and 'folder2_xpath' in sub_doc_info:
            print(f"Accessing nested subfolder: {sub_doc_name}")
            double_click_element(driver, wait, sub_doc_info['folder2_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, sub_doc_info['folder2_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, sub_doc_info['subdocuments2'], year, model, adas_last_row, sub_doc_info['folder2_xpath'])
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)  # Go back one more time to ensure stability
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath)
        else:
            print(f"Retrieving sub-document: {sub_doc_name}")
            document_url = navigate_and_extract(driver, wait, sub_doc_info['xpath'])
            time.sleep(2)
            update_excel(ws, year, model, sub_doc_name, document_url, adas_last_row, sub_doc_info.get('cell_address'))
            time.sleep(2)
            navigate_back_to_element(driver, wait, parent_xpath) # Go back one more time to ensure stability

def process_documents(driver, wait, ws, model_data, year, model, adas_last_row):
    for doc_name, doc_info in model_data['documents'].items():
        if isinstance(doc_info, dict) and 'folder_xpath' in doc_info:
            print(f"Accessing folder: {doc_name}")
            double_click_element(driver, wait, doc_info['folder_xpath'])
            wait.until(EC.visibility_of_element_located((By.XPATH, doc_info['folder_xpath'])))
            time.sleep(2)
            process_subdocuments(driver, wait, ws, doc_info['subdocuments'], year, model, adas_last_row, doc_info['folder_xpath'])
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])
            time.sleep(2)
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])  # Go back one more time to ensure stability
            time.sleep(2)
        else:
            print(f"Retrieving document: {doc_name}")
            document_url = navigate_and_extract(driver, wait, doc_info)
            time.sleep(2)
            update_excel(ws, year, model, doc_name, document_url, adas_last_row)
            time.sleep(2)
            navigate_back_to_element(driver, wait, model_data['model_page_xpath'])

def navigate_and_extract(driver, wait, xpath):
    double_click_element(driver, wait, xpath)
    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.TAG_NAME, "body"))  # Adjust to a reliable element
    )
    time.sleep(1)
    document_url = driver.current_url
    navigate_back_to_element(driver, wait, xpath)
    return document_url

def navigate_back_to_element(driver, wait, xpath):    # javascript language
    driver.execute_script("window.history.go(-1)")
    time.sleep(3)  # Ensure the page fully loads
    # Improved wait mechanism
    for _ in range(10):  # Retry up to 10 times
        try:
            wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
            break
        except TimeoutException:
            time.sleep(1)  # Wait a bit longer if necessary

def update_excel(ws, year, model, doc_name, document_url, adas_last_row, cell_address=None):
    if cell_address:
        cell = ws[cell_address]
    else:
        cell = find_row_in_excel(ws, year, "Alfa Romeo", model, doc_name)
        if cell:
            row = cell.row
        else:
            row = ws.max_row + 1
        if doc_name in adas_last_row:
            row = adas_last_row[doc_name] + 1
        else:
            adas_last_row[doc_name] = row
        cell = ws.cell(row=row, column=12)

    cell.hyperlink = document_url
    cell.value = document_url
    cell.font = Font(color="0000FF", underline='single')
    adas_last_row[doc_name] = cell.row
    ws.parent.save(excel_file_path)
    print(f"Hyperlink for {doc_name} added at {cell.coordinate}")

def find_row_in_excel(ws, year, make, model, adas_system):
    for row in ws.iter_rows(min_row=2, max_col=8):
        year_cell, make_cell, model_cell, adas_cell = row[0], row[1], row[2], row[7]
        if (str(year_cell.value).strip() == str(year).strip() and
            str(make_cell.value).strip().lower() == make.lower().strip() and
            str(model_cell.value).strip().lower() == model.lower().strip() and
            adas_system.lower().strip() in str(adas_cell.value).lower().strip()):
            return ws.cell(row=year_cell.row, column=12)
    return None

def double_click_element(driver, wait, xpath):
    attempts = 5
    while attempts > 0:
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            ActionChains(driver).double_click(element).perform()
            return  # If successful, exit the function
        except (TimeoutException, StaleElementReferenceException) as e:
            print(f"Attempt to double-click element failed: {e}. Retrying... {attempts-1} attempts left.")
            time.sleep(1)  # Wait a bit before retrying
            attempts -= 1
    raise Exception(f"Failed to double-click element after {5} attempts.")

def check_if_chrome_running():
    """Check if any Chrome instances are running."""
    for process in psutil.process_iter(['name']):
        if process.info['name'] == 'chrome.exe':
            return True
    return False

def get_chrome_options(use_existing_profile):
    chrome_options = Options()
    if use_existing_profile:
        # Get the user's home directory dynamically
        home_dir = os.path.expanduser("~")
        
        # Construct the path to the Chrome user data directory
        user_data_dir = os.path.join(home_dir, "AppData", "Local", "Google", "Chrome", "User Data")
        
        # Add the user data directory to Chrome options
        chrome_options.add_argument(f"user-data-dir={user_data_dir}")
        
        # Optionally, specify the profile directory (e.g., "Default" for the default profile)
        profile_dir = "Default"  # Change to the specific profile if needed
        chrome_options.add_argument(f"profile-directory={profile_dir}")
    return chrome_options

def ask_user_choice():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    use_existing_profile = messagebox.askyesno(
        "Chrome Instance",
        "Would you like to use the currently installed version of Chrome?\n"
        "Click 'Yes' for the currently installed version or 'No' for a new instance."
    ) 
    root.destroy()  # Destroy the main window
    return use_existing_profile

def run_alfa_romeo_script(excel_path):
    if check_if_chrome_running():
        raise Exception("The program has detected an instance of Google Chrome running on your system. Please ensure that all Chrome instances are closed before proceeding.")
    
    use_existing_profile = ask_user_choice()
    chrome_options = get_chrome_options(use_existing_profile)
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-extensions")  # Disable extensions to avoid conflicts
    chrome_options.add_argument("--disable-infobars")  # Disable infobars
    chrome_options.add_argument("--disable-browser-side-navigation")  # Disable side navigation issues
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # Avoid detection as bot


    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 10)
    
    years_models_documents = {
        
        '2015': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]'
                    }
                }
            }
        },
        '2016': {
            'year_page_xpath': '//*[@data-grid-row="1"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]'
                    }
                }
            }
        },
        '2017': {
            'year_page_xpath': '//*[@data-grid-row="2"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="1"]/div[3]'
                    }
                },
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="1"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="1"]/div[3]',
                        'AHL': '//*[@data-grid-row="6"]/div[3]',
                        'APA': '//*[@data-grid-row="2"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="3"]/div[3]',
                        'BUC': '//*[@data-grid-row="7"]/div[3]',
                        'LKA': '//*[@data-grid-row="4"]/div[3]',
                    }
                }
            }
        },
        '2018': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',                       ########################################### Work on ARRAY FROM HERE ########################################################################################
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2019': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2020': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2021': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                '4C COUPE': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2022': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2023': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        },
        '2024': {
            'year_page_xpath': '//*[@data-grid-row="0"]/div[3]',
            'models': {
                'GIULIA': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                },
                'STELVIO': {
                    'model_page_xpath': '//*[@data-grid-row="0"]/div[3]',
                    'documents': {
                        'ACC': '//*[@data-grid-row="0"]/div[3]',
                        'AEB': '//*[@data-grid-row="0"]/div[3]',
                        'AHL': '//*[@data-grid-row="0"]/div[3]',
                        'APA': '//*[@data-grid-row="0"]/div[3]',
                        'BSW/RCTW': '//*[@data-grid-row="0"]/div[3]',
                        'BUC': '//*[@data-grid-row="0"]/div[3]',
                        'LKA': '//*[@data-grid-row="0"]/div[3]'
                    }
                }
            }
        }
    }

    try:
        # Navigate to the main SharePoint page for Alfa Romeo
        print("\__ Navigating to Alfa Romeo's main SharePoint page...")
        driver.get('https://calibercollision-my.sharepoint.com/:f:/g/personal/mark_klingenhofer_protechdfw_com/EtJaiKE_WmJInQhx-h6X_KkBB73aha-7sbJJgbv3GWpxIQ?e=evDXxT')
 
        # Give the user up to 120 seconds to log in
        max_wait_time = 120
        start_time = time.time()

        try:
            # Wait until the element with the specified XPath is found, or until 60 seconds have passed
            WebDriverWait(driver, max_wait_time).until(
                EC.presence_of_element_located((By.XPATH, '//*[@data-grid-row="2"]/div[3]'))
            )
        except:
            # If the element is not found within 120 seconds, print a message
            print("\__ The element was not found within 120 seconds.")

        # Calculate the elapsed time
        elapsed_time = time.time() - start_time

        # If less than 60 seconds have passed and the element is found, continue with the rest of the code
        if elapsed_time < max_wait_time:
            # Continue with the code after successful login
            print("\__ Element found, continuing with the code...")
        else:
            # Handle the situation where the element was not found within the allotted time
            print("\__ Proceeding without finding the element...")
        
        # Loads into Alfa Romeo
        print("\__ Loaded into Alfa Romeo Sharepoint Page...")
        time.sleep(1)                       
        adas_last_row = {}
        
        # Ensure the Excel file is valid and can be opened
        try:
            wb = load_workbook(excel_path)
            ws = wb['Model Version']  # Correctly referencing the worksheet
            print(f"\__ Workbook loaded successfully: {excel_path}")
        except Exception as e:
            print(f"\__ Failed to open the Excel file: {e}")
            return

        for year, data in years_models_documents.items():
            print(f"Processing year: {year}")
            year_page_xpath = data['year_page_xpath']
            double_click_element(driver, wait, year_page_xpath)
            time.sleep(2)
            wait.until(EC.visibility_of_element_located((By.XPATH, year_page_xpath)))

            for model, model_data in data['models'].items():
                print(f"Accessing model: {model}")
                model_page_xpath = model_data['model_page_xpath']
                double_click_element(driver, wait, model_page_xpath)
                time.sleep(2)
                wait.until(EC.visibility_of_element_located((By.XPATH, model_page_xpath)))
                adas_last_row = {}  # Reset ADAS last row tracker for each model
                process_documents(driver, wait, ws, model_data, year, model, adas_last_row)
                navigate_back_to_element(driver, wait, year_page_xpath)  # Go back one more time to ensure stability
            navigate_back_to_element(driver, wait, '//*[@data-grid-row="2"]/div[3]')
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()
        print("Script completed, you may exit now.")

if __name__ == "__main__":
    excel_file_path = sys.argv[1]
    run_alfa_romeo_script(excel_file_path)
